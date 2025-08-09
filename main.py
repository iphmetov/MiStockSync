"""
MiStockSync - Приложение для синхронизации прайсов
Версия: 0.9.5
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import sys
import os
import pandas as pd
from datetime import datetime
import logging
import json
import shutil
import warnings

# Отключаем предупреждение PIL о больших изображениях
warnings.filterwarnings("ignore", category=UserWarning, module="PIL")

# Опциональные импорты для точечного обновления Excel
try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Добавляем путь к модулю excel_loader
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "excel_loader"))

try:
    from excel_loader.loader import (
        select_and_load_excel,
        get_available_configs,
        load_largest_file,
    )
except ImportError as e:
    print(f"Ошибка импорта excel_loader: {e}")
    sys.exit(1)


# КОНСТАНТЫ ДЛЯ ФИЛЬТРАЦИИ ДАННЫХ
# ================================

# Для фильтрации баланса Вити - список допустимых статусов
VITYA_BALANCE_AVAILABLE = ["Имеются в нал.", "Распродажа"]

# Для фильтрации баланса Димы
DIMI_BALANCE_EXPECTED = "Ожидается"

# Минимальная цена для фильтрации (исключаем 0 и NaN)
MIN_PRICE_THRESHOLD = 0.01

# Константы для обновления цен (из notebook)
MIN_PRICE_CHANGE_PERCENT = 0.1  # Минимальное изменение для обновления
MAX_PRICE_CHANGE_PERCENT = 100.0  # Максимальное разрешенное изменение
SIGNIFICANT_CHANGE_PERCENT = 20.0  # Порог "значительного" изменения

# Порог схожести для нечеткого поиска (0.3 = 30%)
TRSH = 0.33


class MiStockSyncApp:
    def __init__(self, root):
        self.root = root
        # Заголовок устанавливается в main()

        # Настройка логирования
        self.setup_logging()

        # Загружаем настройки из файла
        self.settings = self.load_settings()

        # Данные
        self.current_df = None
        self.current_config = None
        self.base_df = None
        self.comparison_result = None  # Результаты сравнения
        self.price_updated = False  # Флаг обновления цен
        self.articles_added = False  # Флаг добавления товаров в базу

        # Если root не None, инициализируем GUI
        if self.root is not None:
            # Загружаем размеры основного окна
            main_width = self.settings.get("main_window_width", 1000)
            main_height = self.settings.get("main_window_height", 800)
            self.root.geometry(f"{main_width}x{main_height}")
            self.logger.info(
                f"📐 Размеры окна загружены из настроек: {main_width}x{main_height}"
            )

            self.auto_load_base = tk.BooleanVar(value=True)  # Чекбокс автозагрузки базы

            # Настройки интерфейса (применяем загруженные настройки)
            self.current_font_size = self.settings.get("font_size", "normal")
            self.auto_load_base_enabled = self.settings.get("auto_load_base", True)

            # Создаем интерфейс
            self.create_widgets()

            # Загружаем доступные конфиги
            self.load_available_configs()

            # Автооткрытие диалога выбора файла при запуске приложения
            self.root.after(100, self.auto_open_file_dialog)

            # Добавляем обработчик изменения размера окна
            self.root.bind("<Configure>", self.on_window_resize)
        else:
            # Режим без GUI - только для тестирования
            self.logger.info("🔧 Режим без GUI (для тестирования)")
            self.auto_load_base = None
            self.current_font_size = "normal"
            self.auto_load_base_enabled = True

    def auto_open_file_dialog(self):
        """Автоматически открывает диалог выбора файла при запуске приложения"""
        self.log_info("🚀 Автооткрытие диалога выбора файла...")
        self.select_file()

    def on_window_resize(self, event):
        """Обработчик изменения размера окна"""
        # Проверяем, что это главное окно (не дочерние окна)
        if event.widget == self.root:
            # Получаем новые размеры
            new_width = event.width
            new_height = event.height

            # Проверяем, что размеры действительно изменились значительно
            current_width = self.settings.get("main_window_width", 0)
            current_height = self.settings.get("main_window_height", 0)

            # Минимальное изменение для сохранения (5 пикселей)
            min_change = 5

            if (
                new_width > 100
                and new_height > 100
                and (
                    abs(new_width - current_width) >= min_change
                    or abs(new_height - current_height) >= min_change
                )
            ):
                # Отменяем предыдущий таймер
                if hasattr(self, "_resize_timer"):
                    self.root.after_cancel(self._resize_timer)

                # Устанавливаем новый таймер с передачей размеров (увеличиваем задержку до 2 секунд)
                self._resize_timer = self.root.after(
                    2000, lambda: self._save_window_size(new_width, new_height)
                )

    def _save_window_size(self, width, height):
        """Сохранение размеров окна с задержкой"""
        try:
            # Проверяем, что размеры все еще актуальны
            current_width = self.settings.get("main_window_width", 0)
            current_height = self.settings.get("main_window_height", 0)

            # Обновляем настройки только если они действительно изменились
            if width != current_width or height != current_height:
                self.settings["main_window_width"] = width
                self.settings["main_window_height"] = height

                # Сохраняем настройки
                self.save_settings(self.settings)
                self.log_info(
                    f"💾 Размеры окна автоматически сохранены: {width}x{height}"
                )
        except Exception as e:
            self.log_error(f"❌ Ошибка автоматического сохранения размеров окна: {e}")

    def setup_logging(self):
        """Настройка системы логирования"""
        # Создаем логгер для приложения
        self.logger = logging.getLogger("MiStockSync")
        self.logger.setLevel(logging.INFO)

        # Удаляем существующие обработчики
        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)

        # Консольный обработчик
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        # Формат сообщений
        formatter = logging.Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
        )
        console_handler.setFormatter(formatter)

        # Добавляем обработчик
        self.logger.addHandler(console_handler)

        # Настройка файлового логирования
        logs_dir = "logs"
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir)

        self.log_file_path = os.path.join(
            logs_dir, f"mistocksync_{datetime.now().strftime('%Y%m%d')}.log"
        )
        file_handler = logging.FileHandler(self.log_file_path, encoding="utf-8")
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

        self.logger.info("🚀 MiStockSync запущен")
        self.logger.info("📋 Система логирования настроена")

    def load_settings(self):
        """Загрузка настроек из файла settings.json"""
        settings_file = "settings.json"
        default_settings = {
            "auto_load_base": True,
            "font_size": "normal",
            "main_window_width": 1000,
            "main_window_height": 800,
            "add_articles_dialog_width": 800,
            "add_articles_dialog_height": 533,
            "confirm_exit": True,  # Показывать окно подтверждения выхода при изменениях
        }

        try:
            if os.path.exists(settings_file):
                with open(settings_file, "r", encoding="utf-8") as f:
                    settings = json.load(f)

                # Проверяем наличие всех нужных ключей
                for key, default_value in default_settings.items():
                    if key not in settings:
                        settings[key] = default_value
                        self.logger.info(
                            f"⚙️ Добавлен недостающий ключ настроек: {key} = {default_value}"
                        )

                self.logger.info(f"⚙️ Настройки загружены из {settings_file}")
                return settings
            else:
                self.logger.info(
                    "⚙️ Файл настроек не найден, используются значения по умолчанию"
                )
                self.logger.info(
                    f"📐 Размеры окна по умолчанию: {default_settings['main_window_width']}x{default_settings['main_window_height']}"
                )
                return default_settings

        except Exception as e:
            self.logger.error(f"❌ Ошибка загрузки настроек: {e}")
            return default_settings

    def save_settings(self, settings):
        """Сохранение настроек в файл settings.json"""
        settings_file = "settings.json"

        try:
            with open(settings_file, "w", encoding="utf-8") as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)

            self.logger.info(f"💾 Настройки сохранены в {settings_file}")
            return True

        except Exception as e:
            self.logger.error(f"❌ Ошибка сохранения настроек: {e}")
            return False

    def create_widgets(self):
        """Создание элементов интерфейса"""

        # Создаем главное меню
        self.create_menu()

        # Главный фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Настройка растяжения
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)

        # Заголовок
        title_label = ttk.Label(
            main_frame,
            text="MiStockSync - Синхронизация прайсов",
            font=("Arial", 16, "bold"),
        )
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 10))

        # Мини-панель инструментов
        toolbar_frame = ttk.Frame(main_frame)
        toolbar_frame.grid(
            row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10)
        )

        # Контейнер для инструментов (прижатый к левому краю)
        tools_container = ttk.Frame(toolbar_frame)
        tools_container.grid(row=0, column=0, sticky=tk.W)

        # Оставляем toolbar для будущих быстрых действий (пока пустой)
        # TODO: Добавить кнопки быстрого доступа к основным функциям

        # Выбор конфигурации
        config_frame = ttk.LabelFrame(main_frame, text="Выбор поставщика", padding="10")
        config_frame.grid(
            row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10)
        )
        config_frame.columnconfigure(1, weight=1)

        ttk.Label(config_frame, text="Конфигурация:").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 10)
        )

        self.config_var = tk.StringVar()
        self.config_combo = ttk.Combobox(
            config_frame, textvariable=self.config_var, state="readonly"
        )
        self.config_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))

        # Кнопки загрузки
        buttons_frame = ttk.Frame(config_frame)
        buttons_frame.grid(row=0, column=2, sticky=tk.E)

        ttk.Button(
            buttons_frame, text="📁 Выбрать файл", command=self.select_file
        ).grid(row=0, column=0, padx=(0, 5))

        # Область вывода информации
        info_frame = ttk.LabelFrame(main_frame, text="Информация о файле", padding="10")
        info_frame.grid(
            row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10)
        )
        info_frame.columnconfigure(0, weight=1)
        info_frame.rowconfigure(0, weight=1)

        self.info_text = scrolledtext.ScrolledText(info_frame, width=80, height=15)
        self.info_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Применяем загруженный размер шрифта
        self.apply_font_size(self.current_font_size)

        # Кнопки действий
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(
            row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0)
        )

        self.show_data_button = ttk.Button(
            action_frame,
            text="📋 Показать данные",
            command=self.show_data_sample,
            state="disabled",
        )
        self.show_data_button.grid(row=0, column=0, sticky=tk.W)

        self.save_data_button = ttk.Button(
            action_frame,
            text="💾 Сохранить обработанный",
            command=self.save_data,
            state="disabled",
        )
        self.save_data_button.grid(row=0, column=1, padx=(10, 0))

        self.compare_button = ttk.Button(
            action_frame,
            text="🔍 Сравнить с базой",
            command=self.compare_with_base,
            state="disabled",
        )
        self.compare_button.grid(row=0, column=2, padx=(10, 0))

        # Новые кнопки после сравнения
        self.update_prices_button = ttk.Button(
            action_frame,
            text="🏷️ Обновить цены",
            command=self.update_prices,
            state="disabled",
        )
        self.update_prices_button.grid(row=0, column=3, padx=(10, 0))

        self.report_button = ttk.Button(
            action_frame,
            text="📊 Сохранить отчет",
            command=self.save_report,
            state="disabled",
        )
        self.report_button.grid(row=0, column=4, padx=(10, 0))

        self.add_to_base_button = ttk.Button(
            action_frame,
            text="📥 Добавить новый товар в базу",
            command=self.add_to_base,
            state="disabled",
        )
        self.add_to_base_button.grid(row=0, column=5, padx=(10, 0))

        self.show_log_button = ttk.Button(
            action_frame,
            text="📋 Показать лог",
            command=self.show_log_window,
            state="normal",  # Всегда активна
        )
        self.show_log_button.grid(row=0, column=6, padx=(10, 0))

        # Продвинутый статус-бар
        self.create_advanced_status_bar(main_frame)

    def create_menu(self):
        """Создание главного меню приложения"""

        # Создаем главное меню
        self.menubar = tk.Menu(self.root)

        # === МЕНЮ "ФАЙЛ" ===
        file_menu = tk.Menu(self.menubar, tearoff=0)
        file_menu.add_command(
            label="📁 Открыть файл", command=self.select_file, accelerator="Ctrl+O"
        )
        file_menu.add_separator()
        file_menu.add_command(
            label="⚙️ Настройки", command=self.show_settings, accelerator="Ctrl+,"
        )
        file_menu.add_separator()
        file_menu.add_command(
            label="🚪 Выход", command=self.quit_application, accelerator="Ctrl+Q"
        )
        self.menubar.add_cascade(label="📁 Файл", menu=file_menu)

        # === МЕНЮ "ПРАВКА" ===
        edit_menu = tk.Menu(self.menubar, tearoff=0)
        edit_menu.add_command(
            label="✂️ Вырезать", command=self.cut_text, accelerator="Ctrl+X"
        )
        edit_menu.add_command(
            label="📋 Копировать", command=self.copy_text, accelerator="Ctrl+C"
        )
        edit_menu.add_separator()
        edit_menu.add_command(
            label="🔘 Выделить все", command=self.select_all_text, accelerator="Ctrl+A"
        )
        edit_menu.add_command(
            label="🔄 Инвертировать выделенное",
            command=self.invert_selection,
            accelerator="Ctrl+I",
        )
        self.menubar.add_cascade(label="✏️ Правка", menu=edit_menu)

        # === МЕНЮ "ВИД" ===
        view_menu = tk.Menu(self.menubar, tearoff=0)
        view_menu.add_command(
            label="🧹 Очистить", command=self.clear_info, accelerator="Ctrl+L"
        )
        view_menu.add_command(
            label="🔄 Обновить", command=self.refresh_interface, accelerator="F5"
        )
        view_menu.add_separator()

        # Подменю размеров шрифта
        font_menu = tk.Menu(view_menu, tearoff=0)
        font_menu.add_command(
            label="📝 Обычный шрифт", command=lambda: self.change_font_size("normal")
        )
        font_menu.add_command(
            label="📄 Средний шрифт", command=lambda: self.change_font_size("medium")
        )
        font_menu.add_command(
            label="📊 Крупный шрифт", command=lambda: self.change_font_size("large")
        )
        view_menu.add_cascade(label="🔤 Размер шрифта", menu=font_menu)

        self.menubar.add_cascade(label="👁️ Вид", menu=view_menu)

        # === МЕНЮ "СПРАВКА" ===
        help_menu = tk.Menu(self.menubar, tearoff=0)
        help_menu.add_command(
            label="📖 Помощь", command=self.show_help, accelerator="F1"
        )
        help_menu.add_separator()
        help_menu.add_command(
            label="ℹ️ О программе", command=self.show_about, accelerator="Ctrl+F1"
        )
        self.menubar.add_cascade(label="❓ Справка", menu=help_menu)

        # Привязываем меню к окну
        self.root.config(menu=self.menubar)

        # Горячие клавиши
        self.setup_hotkeys()

    def setup_hotkeys(self):
        """Настройка горячих клавиш"""
        # Файл
        self.root.bind("<Control-o>", lambda e: self.select_file())
        self.root.bind("<Control-comma>", lambda e: self.show_settings())
        self.root.bind("<Control-q>", lambda e: self.quit_application())

        # Правка
        self.root.bind("<Control-x>", lambda e: self.cut_text())
        self.root.bind("<Control-c>", lambda e: self.copy_text())
        self.root.bind("<Control-a>", lambda e: self.select_all_text())
        self.root.bind("<Control-i>", lambda e: self.invert_selection())

        # Вид
        self.root.bind("<Control-l>", lambda e: self.clear_info())
        self.root.bind("<F5>", lambda e: self.refresh_interface())

        # Справка
        self.root.bind("<F1>", lambda e: self.show_help())
        self.root.bind("<Control-F1>", lambda e: self.show_about())

    def load_available_configs(self):
        """Загрузка списка доступных конфигураций"""
        self.log_info("📋 Загрузка доступных конфигураций...")
        try:
            configs = get_available_configs()

            # Проверяем, что мы в режиме GUI
            if hasattr(self, "config_combo"):
                self.config_combo["values"] = configs

                # НОВОЕ: Устанавливаем "auto" по умолчанию
                if "auto" in configs:
                    self.config_combo.set("auto")
                elif configs:
                    self.config_combo.set(configs[0])

            self.log_info(f"✅ Найдено конфигураций: {len(configs)}")
            self.log_info(f"📋 Доступные конфиги: {', '.join(configs)}")
        except Exception as e:
            self.log_error(f"Ошибка загрузки конфигураций: {e}")

    def select_file(self):
        """Выбор и загрузка файла"""
        self.log_info("📁 Выбор файла для загрузки...")

        # Создаем папку data/input если её нет
        input_dir = "data/input"
        if not os.path.exists(input_dir):
            os.makedirs(input_dir)
            self.log_info(f"📁 Создана папка: {input_dir}")

        # Сначала показываем диалог выбора файла
        from tkinter import filedialog

        file_path = filedialog.askopenfilename(
            title="Выберите Excel файл",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            initialdir=input_dir,
        )

        if not file_path:
            self.log_info("ℹ️ Выбор файла отменен")
            return

        self.log_info(f"📁 Выбран файл: {os.path.basename(file_path)}")

        # НОВОЕ: Автоматически определяем конфиг
        if self.config_var.get() == "auto":
            detected_config = self.auto_select_config(file_path)
            config_name = detected_config
        else:
            config_name = self.config_var.get()

        if not config_name:
            messagebox.showwarning(
                "Предупреждение", "Не удалось определить конфигурацию"
            )
            return

        try:
            # Запускаем красивый прогресс-бар для загрузки
            self.start_progress("Загрузка файла", 5, "file")

            # Шаг 1: Подготовка
            self.update_progress(1, "Подготовка к загрузке")
            from excel_loader.loader import load_with_config

            # Шаг 2: Загрузка Excel файла
            self.update_progress(2, "Чтение Excel файла")
            df = load_with_config(file_path, config_name)

            if df is not None:
                # Шаг 3: Обработка данных
                self.update_progress(3, "Обработка данных")
                self.current_df = df
                self.current_config = config_name
                self.current_file_name = os.path.basename(
                    file_path
                )  # Сохраняем имя файла

                # Сбрасываем результаты сравнения при загрузке нового файла
                self.comparison_result = None

                # Сбрасываем состояние обновления цен при загрузке нового прайса поставщика
                if hasattr(self, "price_updated"):
                    self.price_updated = False
                    self.log_info("🔄 Состояние обновления цен сброшено")

                # Сбрасываем состояние добавления товаров при загрузке нового файла
                self.articles_added = False
                self.log_info("🔄 Состояние добавления товаров сброшено")

                # Шаг 4: Отображение информации
                self.update_progress(4, "Подготовка отображения")
                self.show_file_info(df, config_name)

                # Шаг 5: Финализация
                self.update_progress(5, "Завершение загрузки")

                # Обновляем состояние кнопок
                self.update_buttons_state()

                # Обновляем информацию о файлах в статус-баре
                self.update_files_info()

                # Завершаем с красивым сообщением
                rows = len(df)
                cols = len(df.columns)
                size_mb = df.memory_usage(deep=True).sum() / 1024 / 1024
                self.finish_progress(
                    f"✅ Загружено: {rows:,} строк, {cols} столбцов ({size_mb:.1f} МБ)"
                )

                # Сбрасываем конфигурацию на "auto" для следующей загрузки
                if "auto" in self.config_combo["values"]:
                    self.config_combo.set("auto")
                    self.log_info(
                        "🔄 Конфигурация сброшена на 'auto' для следующей загрузки"
                    )
            else:
                self.finish_progress("Файл не был загружен", auto_reset=False)
                self.set_status("Файл не загружен", "error")

        except Exception as e:
            self.log_error(f"Ошибка загрузки файла: {e}")
            self.finish_progress("Ошибка загрузки файла", auto_reset=False)
            self.set_status(f"Ошибка: {str(e)}", "error")

    def load_largest(self):
        """Загрузка самого большого файла"""

        # Директория с данными
        data_dir = "data/input"

        try:
            self.set_status("Поиск самого большого файла...", "loading")
            self.root.update()

            # Находим самый большой файл
            excel_files = []
            for file in os.listdir(data_dir):
                if file.endswith((".xlsx", ".xls")):
                    file_path = os.path.join(data_dir, file)
                    file_size = os.path.getsize(file_path)
                    excel_files.append((file_path, file_size))

            if not excel_files:
                self.log_error("Excel файлы не найдены в data/input")
                self.set_status("Файлы не найдены", "warning")
                return

            # Сортируем по размеру и берем самый большой
            excel_files.sort(key=lambda x: x[1], reverse=True)
            largest_file_path, largest_size = excel_files[0]

            self.log_info(
                f"Найден самый большой файл: {os.path.basename(largest_file_path)} ({largest_size} bytes)"
            )

            # Автоматически определяем конфиг
            config_name = self.auto_select_config(largest_file_path)

            from excel_loader.loader import load_with_config

            df = load_with_config(largest_file_path, config_name)

            if df is not None:
                self.current_df = df
                self.current_config = config_name

                # Сбрасываем результаты сравнения при загрузке нового файла
                self.comparison_result = None

                # Сбрасываем состояние обновления цен при загрузке нового прайса поставщика
                if hasattr(self, "price_updated"):
                    self.price_updated = False
                    self.log_info("🔄 Состояние обновления цен сброшено")

                # Сбрасываем состояние добавления товаров при загрузке нового файла
                self.articles_added = False
                self.log_info("🔄 Состояние добавления товаров сброшено")

                self.show_file_info(df, config_name)
                self.set_status("Самый большой файл загружен", "success")

                # Обновляем состояние кнопок
                self.update_buttons_state()
            else:
                self.set_status("Файл не загружен", "error")

        except Exception as e:
            self.log_error(f"Ошибка загрузки самого большого файла: {e}")
            self.set_status("Ошибка загрузки", "error")

    def show_file_info(self, df, config_name):
        """Показ информации о загруженном файле"""
        self.log_info(f"📊 Отображение информации о файле (конфиг: {config_name})")
        # Очищаем только текстовое поле, НЕ сбрасывая данные
        self.info_text.delete(1.0, tk.END)

        # Основная информация
        info = f"📊 ИНФОРМАЦИЯ О ЗАГРУЖЕННЫХ ФАЙЛАХ\n"
        info += f"{'='*50}\n"

        # Информация о загруженном прайсе поставщика
        info += f"💼 ПРАЙС ПОСТАВЩИКА:\n"
        info += f"   Конфигурация: {config_name}\n"
        if hasattr(self, "current_file_name") and self.current_file_name:
            info += f"   Файл: {self.current_file_name}\n"
        info += f"   Дата загрузки: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        info += f"   Строк: {len(df):,}\n"
        info += f"   Столбцов: {len(df.columns):,}\n"
        info += (
            f"   Размер: {df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB\n\n"
        )

        # Информация о загруженной базе
        info += f"🏢 БАЗА ДАННЫХ:\n"
        if self.base_df is not None:
            info += f"   Статус: ✅ ЗАГРУЖЕНА\n"
            if hasattr(self, "base_file_name") and self.base_file_name:
                info += f"   Файл: {self.base_file_name}\n"
            info += f"   Строк: {len(self.base_df):,}\n"
            info += f"   Столбцов: {len(self.base_df.columns):,}\n"
            info += f"   Размер: {self.base_df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB\n\n"
        else:
            info += f"   Статус: ❌ НЕ ЗАГРУЖЕНА\n\n"

        # Информация о столбцах
        info += f"📋 СТОЛБЦЫ ({len(df.columns)}):\n"
        info += f"{'-'*30}\n"
        for i, col in enumerate(df.columns, 1):
            info += f"{i:2d}. {col}\n"

        # Типы данных
        info += f"\n📊 ТИПЫ ДАННЫХ:\n"
        info += f"{'-'*30}\n"
        for col in df.columns:
            non_null = df[col].notna().sum()
            info += f"{col}: {str(df[col].dtype)} ({non_null:,} не пустых)\n"

        # Статистика по пустым значениям
        info += f"\n❌ ПУСТЫЕ ЗНАЧЕНИЯ:\n"
        info += f"{'-'*30}\n"
        null_counts = df.isnull().sum()
        for col in df.columns:
            if null_counts[col] > 0:
                info += f"{col}: {null_counts[col]:,} пустых\n"

        if null_counts.sum() == 0:
            info += "Пустых значений нет! ✅\n"

        self.info_text.insert(tk.END, info)
        self.log_info(
            f"✅ Файл загружен: {len(df)} строк, {len(df.columns)} столбцов, {df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB"
        )

    def show_data_sample(self):
        """Показ образца данных"""
        if self.current_df is None:
            messagebox.showwarning("Предупреждение", "Сначала загрузите файл")
            return

        # Очищаем только текстовое поле, НЕ сбрасывая данные
        self.info_text.delete(1.0, tk.END)

        df = self.current_df

        info = f"📋 ОБРАЗЕЦ ДАННЫХ (первые 10 строк)\n"
        info += f"{'='*80}\n\n"

        # Показываем первые 10 строк
        sample_df = df.head(10)
        info += sample_df.to_string(max_cols=10, max_colwidth=20) + "\n\n"

        # Уникальные значения для важных столбцов
        important_cols = [
            "article",
            "name",
            "price",
            "article_vitya",
            "price_usd",
            "price_rub",
        ]
        existing_cols = [col for col in important_cols if col in df.columns]

        if existing_cols:
            info += f"📊 УНИКАЛЬНЫЕ ЗНАЧЕНИЯ (топ-10):\n"
            info += f"{'-'*50}\n"
            for col in existing_cols[:3]:  # Показываем только первые 3
                unique_vals = df[col].value_counts().head(10)
                info += f"\n{col.upper()}:\n"
                for val, count in unique_vals.items():
                    info += f"  {val}: {count:,}\n"

        self.info_text.insert(tk.END, info)

    def save_data(self):
        """Сохранение обработанных данных"""
        self.log_info("💾 Начало сохранения обработанных данных...")

        if self.current_df is None:
            self.log_error("Файл не загружен")
            messagebox.showwarning("Предупреждение", "Сначала загрузите файл")
            return

        from tkinter import filedialog

        # Создаем папку data/output если её нет
        output_dir = "data/output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.log_info(f"📁 Создана папка: {output_dir}")

        self.log_info(f"📁 Открываем диалог сохранения в папке: {output_dir}")

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
            initialdir=output_dir,
        )

        if file_path:
            try:
                self.set_status("Предобработка данных...", "save")

                # Предобрабатываем данные перед сохранением
                processed_df = self.preprocess_supplier_data(
                    self.current_df, self.current_config
                )

                if file_path.endswith(".xlsx"):
                    processed_df.to_excel(file_path, index=False)
                elif file_path.endswith(".csv"):
                    processed_df.to_csv(file_path, index=False, encoding="utf-8")

                self.log_info(f"Обработанные данные сохранены: {file_path}")
                self.log_info(
                    f"Исходно: {len(self.current_df)} строк → Обработано: {len(processed_df)} строк"
                )
                messagebox.showinfo(
                    "Успех", f"Обработанные данные сохранены в {file_path}"
                )
                self.set_status("Готов к работе", "info")

            except Exception as e:
                self.log_error(f"Ошибка сохранения: {e}")
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")
                self.set_status("Ошибка сохранения", "error")

    def compare_with_base(self):
        """Сравнение текущего файла с базой данных"""
        try:
            self.log_info("🔍 Начало сравнения с базой данных...")

            # Сбрасываем только флаг добавления товаров, НЕ настройки окон
            self.articles_added = False
            self.log_info("🔄 Состояние добавления товаров сброшено")

            # Проверяем наличие файла поставщика
            if self.current_df is None:
                self.log_error("Файл поставщика не загружен")
                messagebox.showwarning(
                    "Предупреждение", "Сначала загрузите файл поставщика"
                )
                return

            # Запускаем progress bar (5 шагов)
            self.start_progress("Сравнение с базой", 5, "compare")
            self.update_progress(0, "Подготовка к сравнению")

            # Логируем информацию о загруженных файлах
            self.log_info(f"📋 Проверка загруженных файлов:")
            self.log_info(
                f"   💼 Прайс поставщика: {'✅' if self.current_df is not None else '❌'} ({self.current_config or 'неизвестно'})"
            )
            self.log_info(
                f"   🏢 База данных: {'✅' if self.base_df is not None else '❌'}"
            )
            if self.current_df is not None:
                self.log_info(f"   📊 Строк в прайсе: {len(self.current_df):,}")
            if self.base_df is not None:
                self.log_info(f"   📊 Строк в базе: {len(self.base_df):,}")

            # Шаг 1: Проверка и загрузка базы данных
            self.update_progress(1, "Загрузка базы данных")
            if self.auto_load_base_enabled:
                self.set_status("Автозагрузка базы данных...", "loading")
                self.root.update()

                if self.base_df is None:
                    data_dir = "data/input"
                    result = load_largest_file(data_dir, "base")

                    if result is None:
                        self.finish_progress("Ошибка загрузки базы", auto_reset=False)
                        messagebox.showerror(
                            "Ошибка", "Не удалось загрузить базу данных"
                        )
                        return

                    self.base_df, base_file_path = result
                    self.base_file_name = os.path.basename(base_file_path)
                    self.log_info("База данных автоматически загружена")
                    self.update_files_info()
            else:
                if self.base_df is None:
                    self.finish_progress("База не загружена", auto_reset=False)
                    messagebox.showwarning(
                        "Предупреждение",
                        "Сначала загрузите базу данных или включите автозагрузку",
                    )
                    return

            # Шаг 2: Предобработка данных
            self.update_progress(2, "Предобработка данных")
            self.set_status("Предобработка данных...", "loading")
            self.root.update()

            processed_supplier_df = self.preprocess_supplier_data(
                self.current_df, self.current_config
            )

            # Шаг 3: Выполнение сравнения
            self.update_progress(3, "Сравнение данных")
            self.set_status("Сравнение с базой...", "compare")
            self.root.update()

            comparison_result = self.perform_comparison(
                processed_supplier_df, self.base_df
            )

            # Шаг 4: Отображение результатов
            self.update_progress(4, "Формирование отчета")
            self.show_comparison_result(comparison_result)
            self.comparison_result = comparison_result

            if "error" in comparison_result:
                self.log_error(f"Ошибка сравнения: {comparison_result['error']}")
                self.finish_progress("Ошибка сравнения", auto_reset=False)
                return

            # Шаг 5: Завершение
            self.update_progress(5, "Завершение")
            self.update_buttons_state()
            self.finish_progress("Сравнение завершено", auto_reset=True)
            self.set_status("Сравнение завершено", "success")

            # Сохранение размеров окна
            try:
                main_width = self.root.winfo_width()
                main_height = self.root.winfo_height()
                if main_width > 200 and main_height > 200:
                    self.settings["main_window_width"] = main_width
                    self.settings["main_window_height"] = main_height
                    self.save_settings(self.settings)
                    self.log_info("💾 Размеры окна сохранены")
            except Exception as e:
                self.log_error(f"❌ Ошибка сохранения размеров: {e}")

        except Exception as e:
            self.log_error(f"❌ Ошибка при сравнении: {e}")
            self.finish_progress("Ошибка сравнения", auto_reset=False)
            messagebox.showerror("Ошибка", f"Произошла ошибка при сравнении: {e}")

    def perform_comparison(self, supplier_df, base_df):
        """Выполняет сравнение файла поставщика с базой данных"""
        self.set_status("🚀 Начало полного сравнения данных...", "loading")
        self.start_progress("Полное сравнение данных", 10, "comparison")

        # Логируем входные данные
        self.log_info(
            f"🔍 Начинаем сравнение: поставщик {len(supplier_df)} строк, база {len(base_df)} строк"
        )
        self.log_info(f"📊 Колонки поставщика: {list(supplier_df.columns)}")
        self.log_info(f"📊 Колонки базы: {list(base_df.columns)}")

        # 1. ПЕРВЫЙ ЭТАП: Поиск по артикулам
        self.set_status("🔍 Этап 1: Поиск совпадений по артикулам...", "loading")
        self.update_progress(1, "Этап 1: Поиск совпадений по артикулам")
        self.log_info("🔍 Этап 1: Поиск совпадений по артикулам...")
        article_results = self.compare_by_articles(supplier_df, base_df)

        matches = article_results["matches"]
        price_changes = article_results["price_changes"]
        new_items = article_results[
            "new_items"
        ]  # Это товары БЕЗ совпадений по артикулам
        supplier_dict = article_results["supplier_dict"]
        base_dict = article_results["base_dict"]

        # 2. СОЗДАЕМ ДАТАФРЕЙМ ТОВАРОВ БЕЗ СОВПАДЕНИЙ ПО АРТИКУЛАМ
        # Это будут кандидаты для поиска по кодам и дальнейшей обработки ИИ
        self.log_info(
            f"📦 Создаем датафрейм из {len(new_items)} товаров без совпадений по артикулам..."
        )

        unmatched_df = pd.DataFrame(new_items) if new_items else pd.DataFrame()

        if not unmatched_df.empty:
            # Добавляем дополнительную информацию для ИИ обработки
            unmatched_df["search_status"] = "no_article_match"
            unmatched_df["found_by"] = None
            unmatched_df["similarity_score"] = None

            self.log_info(f"📊 unmatched_df колонки: {list(unmatched_df.columns)}")
            if len(unmatched_df) > 0:
                self.log_info(
                    f"📊 Первая строка unmatched_df: {unmatched_df.iloc[0].to_dict()}"
                )

        self.log_info(f"✅ Датафрейм создан: {len(unmatched_df)} строк")

        # 3. ВТОРОЙ ЭТАП: Улучшенный поиск по кодам в скобках с учетом цветов
        self.set_status("🔍 Этап 2: Поиск по кодам в скобках...", "loading")
        self.update_progress(2, "Этап 2: Поиск по кодам в скобках")
        self.log_info(
            "🔍 Этап 2: Улучшенный поиск по кодам в скобках с учетом цветов..."
        )
        bracket_matches = []
        if not unmatched_df.empty:
            bracket_matches = self.compare_by_bracket_codes_advanced(
                supplier_df, base_df, self.current_config, new_items
            )

            # Убираем найденные товары из датафрейма непойсканных
            if bracket_matches:
                found_articles = [
                    match["supplier_article"]
                    for match in bracket_matches
                    if "supplier_article" in match
                ]
                unmatched_df = unmatched_df[
                    ~unmatched_df["article"].isin(found_articles)
                ]
                self.log_info(
                    f"📦 Убрали {len(found_articles)} товаров, найденных по кодам в скобках. Осталось: {len(unmatched_df)}"
                )

        # 4. ТРЕТИЙ ЭТАП: Улучшенный поиск по общим кодам с учетом цветов
        self.set_status("🔍 Этап 3: Поиск по общим кодам...", "loading")
        self.update_progress(3, "Этап 3: Поиск по общим кодам")
        self.log_info("🔍 Этап 3: Улучшенный поиск по общим кодам с учетом цветов...")
        code_matches = []
        if not unmatched_df.empty:
            code_matches = self.compare_by_product_code_advanced(
                supplier_df, base_df, self.current_config, new_items
            )

            # Убираем найденные товары из датафрейма непойсканных
            if code_matches:
                found_articles = [
                    match["supplier_article"]
                    for match in code_matches
                    if "supplier_article" in match
                ]
                unmatched_df = unmatched_df[
                    ~unmatched_df["article"].isin(found_articles)
                ]
                self.log_info(
                    f"📦 Убрали {len(found_articles)} товаров, найденных по общим кодам. Осталось: {len(unmatched_df)}"
                )

        # 5. ЧЕТВЕРТЫЙ ЭТАП: Нечеткий поиск по строкам наименований
        self.set_status("🔍 Этап 4: Нечеткий поиск по наименованиям...", "loading")
        self.update_progress(4, "Этап 4: Нечеткий поиск по наименованиям")
        self.log_info(
            f"🔍 Этап 4: Нечеткий поиск по строкам наименований для {len(unmatched_df)} товаров..."
        )

        fuzzy_candidates = (
            unmatched_df.to_dict("records") if not unmatched_df.empty else []
        )

        self.log_info(
            f"📊 fuzzy_candidates перед вызовом: {len(fuzzy_candidates)} элементов"
        )
        if fuzzy_candidates:
            self.log_info(f"📊 Тип первого элемента: {type(fuzzy_candidates[0])}")
            if isinstance(fuzzy_candidates[0], dict):
                self.log_info(
                    f"📊 Ключи первого элемента: {list(fuzzy_candidates[0].keys())}"
                )

        fuzzy_matches = []

        if fuzzy_candidates:
            self.log_info("🔍 Вызываем compare_by_fuzzy_string_matching...")
            fuzzy_matches = self.compare_by_fuzzy_string_matching(
                fuzzy_candidates, base_df, self.current_config
            )
            self.log_info(
                f"📊 Результат нечеткого поиска: {len(fuzzy_matches)} совпадений"
            )

            # Убираем найденные товары из датафрейма непойсканных
            if fuzzy_matches:
                # Убираем товары по индексу, так как артикул может отсутствовать
                found_indices = [
                    match["supplier_index"]
                    for match in fuzzy_matches
                    if "supplier_index" in match
                ]
                unmatched_df = unmatched_df[~unmatched_df.index.isin(found_indices)]
                self.log_info(
                    f"📦 Убрали {len(found_indices)} товаров, найденных нечетким поиском. Осталось: {len(unmatched_df)}"
                )

        # Обновляем fuzzy_candidates после нечеткого поиска
        self.log_info(
            f"📊 unmatched_df после нечеткого поиска: {len(unmatched_df)} строк"
        )
        if not unmatched_df.empty:
            self.log_info(f"📊 Колонки unmatched_df: {list(unmatched_df.columns)}")
            self.log_info(
                f"📊 Первая строка unmatched_df: {unmatched_df.iloc[0].to_dict()}"
            )

        # Создаем final_unmatched_items для возврата
        final_unmatched_items = (
            unmatched_df.to_dict("records") if not unmatched_df.empty else []
        )

        self.set_status("✅ Сравнение завершено!", "success")
        self.update_progress(5, "Сравнение завершено")

        self.log_info(
            f"📊 final_unmatched_items после преобразования: {len(final_unmatched_items)} элементов"
        )
        if final_unmatched_items:
            self.log_info(f"📊 Тип первого элемента: {type(final_unmatched_items[0])}")
            if isinstance(final_unmatched_items[0], dict):
                self.log_info(
                    f"📊 Ключи первого элемента: {list(final_unmatched_items[0].keys())}"
                )

        return {
            "supplier_total": len(supplier_dict),
            "base_total": len(base_dict),
            "matches": matches,
            "price_changes": price_changes,
            "new_items": new_items,
            "code_matches": code_matches,  # Поиск по общим кодам
            "bracket_matches": bracket_matches,  # Поиск по кодам в скобках
            "fuzzy_matches": fuzzy_matches,  # Нечеткий поиск по строкам наименований
            "fuzzy_candidates": final_unmatched_items,  # Товары без совпадений после всех методов поиска
            "unmatched_count": len(unmatched_df),  # Количество непойсканных товаров
            "match_rate": (
                len(matches) / len(supplier_dict) * 100 if supplier_dict else 0
            ),
        }

    def show_comparison_result(self, result):
        """Показ результатов сравнения"""
        if "error" in result:
            messagebox.showerror("Ошибка", result["error"])
            return

        # Очищаем только текстовое поле, НЕ сбрасывая данные
        self.info_text.delete(1.0, tk.END)

        info = f"🔍 РЕЗУЛЬТАТЫ СРАВНЕНИЯ С БАЗОЙ ДАННЫХ\n"
        info += f"{'='*60}\n"
        info += f"Конфигурация: {self.current_config}\n"
        info += f"Дата сравнения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

        # Общая статистика
        info += f"📊 ОБЩАЯ СТАТИСТИКА:\n"
        info += f"{'-'*40}\n"
        info += f"Товаров у поставщика: {result['supplier_total']:,}\n"
        info += f"Товаров в базе: {result['base_total']:,}\n"
        info += f"Совпадений найдено: {len(result['matches']):,}\n"
        info += f"Процент совпадений: {result['match_rate']:.1f}%\n\n"

        # Изменения цен
        if result["price_changes"]:
            info += f"💰 ЗНАЧИТЕЛЬНЫЕ ИЗМЕНЕНИЯ ЦЕН (>5%):\n"
            info += f"{'-'*50}\n"
            for i, item in enumerate(result["price_changes"][:10], 1):
                change_sign = "📈" if item["price_diff"] > 0 else "📉"
                info += f"{i:2d}. {item['article']}: {item['base_price']:.2f} → {item['supplier_price']:.2f} "
                info += f"({item['price_change_percent']:+.1f}%) {change_sign}\n"
            if len(result["price_changes"]) > 10:
                info += f"... и еще {len(result['price_changes']) - 10} изменений\n"
        else:
            info += f"💰 ИЗМЕНЕНИЯ ЦЕН: Значительных изменений не найдено ✅\n"

        info += "\n"

        # Новые товары
        if result["new_items"]:
            info += f"🆕 НОВЫЕ ТОВАРЫ У ПОСТАВЩИКА:\n"
            info += f"{'-'*40}\n"
            for i, item in enumerate(result["new_items"][:10], 1):
                info += f"{i:2d}. {item['article']}: {item['price']:.2f} - {item['name'][:30]}\n"

                # Показываем информацию о возможном совпадении
                if (
                    item.get("fuzzy_match_name")
                    and item.get("fuzzy_match_similarity", 0) > 0
                ):
                    similarity_percent = item["fuzzy_match_similarity"] * 100
                    info += f"    🔍 Возможное совпадение: {item['fuzzy_match_name'][:40]}...\n"
                    info += f"    📍 Строка в базе: {item['fuzzy_match_row']}, Цвет: {item['fuzzy_match_color']}, Цена: {item['fuzzy_match_price']}\n"
                    info += f"    📊 Схожесть: {similarity_percent:.1f}%\n"
                else:
                    info += f"    ❌ Совпадений не найдено\n"
                info += "\n"

            if len(result["new_items"]) > 10:
                info += f"... и еще {len(result['new_items']) - 10} новых товаров\n"
        else:
            info += f"🆕 НОВЫЕ ТОВАРЫ: Не найдено\n"

        # Результаты поиска по кодам
        if result.get("code_matches"):
            info += f"\n🔍 СОВПАДЕНИЯ ПО КОДАМ В НАИМЕНОВАНИЯХ (общие коды):\n"
            info += f"{'-'*50}\n"
            for i, match in enumerate(result["code_matches"][:5], 1):
                info += f"{i:2d}. Код: {match['code']}\n"
                info += f"    Поставщик: {match['supplier_name'][:40]}...\n"
                info += f"    База: {match['base_name'][:40]}...\n"
                info += f"    Цены: {match['supplier_price']:.2f} ↔ {match['base_price']:.2f}\n\n"
            if len(result["code_matches"]) > 5:
                info += (
                    f"... и еще {len(result['code_matches']) - 5} совпадений по кодам\n"
                )

        # Результаты поиска по кодам в скобках
        if result.get("bracket_matches"):
            info += f"\n🔍 СОВПАДЕНИЯ ПО КОДАМ В СКОБКАХ (наименования):\n"
            info += f"{'-'*50}\n"
            for i, match in enumerate(result["bracket_matches"][:5], 1):
                info += f"{i:2d}. Код: {match['code']}\n"
                info += f"    Поставщик: {match['supplier_name'][:40]}...\n"
                info += f"    База: {match['base_name'][:40]}...\n"
                info += f"    Найдено в: {match['matched_in']}\n"
                info += f"    Цены: {match['supplier_price']:.2f} ↔ {match['base_price']:.2f}\n\n"
            if len(result["bracket_matches"]) > 5:
                info += f"... и еще {len(result['bracket_matches']) - 5} совпадений по кодам в скобках\n"

        # Результаты нечеткого поиска
        if result.get("fuzzy_matches"):
            info += f"\n🔍 СОВПАДЕНИЯ ПО НЕЧЕТКОМУ ПОИСКУ (строки наименований):\n"
            info += f"{'-'*50}\n"
            for i, match in enumerate(result["fuzzy_matches"][:5], 1):
                info += f"{i:2d}. Схожесть: {match['similarity_ratio']:.1%}\n"
                info += f"    Поставщик: {match['supplier_name'][:40]}...\n"
                info += f"    База: {match['base_name'][:40]}...\n"
                info += f"    Цены: {match['supplier_price']:.2f} ↔ {match['base_price']:.2f}\n\n"
            if len(result["fuzzy_matches"]) > 5:
                info += f"... и еще {len(result['fuzzy_matches']) - 5} совпадений по нечеткому поиску\n"

        # Информация о товарах без совпадений
        if result.get("fuzzy_candidates"):
            info += f"\n🤖 ТОВАРЫ БЕЗ СОВПАДЕНИЙ (после всех методов поиска):\n"
            info += f"{'-'*50}\n"
            info += f"Товаров без совпадений: {result.get('unmatched_count', 0)}\n"
            info += f"Остались после всех методов поиска (артикулы, коды, нечеткий поиск): {len(result['fuzzy_candidates'])}\n"

            # Показываем примеры
            for i, candidate in enumerate(result["fuzzy_candidates"][:3], 1):
                info += f"{i:2d}. {candidate['article']}: ${candidate['price']:.2f}\n"
                info += f"    {candidate['name'][:50]}...\n"

            if len(result["fuzzy_candidates"]) > 3:
                info += f"... и еще {len(result['fuzzy_candidates']) - 3} товаров\n"
        elif result.get("unmatched_count", 0) == 0:
            info += f"\n🎉 ВСЕ ТОВАРЫ НАЙДЕНЫ! Нет товаров без совпадений после всех методов поиска.\n"
        else:
            info += f"\n📝 Непойсканных товаров (после всех методов поиска): {result.get('unmatched_count', 0)}\n"

        info += f"\n🎉 СРАВНЕНИЕ ЗАВЕРШЕНО!"

        self.info_text.insert(tk.END, info)
        self.log_info(
            f"Сравнение завершено: {len(result['matches'])} совпадений из {result['supplier_total']} товаров"
        )

    def clean_invisible_chars(self, text):
        """Убирает невидимые и непечатаемые символы из текста"""
        if pd.isna(text):
            return None

        # Преобразуем в строку
        text_str = str(text).strip()

        cleaned = " ".join(text_str.split())

        return cleaned if cleaned else None

    def _calculate_similarity(self, text1, text2):
        """Вычисляет схожесть между двумя текстами"""
        if not text1 or not text2:
            return 0.0

        import difflib

        return difflib.SequenceMatcher(
            None, str(text1).lower(), str(text2).lower()
        ).ratio()

    def clean_article_vitya_simple(self, article):
        """Простая очистка артикула Вити - убираем ТОЛЬКО апострофы и префикс '000, результат ВСЕГДА int"""
        # Проверяем на NaN или None
        if pd.isna(article) or article is None:
            return None

        # Преобразуем в строку
        cleaned = str(article).strip()

        # Проверяем, что получилась не пустая строка и не 'nan'
        if not cleaned or cleaned.lower() == "nan":
            return None

        # 1. Убираем ТОЛЬКО апострофы
        cleaned = cleaned.replace("'", "")

        # 2. Убираем префикс '000 если есть
        if cleaned.startswith("000"):
            cleaned = cleaned[3:]

        # 3. ВСЕГДА преобразуем в int
        if cleaned.isdigit():
            return int(cleaned) if cleaned else 0
        elif cleaned == "":
            return 0
        else:
            # Если есть нецифровые символы, извлекаем только цифры
            import re

            digits = re.findall(r"\d+", cleaned)
            if digits:
                return int("".join(digits))
            else:
                return 0

    def filter_by_price(self, df, price_column="price_usd"):
        """
        Фильтрация данных по цене - убирает строки где price_usd является NaN, пустой или <= 0

        Args:
            df: DataFrame для фильтрации
            price_column: название столбца с ценой (по умолчанию 'price_usd')

        Returns:
            DataFrame с отфильтрованными данными
        """
        if price_column not in df.columns:
            self.log_info(
                f"⚠️ Столбец '{price_column}' не найден, фильтрация по цене пропущена"
            )
            return df

        initial_count = len(df)

        # Фильтруем: убираем NaN, пустые значения и цены <= MIN_PRICE_THRESHOLD
        filtered_df = df[
            (df[price_column].notna()) & (df[price_column] > MIN_PRICE_THRESHOLD)
        ].copy()

        final_count = len(filtered_df)
        removed_count = initial_count - final_count

        if removed_count > 0:
            self.log_info(f"💰 Фильтрация по цене ({price_column}):")
            self.log_info(f"   Удалено строк: {removed_count}")
            self.log_info(f"   Осталось строк: {final_count}")

            # Показываем статистику удаленных строк
            nan_count = df[price_column].isna().sum()
            zero_count = (df[price_column] == 0).sum()
            low_price_count = (
                (df[price_column] > 0) & (df[price_column] <= MIN_PRICE_THRESHOLD)
            ).sum()

            self.log_info(f"   📊 Причины удаления:")
            if nan_count > 0:
                self.log_info(f"      NaN/пустые: {nan_count}")
            if zero_count > 0:
                self.log_info(f"      Нулевые цены: {zero_count}")
            if low_price_count > 0:
                self.log_info(
                    f"      Слишком низкие (<={MIN_PRICE_THRESHOLD}): {low_price_count}"
                )
        else:
            self.log_info(
                f"✅ Фильтрация по цене: все {final_count} строк прошли фильтр"
            )

        return filtered_df

    def preprocess_vitya_fixed_v3(self, df):
        """ИСПРАВЛЕННАЯ предобработка для Вити с фильтрацией по цене и балансу"""
        self.log_info("🔧 Запуск предобработки для Витя...")

        # Копируем данные для безопасности
        processed_df = df.copy()
        initial_count = len(processed_df)

        # 1. Фильтрация по цене - убираем строки с NaN, пустыми или нулевыми ценами
        self.log_info("💰 Фильтруем по цене...")
        processed_df = self.filter_by_price(processed_df, "price_usd")

        # 2. Фильтрация по балансу - оставляем только товары в наличии И на распродаже
        if "balance" in processed_df.columns:
            self.log_info(
                f"📦 Фильтруем по балансу (оставляем только {VITYA_BALANCE_AVAILABLE})..."
            )

            balance_before = len(processed_df)
            # Новая логика: фильтруем по списку значений
            processed_df = processed_df[
                processed_df["balance"].isin(VITYA_BALANCE_AVAILABLE)
            ].copy()
            balance_after = len(processed_df)

            removed_balance = balance_before - balance_after
            if removed_balance > 0:
                self.log_info(f"   📦 Удалено строк без наличия: {removed_balance}")
                self.log_info(f"   📦 Осталось строк в наличии: {balance_after}")

                # Показываем статистику по каждому типу баланса
                for status in VITYA_BALANCE_AVAILABLE:
                    status_count = (processed_df["balance"] == status).sum()
                    if status_count > 0:
                        self.log_info(f"      '{status}': {status_count} товаров")
            else:
                self.log_info(f"   📦 Все {balance_after} строк имеют товары в наличии")
        else:
            self.log_info(
                "⚠️ Столбец 'balance' не найден, фильтрация по наличию пропущена"
            )

        # 3. Очистка артикулов - активируем очистку
        if "article_vitya" in processed_df.columns:
            self.log_info("🧹 Очистка артикулов Витя...")

            processed_df["article_vitya"] = processed_df["article_vitya"].apply(
                self.clean_article_vitya_simple
            )

        # 4. Добавляем метку поставщика
        processed_df["supplier_name"] = "Витя"

        # 5. Отладочная информация
        if "article_vitya" in processed_df.columns:
            valid_articles = processed_df["article_vitya"].notna().sum()
            self.log_info(f"🔢 Валидных артикулов Витя: {valid_articles}")
            if valid_articles > 0:
                sample_articles = (
                    processed_df["article_vitya"].dropna().head(5).tolist()
                )
                self.log_info(f"📝 Примеры артикулов: {sample_articles}")

        # 6. Финальная статистика
        final_count = len(processed_df)
        total_removed = initial_count - final_count

        self.log_info(f"✅ Предобработка Витя завершена:")
        self.log_info(f"   📊 Исходно: {initial_count} строк")
        self.log_info(f"   📊 Итого: {final_count} строк")
        self.log_info(f"   📊 Удалено: {total_removed} строк")

        return processed_df

    def preprocess_dimi_fixed(self, df):
        """УПРОЩЕННАЯ предобработка данных для поставщика Дима с фильтрацией"""
        self.log_info("🔧 Запуск предобработки для Дима...")

        # Копируем данные
        processed_df = df.copy()
        initial_count = len(processed_df)

        # 1. Фильтрация по цене - убираем строки с NaN, пустыми или нулевыми ценами
        self.log_info("💰 Фильтруем по цене...")
        processed_df = self.filter_by_price(processed_df, "price_usd")

        # 2. Фильтрация по балансу - убираем строки где balance или balance1 = "Ожидается"
        balance_columns = ["balance", "balance1"]
        found_balance_columns = [
            col for col in balance_columns if col in processed_df.columns
        ]

        if found_balance_columns:
            self.log_info(
                f"📦 Фильтруем по балансу (убираем '{DIMI_BALANCE_EXPECTED}')..."
            )

            balance_before = len(processed_df)

            # Создаем условие: НИ balance, НИ balance1 не должны быть "Ожидается"
            for col in found_balance_columns:
                processed_df = processed_df[processed_df[col] != DIMI_BALANCE_EXPECTED]

            processed_df = processed_df.copy()
            balance_after = len(processed_df)

            removed_balance = balance_before - balance_after
            if removed_balance > 0:
                self.log_info(
                    f"   📦 Удалено строк с '{DIMI_BALANCE_EXPECTED}': {removed_balance}"
                )
                self.log_info(f"   📦 Осталось строк в наличии: {balance_after}")

                # Показываем детализацию по столбцам
                for col in found_balance_columns:
                    expected_count = (df[col] == DIMI_BALANCE_EXPECTED).sum()
                    if expected_count > 0:
                        self.log_info(
                            f"      {col}: {expected_count} строк с '{DIMI_BALANCE_EXPECTED}'"
                        )
            else:
                self.log_info(
                    f"   📦 Все {balance_after} строк прошли фильтр по балансу"
                )
        else:
            self.log_info(
                "⚠️ Столбцы balance/balance1 не найдены, фильтрация по наличию пропущена"
            )

        # 3. Очистка артикулов - активируем очистку
        if "article_dimi" in processed_df.columns:
            self.log_info("🧹 Очистка артикулов Дима...")

            def clean_article_dimi_simple(article):
                """Упрощенная очистка артикула Димы - ТОЛЬКО апострофы и префикс '000"""
                if pd.isna(article):
                    return None

                # Преобразуем в строку и убираем лишние пробелы
                cleaned = str(article).strip()

                if not cleaned or cleaned.lower() == "nan":
                    return None

                # Убираем апострофы
                cleaned = cleaned.replace("'", "")

                # ДЛЯ ДИМЫ: Убираем префикс '000 если есть
                if cleaned.startswith("000"):
                    cleaned = cleaned[3:]

                return cleaned if cleaned else None

            processed_df["article_dimi"] = processed_df["article_dimi"].apply(
                clean_article_dimi_simple
            )

        # 4. Добавляем метку поставщика
        processed_df["supplier_name"] = "Дима"

        # 5. Отладочная информация
        if "article_dimi" in processed_df.columns:
            valid_articles = processed_df["article_dimi"].notna().sum()
            self.log_info(f"🔢 Валидных артикулов Дима: {valid_articles}")
            if valid_articles > 0:
                sample_articles = processed_df["article_dimi"].dropna().head(5).tolist()
                self.log_info(f"📝 Примеры артикулов: {sample_articles}")

        # 6. Финальная статистика
        final_count = len(processed_df)
        total_removed = initial_count - final_count

        self.log_info(f"✅ Предобработка Дима завершена:")
        self.log_info(f"   📊 Исходно: {initial_count} строк")
        self.log_info(f"   📊 Итого: {final_count} строк")
        self.log_info(f"   📊 Удалено: {total_removed} строк")

        return processed_df

    def preprocess_supplier_data(self, df, config_name):
        """Универсальная предобработка в зависимости от конфига"""

        if config_name == "vitya":
            return self.preprocess_vitya_fixed_v3(df)
        elif config_name == "dimi":
            return self.preprocess_dimi_fixed(df)
        else:
            self.log_info(f"📋 Предобработка для {config_name} не требуется")
            return df

    def detect_config_by_filename(self, file_path):
        """Автоматическое определение конфига по имени файла"""

        filename = os.path.basename(file_path).upper()  # Имя файла в верхнем регистре

        self.log_info(f"🔍 Определение конфига для файла: {filename}")

        # Правила определения конфига
        if "JHT" in filename:
            detected_config = "vitya"
            self.log_info("✅ Обнаружен прайс Вити (содержит JHT)")

        elif "DIMI" in filename or "DIMA" in filename:
            detected_config = "dimi"
            self.log_info("✅ Обнаружен прайс Димы (содержит DiMi/DiMa)")

        elif "BASE" in filename or "БАЗА" in filename:
            detected_config = "base"
            self.log_info("✅ Обнаружена база данных (содержит BASE/БАЗА)")

        else:
            detected_config = "auto"  # По умолчанию
            self.log_info("ℹ️ Конфиг не определен, используется AUTO")

        return detected_config

    def find_product_code_in_name(self, product_name):
        """Извлечение кода товара из наименования"""
        if pd.isna(product_name) or not isinstance(product_name, str):
            return None

        import re

        # Улучшенные паттерны для поиска кодов (только заглавные буквы, цифры и тире)
        patterns = [
            # Коды с тире: AC-M25-SC, P27QDA-RGP и т.д.
            r"\b[A-Z0-9]+(?:-[A-Z0-9]+)+\b",  # Один или более блоков через тире
            # Коды без тире: MJKDDYJ02HT, XM123 и т.д.
            r"\b[A-Z]{2,}[0-9]{2,}[A-Z]*\b",  # Буквы-цифры-буквы (XM123, MJKDDYJ02HT)
            r"\b[0-9]{3,}[A-Z]{1,3}\b",  # Цифры-буквы (123XM)
            r"\b[A-Z][0-9]{4,}[A-Z][0-9]+\b",  # Паттерн M2319E1
            # Бренды и коды только из заглавных букв: GREENOE, XIAOMI, SAMSUNG и т.д.
            r"\b[A-Z]{4,8}\b",  # 4-8 заглавные буквы подряд (GREENOE, XIAOMI, SAMSUNG)
            # Коды типа C60, C20, D30 и т.д. (буква + цифры)
            r"\b[A-Z][0-9]{1,3}\b",  # C60, C20, D30
            # Коды типа С20 (кириллица + цифры)
            r"\b[А-Я][0-9]{1,3}\b",  # С20, Д30
        ]

        for pattern in patterns:
            matches = re.findall(pattern, product_name.upper())
            if matches:
                # Фильтруем найденные совпадения, исключая характеристики товара
                for match in matches:
                    # Исключаем характеристики типа "60000mah", "100w", "20000mah" и т.д.
                    if re.match(r"^\d+mah$|^\d+w$|^\d+wh$|^\d+ma$", match.lower()):
                        continue

                    # Исключаем слишком длинные числовые коды (характеристики)
                    if re.match(r"^\d{5,}$", match):
                        continue

                    # Исключаем коды типа "USB-C", "POWER" и т.д.
                    if match in ["USB-C", "POWER", "PORTABLE", "CHARGER", "BANK"]:
                        continue

                    # Для брендов (только заглавные буквы) исключаем слишком длинные названия
                    if re.match(r"^[A-Z]{4,8}$", match) and len(match) > 8:
                        continue

                    # Возвращаем первое подходящее совпадение
                    return match

        return None

    def find_product_code_in_brackets(self, product_name):
        """
        Поиск кода товара в скобках из наименования
        Ищет коды типа (P27QDA-RGP), (XM123) и т.д.
        Принимает заглавные/строчные буквы, цифры, тире
        """

        if pd.isna(product_name) or not isinstance(product_name, str):
            return None

        import re

        # Ищем коды в скобках
        # Паттерн для кодов в скобках: (любые символы кроме скобок)
        bracket_pattern = r"\(([^)]+)\)"

        matches = re.findall(bracket_pattern, product_name)

        if matches:
            # Берем первое найденное совпадение
            code = matches[0].strip().upper()

            # Код должен содержать только буквы (заглавные/строчные/кириллица), цифры и тире, минимум 4 символа
            if re.match(r"^[A-Za-zА-Яа-я0-9\-]+$", code) and len(code) >= 4:
                return code

        return None

    def find_battery_capacity(self, product_name):
        """
        Извлекает емкость батареи (mAh) из наименования товара

        Примеры:
        - "Повербанк GREENOE Protable Power bank 60000mah 100w (C60)" -> "60000"
        - "Повербанк GREENOE Protable Power bank 20000MAH 25w (цвет white)" -> "20000"
        - "Power Bank 10000mAh Portable Charger" -> "10000"
        """
        if pd.isna(product_name) or not isinstance(product_name, str):
            return None

        # Приводим к нижнему регистру для унификации
        name_lower = product_name.lower()

        # Ищем паттерны емкости батареи
        import re

        # Паттерны для поиска емкости: число + mah/mAh/MAH
        patterns = [
            r"(\d+)\s*mah",  # 60000 mah
            r"(\d+)mah",  # 60000mah
            r"(\d+)\s*mah",  # 60000mAh
            r"(\d+)mah",  # 60000mAh
            r"(\d+)\s*мач",  # 60000 мач (кириллица)
            r"(\d+)мач",  # 60000мач
            r"(\d+)\s*мч",  # 60000 мч (сокращение)
            r"(\d+)мч",  # 60000мч
        ]

        for pattern in patterns:
            match = re.search(pattern, name_lower)
            if match:
                capacity = match.group(1)
                # Проверяем, что это разумное значение емкости (от 50 до 999999)
                try:
                    capacity_int = int(capacity)
                    if 5 <= capacity_int <= 999999:
                        return capacity
                except ValueError:
                    continue

        return None

    def find_product_code_unified(self, product_name):
        """
        Объединенная функция поиска кодов товара
        Приоритет: 1) Коды в скобках (≥4 символов), 2) Бренды в наименовании (GREENOE, XIAOMI, SAMSUNG), 3) Другие коды (≥4 символов)
        """
        if pd.isna(product_name) or not isinstance(product_name, str):
            return None

        # 1. Сначала ищем коды в скобках (приоритет, но только ≥4 символов)
        bracket_code = self.find_product_code_in_brackets(product_name)
        if bracket_code:
            return bracket_code

        # 2. Ищем бренды в наименовании (GREENOE, XIAOMI, SAMSUNG)
        # Приводим к верхнему регистру для поиска
        name_upper = product_name.upper()

        # Список известных брендов
        brands = [
            "GREENOE",
            "XIAOMI",
            "SAMSUNG",
            "APPLE",
            "HUAWEI",
            "OPPO",
            "VIVO",
            "ONEPLUS",
        ]

        for brand in brands:
            if brand in name_upper:
                return brand

        # 3. Ищем другие коды в наименовании (только заглавные, ≥4 символов)
        name_code = self.find_product_code_in_name(product_name)
        if name_code and len(name_code) >= 4:
            return name_code

        return None

    def get_supplier_article_column(self):
        """Получение столбца артикула поставщика для текущего конфига"""
        if self.current_config == "vitya":
            return "article_vitya"
        elif self.current_config == "dimi":
            return "article_dimi"
        else:
            return "article"

    def get_supplier_price_column(self):
        """Получение столбца цены поставщика для текущего конфига"""
        if self.current_config == "vitya":
            return "price_usd"
        elif self.current_config == "dimi":
            return "price_usd"
        else:
            return "price"

    def get_base_article_column(self):
        """Получение столбца артикула в базе для текущего конфига"""
        if self.current_config == "vitya":
            return "article_vitya"
        elif self.current_config == "dimi":
            return "article_dimi"
        else:
            return "article"

    def get_base_price_column(self):
        """Получение столбца цены в базе для текущего конфига"""
        if self.current_config == "vitya":
            return "price_vitya_usd"
        elif self.current_config == "dimi":
            return "price_dimi_usd"
        else:
            return "price"

    def get_column_data_type(self, column_name):
        """Получение типа данных столбца из конфигурации базы"""
        base_config_path = "excel_loader/configs/base_config.json"
        try:
            with open(base_config_path, "r", encoding="utf-8") as f:
                base_config = json.load(f)
            data_types = base_config.get("data_types", {})
            return data_types.get(column_name, "int")  # По умолчанию int
        except Exception as e:
            self.log_error(f"Ошибка загрузки конфига базы: {e}")
            return "int"

    def get_excel_column_name_from_config(self, pandas_column_name):
        """Получить оригинальное имя столбца Excel из конфигурации базы"""
        try:
            # Загружаем конфигурацию базы
            base_config_path = "excel_loader/configs/base_config.json"
            with open(base_config_path, "r", encoding="utf-8") as f:
                base_config = json.load(f)

            # Ищем в column_mapping обратное соответствие
            column_mapping = base_config.get("column_mapping", {})
            for excel_name, pandas_name in column_mapping.items():
                if pandas_name == pandas_column_name:
                    return excel_name

            # Если не найдено, возвращаем исходное имя
            return pandas_column_name

        except Exception as e:
            self.log_error(f"Ошибка чтения конфигурации базы: {e}")
            return pandas_column_name

    def get_pandas_column_name_from_excel_name(self, excel_column_name):
        """Получить pandas название столбца из оригинального названия Excel"""
        try:
            # Загружаем конфигурацию базы
            base_config_path = "excel_loader/configs/base_config.json"
            with open(base_config_path, "r", encoding="utf-8") as f:
                base_config = json.load(f)

            # Ищем в column_mapping соответствие
            column_mapping = base_config.get("column_mapping", {})
            for excel_name, pandas_name in column_mapping.items():
                if excel_name.lower().strip() == excel_column_name.lower().strip():
                    return pandas_name

            # Если не найдено, возвращаем исходное имя
            return excel_column_name

        except Exception as e:
            self.log_error(f"Ошибка чтения конфигурации базы: {e}")
            return excel_column_name

    def _get_supplier_name_column(self, supplier_df):
        """
        Определяет название колонки с названиями товаров в данных поставщика

        Args:
            supplier_df: DataFrame поставщика

        Returns:
            str: Название колонки или None если не найдено
        """
        # Проверяем, что supplier_df не пустой и имеет колонки
        if (
            supplier_df is None
            or supplier_df.empty
            or not hasattr(supplier_df, "columns")
        ):
            return None

        # Сначала ищем стандартную колонку 'name'
        if "name" in supplier_df.columns:
            return "name"

        # Если не найдено, ищем по текущему конфигу
        if self.current_config == "vitya":
            # Для Вити ищем колонку 'Unnamed: 1' (как в конфиге)
            for col in supplier_df.columns:
                if "Unnamed: 1" in str(col):
                    return col
        elif self.current_config == "dimi":
            # Для Димы ищем колонку с названиями (обычно вторая колонка)
            if len(supplier_df.columns) > 1:
                return supplier_df.columns[1]

        # Если ничего не найдено, возвращаем None
        return None

    def _get_base_name_column(self, base_df):
        """
        Определяет название колонки с названиями товаров в базе данных

        Args:
            base_df: DataFrame базы данных

        Returns:
            str: Название колонки или None если не найдено
        """
        # Проверяем, что base_df не пустой и имеет колонки
        if base_df is None or base_df.empty or not hasattr(base_df, "columns"):
            return None

        # Сначала ищем стандартную колонку 'name'
        if "name" in base_df.columns:
            return "name"

        # Если не найдено, ищем колонку 'Наименование' (как в конфиге базы)
        if "Наименование" in base_df.columns:
            return "Наименование"

        # Если ничего не найдено, возвращаем None
        return None

    def safe_color_processing(self, color_value):
        """
        Безопасная обработка цвета с учетом NaN значений

        Args:
            color_value: Значение цвета (может быть строкой, NaN, None)

        Returns:
            str: Обработанный цвет в нижнем регистре или пустая строка
        """
        if color_value is None or pd.isna(color_value):
            return ""

        # Преобразуем в строку и обрабатываем
        color_str = str(color_value).strip()
        if not color_str or color_str.lower() in ["nan", "none", ""]:
            return ""

        return color_str.lower()

    def get_base_price_from_config(self, row):
        """
        Получает цену из конфигурации (например, price_vitya_usd для конфига vitya)
        Используется для сравнения и обновления цен

        Args:
            row: Строка DataFrame

        Returns:
            float: Цена из конфигурации или 0.0 если цена пустая
        """
        base_price_col = self.get_base_price_column()
        if (
            base_price_col in row
            and pd.notna(row[base_price_col])
            and row[base_price_col] > 0
        ):
            return float(row[base_price_col])
        return 0.0

    def get_min_base_price(self, row):
        """
        Получает минимальную цену из колонок price_dimi_usd, price_vitya_usd, price_mila_usd
        Используется только для новых товаров при добавлении в базу

        Args:
            row: Строка DataFrame

        Returns:
            float: Минимальная цена или 0.0 если все цены пустые
        """
        prices = []
        for col in ["price_dimi_usd", "price_vitya_usd", "price_mila_usd"]:
            if col in row and pd.notna(row[col]) and row[col] > 0:
                # Приводим к float для единообразия типов
                price = float(row[col])
                prices.append(price)

        return min(prices) if prices else 0.0

    def compare_by_articles(self, supplier_df, base_df):
        """Поиск совпадений строго по артикулам"""
        self.set_status("🔍 Начало сравнения по артикулам...", "loading")
        self.update_progress(1, "Начало сравнения по артикулам")

        # Получаем столбцы для текущего конфига
        supplier_article_col = self.get_supplier_article_column()
        supplier_price_col = self.get_supplier_price_column()
        base_article_col = self.get_base_article_column()
        base_price_col = self.get_base_price_column()

        # Удаляем строки с пустыми артикулами и ценами
        supplier_clean = supplier_df.dropna(
            subset=[supplier_article_col, supplier_price_col]
        )
        base_clean = base_df.dropna(subset=[base_article_col])

        # Создаем словари для быстрого поиска
        self.set_status("📊 Создание словаря товаров поставщика...", "loading")
        self.update_progress(2, "Создание словаря товаров поставщика")

        supplier_dict = {}
        for _, row in supplier_clean.iterrows():
            article_value = row[supplier_article_col]
            # Для article_vitya используем int значение напрямую, для других - строку
            if self.current_config == "vitya" and isinstance(article_value, int):
                article = str(article_value)
            else:
                article = str(article_value).strip()

            if article and article != "nan" and article != "None":
                # Приводим цену к float для единообразия типов
                price = (
                    float(row[supplier_price_col])
                    if pd.notna(row[supplier_price_col])
                    else 0.0
                )
                supplier_dict[article] = {
                    "price": price,
                    "name": row.get("name", ""),
                    "index": row.name,
                    "color": self.safe_color_processing(
                        row.get("color")
                    ),  # Сохраняем цвет для проверки
                }

        self.set_status("📊 Создание словаря базы данных...", "loading")
        self.update_progress(2, "Создание словаря базы данных")

        base_dict = {}
        for _, row in base_clean.iterrows():
            article_value = row[base_article_col]
            # Для article_vitya используем int значение напрямую, для других - строку
            if self.current_config == "vitya" and isinstance(article_value, int):
                article = str(article_value)
            else:
                article = str(article_value).strip()

            if article and article != "nan" and article != "None":
                base_dict[article] = {
                    "price": self.get_base_price_from_config(
                        row
                    ),  # Используем цену из конфигурации
                    "name": row.get("name", ""),
                    "index": row.name,
                    "color": self.safe_color_processing(
                        row.get("color")
                    ),  # Сохраняем цвет для проверки
                }

        # Анализируем совпадения
        self.set_status("🔍 Анализ совпадений по артикулам...", "loading")
        self.update_progress(3, "Анализ совпадений по артикулам")

        matches = []
        price_changes = []
        new_items = []

        for article, supplier_data in supplier_dict.items():
            if article in base_dict:
                base_data = base_dict[article]
                # Проверяем, действительно ли цены отличаются
                price_diff = abs(supplier_data["price"] - base_data["price"])
                prices_equal = price_diff < 0.001

                self.log_info(
                    f"🔍 Сравнение {article}: supplier={supplier_data['price']}, base={base_data['price']}, diff={price_diff:.6f}, equal={prices_equal}"
                )

                match_info = {
                    "article": article,
                    "supplier_price": supplier_data["price"],
                    "base_price": base_data["price"],
                    "name": supplier_data["name"] or base_data["name"],
                    "price_diff": supplier_data["price"] - base_data["price"],
                    "price_change_percent": 0,
                    "base_index": base_data[
                        "index"
                    ],  # Индекс строки в базе для прямого обновления
                }

                if base_data["price"] > 0:
                    match_info["price_change_percent"] = (
                        (supplier_data["price"] - base_data["price"])
                        / base_data["price"]
                        * 100
                    )

                matches.append(match_info)

                # Логируем созданное совпадение
                self.log_info(
                    f"🔍 Создано совпадение {article}: supplier={supplier_data['price']} ({type(supplier_data['price'])}), base={base_data['price']} ({type(base_data['price'])}), change={match_info['price_change_percent']:.1f}%"
                )

                # Значительные изменения цены (больше 5%)
                if abs(match_info["price_change_percent"]) > 5:
                    price_changes.append(match_info)
            else:
                # Ищем возможные совпадения по нечеткому поиску для новых товаров
                (
                    fuzzy_match_name,
                    fuzzy_match_row,
                    fuzzy_match_color,
                    fuzzy_match_price,
                ) = self.find_item_by_fuzzy_matching(supplier_data["name"])

                new_items.append(
                    {
                        "article": article,  # Артикул поставщика
                        "price": supplier_data["price"],
                        "name": supplier_data["name"],
                        "color": supplier_data.get("color", ""),  # Добавляем цвет
                        "supplier_article": article,  # Артикул поставщика (для отчета)
                        "base_article": "",  # Артикул в базе (пустой для новых товаров)
                        "supplier_article_col": self.get_supplier_article_column(),  # Название столбца поставщика
                        "base_article_col": self.get_base_article_column(),  # Название столбца базы
                        # Информация о возможном совпадении по нечеткому поиску
                        "fuzzy_match_name": (
                            fuzzy_match_name if fuzzy_match_name != "Не найдено" else ""
                        ),
                        "fuzzy_match_row": (
                            fuzzy_match_row if fuzzy_match_row != "N/A" else ""
                        ),
                        "fuzzy_match_color": (
                            fuzzy_match_color if fuzzy_match_color != "N/A" else ""
                        ),
                        "fuzzy_match_price": (
                            fuzzy_match_price if fuzzy_match_price != "N/A" else ""
                        ),
                        "fuzzy_match_similarity": (
                            self._calculate_similarity(
                                supplier_data["name"], fuzzy_match_name
                            )
                            if fuzzy_match_name != "Не найдено"
                            else 0.0
                        ),
                    }
                )

        self.set_status("✅ Сравнение по артикулам завершено!", "success")
        self.update_progress(4, "Сравнение по артикулам завершено")

        self.log_info(f"✅ Найдено совпадений по артикулам: {len(matches)}")
        return {
            "matches": matches,
            "price_changes": price_changes,
            "new_items": new_items,
            "supplier_dict": supplier_dict,
            "base_dict": base_dict,
        }

    def compare_by_product_code_advanced(
        self, supplier_df, base_df, supplier_config, new_items_list=None
    ):
        """
        Улучшенный поиск совпадений по кодам товаров с учетом цветов и емкости батареи

        Логика:
        1. Ищем совпадения по кодам среди новых товаров
        2. Сравниваем цвета
        3. Сравниваем емкость батареи (mAh)
        4. Если цвета или емкость не совпадают, ищем тот же код с другими параметрами
        5. Проверяем все варианты кодов
        """

        self.log_info(
            "🔍 Улучшенный поиск совпадений по кодам с учетом цветов и емкости батареи..."
        )

        code_matches = []

        # Создаем множество артикулов новых товаров для быстрой проверки
        new_articles_set = set()
        if new_items_list:
            new_articles_set = {item["article"] for item in new_items_list}
            self.log_info(
                f"🔍 Ищем совпадения только среди {len(new_articles_set)} новых товаров"
            )

        # Извлекаем коды из наименований поставщика (только новые товары)
        supplier_codes = {}
        for idx, row in supplier_df.iterrows():
            if "name" in row and pd.notna(row["name"]):
                # Проверяем, что товар является новым
                article_key = str(row.get(f"article_{supplier_config}", ""))
                if new_items_list and article_key not in new_articles_set:
                    continue  # Пропускаем товары, которые не являются новыми

                code = self.find_product_code_unified(row["name"])
                if code:
                    # Приводим цену к правильному типу данных
                    price_raw = (
                        row.get("price_usd", 0)
                        if supplier_config == "vitya"
                        else row.get("price_usd", 0)
                    )
                    try:
                        price_float = float(price_raw) if price_raw is not None else 0.0
                    except (ValueError, TypeError):
                        price_float = 0.0

                    supplier_color = self.safe_color_processing(row.get("color"))
                    supplier_capacity = self.find_battery_capacity(row["name"])

                    # Группируем по коду, но сохраняем все варианты с разными цветами и емкостями
                    if code not in supplier_codes:
                        supplier_codes[code] = []

                    supplier_codes[code].append(
                        {
                            "index": idx,
                            "name": row["name"],
                            "price": price_float,
                            "article": row.get(f"article_{supplier_config}", ""),
                            "color": supplier_color,
                            "capacity": supplier_capacity,
                        }
                    )

        # Извлекаем коды из наименований базы
        base_codes = {}
        for idx, row in base_df.iterrows():
            if "name" in row and pd.notna(row["name"]):
                code = self.find_product_code_unified(row["name"])
                if code:
                    # Приводим цены к правильному типу данных
                    price_raw = row.get("price", 0)
                    price_vitya_raw = row.get("price_vitya_usd", 0)
                    price_dimi_raw = row.get("price_dimi_usd", 0)
                    price_mila_raw = row.get("price_mila_usd", 0)

                    try:
                        price_float = float(price_raw) if price_raw is not None else 0.0
                        price_vitya_float = (
                            float(price_vitya_raw)
                            if price_vitya_raw is not None
                            else 0.0
                        )
                        price_dimi_float = (
                            float(price_dimi_raw) if price_dimi_raw is not None else 0.0
                        )
                        price_mila_float = (
                            float(price_mila_raw) if price_mila_raw is not None else 0.0
                        )
                    except (ValueError, TypeError):
                        price_float = 0.0
                        price_vitya_float = 0.0
                        price_dimi_float = 0.0
                        price_mila_float = 0.0

                    base_color = self.safe_color_processing(row.get("color"))
                    base_capacity = self.find_battery_capacity(row["name"])

                    # Группируем по коду, но сохраняем все варианты с разными цветами и емкостями
                    if code not in base_codes:
                        base_codes[code] = []

                    base_codes[code].append(
                        {
                            "index": idx,
                            "name": row["name"],
                            "price": price_float,
                            "article": row.get("article", ""),
                            "color": base_color,
                            "capacity": base_capacity,
                            "price_vitya_usd": price_vitya_float,
                            "price_dimi_usd": price_dimi_float,
                            "price_mila_usd": price_mila_float,
                        }
                    )

            # Также ищем коды в столбцах артикулов поставщиков
            for supplier in ["vitya", "dimi", "mila"]:
                article_col = f"article_{supplier}"
                if article_col in row and pd.notna(row[article_col]):
                    article_str = str(row[article_col])
                    code = self.find_product_code_unified(article_str)
                    if code:
                        # Приводим цену к правильному типу данных
                        price_raw = row.get("price", 0)
                        try:
                            price_float = (
                                float(price_raw) if price_raw is not None else 0.0
                            )
                        except (ValueError, TypeError):
                            price_float = 0.0

                        base_color = self.safe_color_processing(row.get("color"))
                        base_capacity = self.find_battery_capacity(row["name"])

                        if code not in base_codes:
                            base_codes[code] = []

                        base_codes[code].append(
                            {
                                "index": idx,
                                "name": row["name"],
                                "price": price_float,
                                "article": row.get("article", ""),
                                "color": base_color,
                                "capacity": base_capacity,
                                "price_vitya_usd": price_vitya_float,
                                "price_dimi_usd": price_dimi_float,
                                "price_mila_usd": price_mila_float,
                            }
                        )

        self.log_info(
            f"📋 Извлечено кодов: поставщик {len(supplier_codes)}, база {len(base_codes)}"
        )

        # Ищем совпадения с улучшенной логикой
        for code, supplier_variants in supplier_codes.items():
            if code in base_codes:
                base_variants = base_codes[code]

                self.log_info(
                    f"🔍 Проверяем код {code}: {len(supplier_variants)} вариантов поставщика, {len(base_variants)} вариантов базы"
                )

                # Для каждого варианта поставщика ищем подходящий вариант в базе
                for supplier_variant in supplier_variants:
                    supplier_color = supplier_variant["color"]
                    supplier_capacity = supplier_variant["capacity"]
                    best_match = None
                    best_color_match = False
                    best_capacity_match = False

                    # Сначала ищем точное совпадение по цвету И емкости
                    for base_variant in base_variants:
                        base_color = base_variant["color"]
                        base_capacity = base_variant["capacity"]

                        color_match = supplier_color == base_color
                        capacity_match = supplier_capacity == base_capacity

                        if color_match and capacity_match:
                            best_match = base_variant
                            best_color_match = True
                            best_capacity_match = True
                            self.log_info(
                                f"✅ Найдено точное совпадение по цвету и емкости: {supplier_color}, {supplier_capacity}mAh"
                            )
                            break

                    # Если точное совпадение не найдено, ищем совпадение только по цвету
                    if not best_match:
                        for base_variant in base_variants:
                            base_color = base_variant["color"]
                            if supplier_color == base_color:
                                best_match = base_variant
                                best_color_match = True
                                best_capacity_match = False
                                self.log_info(
                                    f"✅ Найдено совпадение по цвету (емкость отличается): {supplier_color}, {supplier_capacity}mAh vs {base_variant['capacity']}mAh"
                                )
                                break

                    # Если совпадение по цвету не найдено, ищем совпадение только по емкости
                    if not best_match:
                        for base_variant in base_variants:
                            base_capacity = base_variant["capacity"]
                            if supplier_capacity == base_capacity:
                                best_match = base_variant
                                best_color_match = False
                                best_capacity_match = True
                                self.log_info(
                                    f"✅ Найдено совпадение по емкости (цвет отличается): {supplier_capacity}mAh, {supplier_color} vs {base_variant['color']}"
                                )
                                break

                    # Если точное совпадение не найдено, берем первый вариант из базы
                    if not best_match and base_variants:
                        best_match = base_variants[0]
                        best_color_match = False
                        best_capacity_match = False
                        self.log_info(
                            f"⚠️ Точное совпадение не найдено, берем первый вариант: цвет {supplier_color} vs {best_match['color']}, емкость {supplier_capacity}mAh vs {best_match['capacity']}mAh"
                        )

                    if best_match:
                        # Получаем цену из конфигурации
                        base_price = self.get_base_price_from_config(best_match)

                        # Проверяем, действительно ли цены отличаются
                        price_diff = abs(supplier_variant["price"] - base_price)
                        prices_equal = price_diff < 0.001

                        self.log_info(
                            f"🔍 Сравнение {code}: supplier={supplier_variant['price']} ({type(supplier_variant['price'])}), base={base_price} ({type(base_price)}), diff={price_diff:.6f}, equal={prices_equal}"
                        )

                        match_info = {
                            "code": code,
                            "supplier_name": supplier_variant["name"],
                            "base_name": best_match["name"],
                            "supplier_price": supplier_variant["price"],
                            "base_price": base_price,
                            "supplier_article": supplier_variant["article"],
                            "base_article": best_match["article"],
                            "supplier_color": supplier_variant["color"],
                            "base_color": best_match["color"],
                            "supplier_capacity": supplier_variant["capacity"],
                            "base_capacity": best_match["capacity"],
                            "base_index": best_match["index"],
                            "match_type": "product_code",
                            "color_match": best_color_match,
                            "capacity_match": best_capacity_match,
                            "price_change_percent": 0,
                        }

                        if base_price > 0:
                            match_info["price_change_percent"] = (
                                (supplier_variant["price"] - base_price)
                                / base_price
                                * 100
                            )

                        code_matches.append(match_info)

                        # Логируем созданное совпадение
                        self.log_info(
                            f"🔍 Создано совпадение {code}: supplier={supplier_variant['price']} ({type(supplier_variant['price'])}), base={base_price} ({type(base_price)}), color_match={best_color_match}, capacity_match={best_capacity_match}, change={match_info['price_change_percent']:.1f}%"
                        )

        self.log_info(f"✅ Найдено совпадений по кодам: {len(code_matches)}")
        return code_matches

    def compare_by_bracket_codes_advanced(
        self, supplier_df, base_df, supplier_config, new_items_list=None
    ):
        """
        Улучшенный поиск совпадений по кодам в скобках с учетом цветов и емкости батареи

        Логика:
        1. Ищем совпадения по кодам в скобках среди новых товаров
        2. Сравниваем цвета
        3. Сравниваем емкость батареи (mAh)
        4. Если цвета или емкость не совпадают, ищем тот же код с другими параметрами
        5. Проверяем все варианты кодов
        """

        self.log_info(
            "🔍 Улучшенный поиск совпадений по кодам в скобках с учетом цветов и емкости батареи..."
        )

        bracket_matches = []

        # Создаем множество артикулов новых товаров для быстрой проверки
        new_articles_set = set()
        if new_items_list:
            new_articles_set = {item["article"] for item in new_items_list}
            self.log_info(
                f"🔍 Ищем совпадения только среди {len(new_articles_set)} новых товаров"
            )

        # Извлекаем коды в скобках из наименований поставщика (только новые товары)
        supplier_bracket_codes = {}
        for idx, row in supplier_df.iterrows():
            if "name" in row and pd.notna(row["name"]):
                # Проверяем, что товар является новым
                article_key = str(row.get(f"article_{supplier_config}", ""))
                if new_items_list and article_key not in new_articles_set:
                    continue  # Пропускаем товары, которые не являются новыми

                code = self.find_product_code_in_brackets(row["name"])
                if code:
                    # Приводим цену к правильному типу данных
                    price_raw = (
                        row.get("price_usd", 0)
                        if supplier_config == "vitya"
                        else row.get("price_usd", 0)
                    )
                    try:
                        price_float = float(price_raw) if price_raw is not None else 0.0
                    except (ValueError, TypeError):
                        price_float = 0.0

                    supplier_color = self.safe_color_processing(row.get("color"))
                    supplier_capacity = self.find_battery_capacity(row["name"])

                    # Группируем по коду, но сохраняем все варианты с разными цветами и емкостями
                    if code not in supplier_bracket_codes:
                        supplier_bracket_codes[code] = []

                    supplier_bracket_codes[code].append(
                        {
                            "index": idx,
                            "name": row["name"],
                            "price": price_float,
                            "article": row.get(f"article_{supplier_config}", ""),
                            "color": supplier_color,
                            "capacity": supplier_capacity,
                        }
                    )

        # Извлекаем коды в скобках из наименований базы
        base_bracket_codes = {}
        for idx, row in base_df.iterrows():
            if "name" in row and pd.notna(row["name"]):
                code = self.find_product_code_in_brackets(row["name"])
                if code:
                    # Приводим цену к правильному типу данных
                    price_raw = row.get("price", 0)
                    try:
                        price_float = float(price_raw) if price_raw is not None else 0.0
                    except (ValueError, TypeError):
                        price_float = 0.0

                    base_color = self.safe_color_processing(row.get("color"))
                    base_capacity = self.find_battery_capacity(row["name"])

                    # Группируем по коду, но сохраняем все варианты с разными цветами и емкостями
                    if code not in base_bracket_codes:
                        base_bracket_codes[code] = []

                    base_bracket_codes[code].append(
                        {
                            "index": idx,
                            "name": row["name"],
                            "price": price_float,
                            "article": row.get("article", ""),
                            "matched_in": "name",
                            "color": base_color,
                            "capacity": base_capacity,
                        }
                    )

            # Также ищем коды в скобках в столбцах артикулов поставщиков
            for supplier in ["vitya", "dimi", "mila"]:
                article_col = f"article_{supplier}"
                if article_col in row and pd.notna(row[article_col]):
                    article_str = str(row[article_col])
                    code = self.find_product_code_in_brackets(article_str)
                    if code:
                        # Приводим цену к правильному типу данных
                        price_raw = row.get("price", 0)
                        try:
                            price_float = (
                                float(price_raw) if price_raw is not None else 0.0
                            )
                        except (ValueError, TypeError):
                            price_float = 0.0

                        base_color = self.safe_color_processing(row.get("color"))
                        base_capacity = self.find_battery_capacity(row["name"])

                        if code not in base_bracket_codes:
                            base_bracket_codes[code] = []

                        base_bracket_codes[code].append(
                            {
                                "index": idx,
                                "name": row["name"],
                                "price": price_float,
                                "article": row.get("article", ""),
                                "matched_in": article_col,
                                "color": base_color,
                                "capacity": base_capacity,
                            }
                        )

        self.log_info(
            f"📋 Извлечено кодов в скобках: поставщик {len(supplier_bracket_codes)}, база {len(base_bracket_codes)}"
        )

        # Ищем совпадения с улучшенной логикой
        for code, supplier_variants in supplier_bracket_codes.items():
            if code in base_bracket_codes:
                base_variants = base_bracket_codes[code]

                self.log_info(
                    f"🔍 Проверяем код в скобках {code}: {len(supplier_variants)} вариантов поставщика, {len(base_variants)} вариантов базы"
                )

                # Для каждого варианта поставщика ищем подходящий вариант в базе
                for supplier_variant in supplier_variants:
                    supplier_color = supplier_variant["color"]
                    supplier_capacity = supplier_variant["capacity"]
                    best_match = None
                    best_color_match = False
                    best_capacity_match = False

                    # Сначала ищем точное совпадение по цвету И емкости
                    for base_variant in base_variants:
                        base_color = base_variant["color"]
                        base_capacity = base_variant["capacity"]

                        color_match = supplier_color == base_color
                        capacity_match = supplier_capacity == base_capacity

                        if color_match and capacity_match:
                            best_match = base_variant
                            best_color_match = True
                            best_capacity_match = True
                            self.log_info(
                                f"✅ Найдено точное совпадение по цвету и емкости: {supplier_color}, {supplier_capacity}mAh"
                            )
                            break

                    # Если точное совпадение не найдено, ищем совпадение только по цвету
                    if not best_match:
                        for base_variant in base_variants:
                            base_color = base_variant["color"]
                            if supplier_color == base_color:
                                best_match = base_variant
                                best_color_match = True
                                best_capacity_match = False
                                self.log_info(
                                    f"✅ Найдено совпадение по цвету (емкость отличается): {supplier_color}, {supplier_capacity}mAh vs {base_variant['capacity']}mAh"
                                )
                                break

                    # Если совпадение по цвету не найдено, ищем совпадение только по емкости
                    if not best_match:
                        for base_variant in base_variants:
                            base_capacity = base_variant["capacity"]
                            if supplier_capacity == base_capacity:
                                best_match = base_variant
                                best_color_match = False
                                best_capacity_match = True
                                self.log_info(
                                    f"✅ Найдено совпадение по емкости (цвет отличается): {supplier_capacity}mAh, {supplier_color} vs {base_variant['color']}"
                                )
                                break

                    # Если точное совпадение не найдено, берем первый вариант из базы
                    if not best_match and base_variants:
                        best_match = base_variants[0]
                        best_color_match = False
                        best_capacity_match = False
                        self.log_info(
                            f"⚠️ Точное совпадение не найдено, берем первый вариант: цвет {supplier_color} vs {best_match['color']}, емкость {supplier_capacity}mAh vs {best_match['capacity']}mAh"
                        )

                    if best_match:
                        # Получаем цену из конфигурации
                        base_price = self.get_base_price_from_config(best_match)

                        # Проверяем, действительно ли цены отличаются
                        price_diff = abs(supplier_variant["price"] - base_price)
                        prices_equal = price_diff < 0.001

                        self.log_info(
                            f"🔍 Сравнение {code}: supplier={supplier_variant['price']} ({type(supplier_variant['price'])}), base={base_price} ({type(base_price)}), diff={price_diff:.6f}, equal={prices_equal}"
                        )

                        match_info = {
                            "code": code,
                            "supplier_name": supplier_variant["name"],
                            "base_name": best_match["name"],
                            "supplier_price": supplier_variant["price"],
                            "base_price": base_price,
                            "supplier_article": supplier_variant["article"],
                            "base_article": best_match["article"],
                            "supplier_color": supplier_variant["color"],
                            "base_color": best_match["color"],
                            "supplier_capacity": supplier_variant["capacity"],
                            "base_capacity": best_match["capacity"],
                            "matched_in": best_match["matched_in"],
                            "base_index": best_match["index"],
                            "match_type": "bracket_code",
                            "color_match": best_color_match,
                            "capacity_match": best_capacity_match,
                            "price_change_percent": 0,
                        }

                        if base_price > 0:
                            match_info["price_change_percent"] = (
                                (supplier_variant["price"] - base_price)
                                / base_price
                                * 100
                            )

                        bracket_matches.append(match_info)

                        # Логируем созданное совпадение
                        self.log_info(
                            f"🔍 Создано совпадение {code}: supplier={supplier_variant['price']} ({type(supplier_variant['price'])}), base={base_price} ({type(base_price)}), color_match={best_color_match}, capacity_match={best_capacity_match}, change={match_info['price_change_percent']:.1f}%"
                        )

        self.log_info(
            f"✅ Найдено совпадений по кодам в скобках: {len(bracket_matches)}"
        )
        return bracket_matches

    def compare_by_fuzzy_string_matching(
        self, fuzzy_candidates, base_df, supplier_config
    ):
        """
        Поиск совпадений по нечеткому сопоставлению строк наименований

        Эта функция ищет схожие названия товаров когда не удается найти совпадения
        по артикулам и кодам.

        Параметры:
        - fuzzy_candidates: список кандидатов для нечеткого поиска (товары без совпадений)
        - base_df: база данных
        - supplier_config: конфигурация поставщика

        Возвращает:
        - список найденных совпадений
        """

        self.log_info("🔍 Нечеткий поиск по совпадению строк...")
        self.log_info(
            f"📊 Кандидатов для поиска: {len(fuzzy_candidates) if isinstance(fuzzy_candidates, list) else len(fuzzy_candidates) if isinstance(fuzzy_candidates, pd.DataFrame) else 'N/A'}"
        )

        import difflib

        fuzzy_matches = []

        # Проверяем, что fuzzy_candidates не пустой (может быть список или DataFrame)
        if (isinstance(fuzzy_candidates, list) and len(fuzzy_candidates) == 0) or (
            isinstance(fuzzy_candidates, pd.DataFrame) and fuzzy_candidates.empty
        ):
            self.log_info("ℹ️ Нет кандидатов для нечеткого поиска")
            return fuzzy_matches

        # Если это DataFrame, конвертируем в список словарей
        if isinstance(fuzzy_candidates, pd.DataFrame):
            fuzzy_candidates = fuzzy_candidates.to_dict("records")
            self.log_info(
                f"🔄 DataFrame конвертирован в список из {len(fuzzy_candidates)} элементов"
            )

        # Проверяем, что base_df не пустой и является DataFrame
        if base_df is None or base_df.empty:
            self.log_error("❌ База данных пуста или не загружена")
            return fuzzy_matches

        # Определяем названия колонок для названий товаров
        base_name_col = self._get_base_name_column(base_df)

        if not base_name_col:
            self.log_error(
                "❌ Не удалось определить колонку с названиями товаров в базе"
            )
            return fuzzy_matches

        # Получаем названия товаров из базы
        base_names = []
        for idx, row in base_df.iterrows():
            # Проверяем, что колонка существует и значение не пустое
            if base_name_col in row.index and pd.notna(row[base_name_col]):
                base_names.append(
                    {
                        "index": idx,
                        "name": str(row[base_name_col]).strip(),
                        "price": row.get("price", 0),
                        "article": row.get("article", ""),
                        "color": self.safe_color_processing(row.get("color")),
                    }
                )

        # Порог схожести (0.3 = 30%)
        similarity_threshold = TRSH

        # Ищем совпадения для каждого кандидата
        for candidate in fuzzy_candidates:
            best_match = None
            best_ratio = 0

            # Получаем название товара кандидата
            supplier_name_col = self._get_supplier_name_column(
                pd.DataFrame([candidate])
            )
            if not supplier_name_col or supplier_name_col not in candidate:
                continue

            candidate_name = str(candidate[supplier_name_col]).strip()

            for base_item in base_names:
                # Вычисляем схожесть названий
                ratio = difflib.SequenceMatcher(
                    None, candidate_name.lower(), base_item["name"].lower()
                ).ratio()

                # Если схожесть выше порога и лучше предыдущего
                if ratio >= similarity_threshold and ratio > best_ratio:
                    best_ratio = ratio
                    best_match = base_item

            # Если нашли хорошее совпадение
            if best_match:
                match_info = {
                    "supplier_index": candidate.get("index", 0),
                    "supplier_name": candidate_name,
                    "supplier_price": candidate.get("price_usd", 0),
                    "supplier_article": candidate.get("article", ""),
                    "supplier_color": self.safe_color_processing(
                        candidate.get("color")
                    ),
                    "base_index": best_match["index"],
                    "base_name": best_match["name"],
                    "base_price": best_match["price"],
                    "base_article": best_match["article"],
                    "base_color": best_match["color"],
                    "similarity_ratio": best_ratio,
                    "match_type": "fuzzy_string",
                    "matched_in": "name",
                }
                fuzzy_matches.append(match_info)

        self.log_info(f"🔍 Найдено {len(fuzzy_matches)} совпадений по нечеткому поиску")
        return fuzzy_matches

    def auto_select_config(self, file_path):
        """Автоматический выбор и установка конфига"""

        detected_config = self.detect_config_by_filename(file_path)

        # Проверяем, есть ли такой конфиг в списке доступных
        available_configs = self.config_combo["values"]

        if detected_config in available_configs:
            self.config_combo.set(detected_config)
            self.log_info(f"🎯 Конфиг автоматически изменен на: {detected_config}")
            return detected_config
        else:
            self.log_info(f"⚠️ Конфиг {detected_config} не найден, оставляем текущий")
            return self.config_var.get()

    def clear_info(self):
        """Очистка области информации и сброс состояния"""
        self.info_text.delete(1.0, tk.END)

        # Сбрасываем только результаты сравнения, НЕ данные поставщика
        self.comparison_result = None

        # Сбрасываем состояние обновления цен при очистке интерфейса
        if hasattr(self, "price_updated"):
            self.price_updated = False
            self.log_info("🔄 Состояние обновления цен сброшено")

        # Сбрасываем состояние добавления товаров при очистке интерфейса
        self.articles_added = False
        self.log_info("🔄 Состояние добавления товаров сброшено")

        # Обновляем состояние кнопок
        self.update_buttons_state()

        # Сбрасываем статус
        self.set_status("Готов к работе", "info")

        # Обновляем информацию о файлах в статус-баре
        self.update_files_info()

        self.log_info("🧹 Интерфейс очищен, результаты сравнения сброшены")

    def update_buttons_state(self, log_changes=True, articles_added_this_run=False):
        """Обновление состояния кнопок в зависимости от загруженных данных"""
        # Кнопки, которые зависят от загруженного файла поставщика
        file_loaded = self.current_df is not None
        file_state = "normal" if file_loaded else "disabled"

        self.show_data_button.config(state=file_state)
        self.save_data_button.config(state=file_state)
        self.compare_button.config(state=file_state)

        # Кнопки, которые зависят от выполненного сравнения
        comparison_done = self.comparison_result is not None
        comparison_state = "normal" if comparison_done else "disabled"

        self.report_button.config(state=comparison_state)

        # Кнопка "Обновить цены" активна если загружен прайс И (цены еще не обновлены ИЛИ были добавлены новые товары)
        update_prices_state = (
            "normal"
            if (file_loaded and (not self.price_updated or self.articles_added))
            else "disabled"
        )
        self.update_prices_button.config(state=update_prices_state)

        # Кнопка "Добавить новый товар в базу" активна только если есть новые товары
        has_new_items = False
        new_items_count = 0
        if self.comparison_result is not None:
            new_items = self.comparison_result.get("new_items", [])
            new_items_count = len(new_items)
            has_new_items = new_items_count > 0

        # Кнопка "Добавить в базу" активна только если есть новые товары И товары еще не были добавлены
        add_to_base_state = (
            "normal" if (has_new_items and not self.articles_added) else "disabled"
        )
        self.add_to_base_button.config(state=add_to_base_state)

        # Логирование изменений (опционально)
        if log_changes:
            if file_loaded:
                self.log_info("✅ Файл загружен - основные кнопки активны")
            if comparison_done:
                self.log_info("✅ Сравнение выполнено - кнопки отчетов активны")
            if file_loaded and not self.price_updated and not self.articles_added:
                self.log_info("✅ Прайс загружен - кнопка обновления цен активна")
            elif file_loaded and self.price_updated and not self.articles_added:
                self.log_info(
                    "🔒 Прайс загружен, но цены уже обновлены - кнопка обновления цен неактивна"
                )
            elif file_loaded and self.articles_added:
                self.log_info(
                    "🔄 Прайс загружен, новые товары добавлены - кнопка обновления цен активна"
                )
            if has_new_items and not self.articles_added:
                self.log_info(
                    f"📥 Обнаружено новых товаров: {new_items_count} - кнопка добавления активна"
                )
            elif has_new_items and self.articles_added:
                self.log_info(
                    f"🔒 Обнаружено новых товаров: {new_items_count}, но товары уже добавлены - кнопка добавления неактивна"
                )
            elif comparison_done and not has_new_items:
                self.log_info(
                    "ℹ️ Новых товаров не найдено - кнопка добавления неактивна"
                )
            if not file_loaded and not comparison_done:
                self.log_info("⚪ Данные отсутствуют - кнопки деактивированы")

        # Обновляем информацию о файлах в статус-баре
        self.update_files_info()

    def log_info(self, message):
        """Логирование информации"""
        # Логируем в консоль и файл
        self.logger.info(message)

        # Также выводим в GUI (если доступен)
        if hasattr(self, "info_text") and self.info_text is not None:
            timestamp = datetime.now().strftime("%H:%M:%S")
            log_message = f"[{timestamp}] {message}\n"
            self.info_text.insert(tk.END, log_message)
            self.info_text.see(tk.END)

    def log_error(self, message):
        """Логирование ошибок"""
        # Логируем в консоль и файл
        self.logger.error(f"❌ ОШИБКА: {message}")

        # Также выводим в GUI (если доступен)
        if hasattr(self, "info_text") and self.info_text is not None:
            timestamp = datetime.now().strftime("%H:%M:%S")
            log_message = f"[{timestamp}] ❌ ОШИБКА: {message}\n"
            self.info_text.insert(tk.END, log_message)
            self.info_text.see(tk.END)

    def save_report(self):
        """Сохранение отчета о сравнении в Excel"""
        self.log_info("🔘 Нажата кнопка 'Сохранить отчет'")

        if self.comparison_result is None:
            self.log_info("❌ Результат сравнения отсутствует")
            messagebox.showwarning(
                "Предупреждение", "Сначала выполните сравнение с базой"
            )
            return

        self.log_info("✅ Результат сравнения найден, открываем диалог сохранения...")

        try:
            from tkinter import filedialog
            import os

            # Создаем папку data/output если её нет
            output_dir = "data/output"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                self.log_info(f"📁 Создана папка: {output_dir}")

            # Создаем имя файла с временной меткой
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"summary_report_{timestamp}.xlsx"
            self.log_info(f"📁 Предлагаемое имя файла: {default_filename}")

            file_path = filedialog.asksaveasfilename(
                title="Сохранить отчет о сравнении",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
                initialdir=output_dir,
            )

            self.log_info(
                f"📁 Выбранный путь: {file_path if file_path else 'Отменено пользователем'}"
            )
        except Exception as e:
            self.log_error(f"Ошибка при открытии диалога: {e}")
            return

        if file_path:
            try:
                self.log_info("💾 Начинаем сохранение отчета...")
                self.set_status("Сохранение отчета...", "save")
                self.root.update()

                # Создаем сводную таблицу
                self.log_info("📊 Создаем сводную таблицу...")
                summary_data = [
                    {
                        "Поставщик": self.current_config.upper(),
                        "Товаров": self.comparison_result["supplier_total"],
                        "Совпадений": len(self.comparison_result["matches"]),
                        "Процент совпадений": f"{self.comparison_result['match_rate']:.1f}%",
                        "Изменений цен": len(self.comparison_result["price_changes"]),
                        "Новых товаров": len(self.comparison_result["new_items"]),
                        "Совпадений по кодам": len(
                            self.comparison_result.get("code_matches", [])
                        ),
                    }
                ]
                self.log_info(f"✅ Сводная таблица создана: {summary_data[0]}")

                # Сохраняем в Excel с несколькими листами
                self.log_info("📝 Создаем Excel файл...")
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    # Лист с общей сводкой
                    self.log_info("📄 Создаем лист 'Сводка'...")
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name="Сводка", index=False)

                    # Настраиваем ширину столбцов для Сводки
                    worksheet = writer.sheets["Сводка"]
                    worksheet.column_dimensions["A"].width = 20  # Поставщик
                    worksheet.column_dimensions["B"].width = 12  # Товаров
                    worksheet.column_dimensions["C"].width = 15  # Совпадений
                    worksheet.column_dimensions["D"].width = 18  # Процент совпадений
                    worksheet.column_dimensions["E"].width = 15  # Изменений цен
                    worksheet.column_dimensions["F"].width = 15  # Новых товаров
                    worksheet.column_dimensions["G"].width = 20  # Совпадений по кодам

                    # Лист с совпадениями
                    if self.comparison_result["matches"]:
                        self.log_info(
                            f"📄 Создаем лист 'Совпадения' ({len(self.comparison_result['matches'])} записей)..."
                        )

                        # Добавляем цвета и артикулы в совпадения
                        matches_with_details = []
                        for match in self.comparison_result["matches"]:
                            match_with_details = match.copy()
                            # Безопасно получаем article, code или supplier_article
                            article = match.get("article", "")
                            if not article:
                                article = match.get("code", "")
                            if not article:
                                article = match.get("supplier_article", "")
                            if not article:
                                article = "N/A"  # Fallback

                            # Добавляем данные из базы
                            if (
                                article
                                and article != "N/A"
                                and self.base_df is not None
                            ):
                                # Для совпадений по артикулам ищем по артикулу
                                if "base_index" in match:
                                    # Для совпадений по кодам используем base_index
                                    base_index = match.get("base_index")
                                    if (
                                        base_index is not None
                                        and base_index in self.base_df.index
                                    ):
                                        base_row = self.base_df.loc[base_index]
                                        match_with_details["base_article"] = (
                                            base_row.get("article", "")
                                        )
                                        match_with_details["base_color"] = (
                                            self.safe_color_processing(
                                                base_row.get("color", "")
                                            )
                                        )
                                        match_with_details["base_price"] = (
                                            self.get_base_price_from_config(base_row)
                                        )
                                    else:
                                        match_with_details["base_article"] = article
                                        match_with_details["base_color"] = ""
                                        match_with_details["base_price"] = 0
                                else:
                                    # Для обычных совпадений по артикулам
                                    try:
                                        # Получаем правильное название столбца артикула для базы
                                        base_article_col = (
                                            self.get_base_article_column()
                                        )
                                        # Проверяем, что столбец существует в базе
                                        if base_article_col in self.base_df.columns:
                                            base_matches = self.base_df[
                                                self.base_df[base_article_col]
                                                == article
                                            ]
                                            if len(base_matches) > 0:
                                                base_row = base_matches.iloc[0]
                                                match_with_details["base_article"] = (
                                                    article
                                                )
                                                match_with_details["base_color"] = (
                                                    self.safe_color_processing(
                                                        base_row.get("color", "")
                                                    )
                                                )
                                                match_with_details["base_price"] = (
                                                    self.get_base_price_from_config(
                                                        base_row
                                                    )
                                                )
                                            else:
                                                match_with_details["base_article"] = (
                                                    article
                                                )
                                                match_with_details["base_color"] = ""
                                                match_with_details["base_price"] = 0
                                        else:
                                            match_with_details["base_article"] = article
                                            match_with_details["base_color"] = ""
                                            match_with_details["base_price"] = 0
                                    except Exception as e:
                                        # Если произошла ошибка при поиске, используем fallback
                                        match_with_details["base_article"] = article
                                        match_with_details["base_color"] = ""
                                        match_with_details["base_price"] = 0
                            else:
                                match_with_details["base_article"] = article
                                match_with_details["base_color"] = ""
                                match_with_details["base_price"] = 0

                            # Добавляем данные из поставщика
                            if (
                                article
                                and article != "N/A"
                                and self.current_df is not None
                            ):
                                # Для совпадений по кодам используем supplier_article из match
                                if "supplier_article" in match:
                                    supplier_article = match.get("supplier_article", "")
                                    if supplier_article:
                                        try:
                                            # Получаем правильное название столбца артикула для поставщика
                                            supplier_article_col = (
                                                self.get_supplier_article_column()
                                            )
                                            # Проверяем, что столбец существует в поставщике
                                            if (
                                                supplier_article_col
                                                in self.current_df.columns
                                            ):
                                                supplier_matches = self.current_df[
                                                    self.current_df[
                                                        supplier_article_col
                                                    ]
                                                    == supplier_article
                                                ]
                                                if len(supplier_matches) > 0:
                                                    supplier_row = (
                                                        supplier_matches.iloc[0]
                                                    )
                                                    match_with_details[
                                                        "supplier_article"
                                                    ] = supplier_article
                                                    match_with_details[
                                                        "supplier_color"
                                                    ] = self.safe_color_processing(
                                                        supplier_row.get("color", "")
                                                    )
                                                else:
                                                    match_with_details[
                                                        "supplier_article"
                                                    ] = supplier_article
                                                    match_with_details[
                                                        "supplier_color"
                                                    ] = ""
                                            else:
                                                match_with_details[
                                                    "supplier_article"
                                                ] = supplier_article
                                                match_with_details["supplier_color"] = (
                                                    ""
                                                )
                                        except Exception as e:
                                            # Если произошла ошибка при поиске, используем fallback
                                            match_with_details["supplier_article"] = (
                                                supplier_article
                                            )
                                            match_with_details["supplier_color"] = ""
                                    else:
                                        match_with_details["supplier_article"] = article
                                        match_with_details["supplier_color"] = ""
                                else:
                                    # Для обычных совпадений по артикулам
                                    try:
                                        # Получаем правильное название столбца артикула для поставщика
                                        supplier_article_col = (
                                            self.get_supplier_article_column()
                                        )
                                        # Проверяем, что столбец существует в поставщике
                                        if (
                                            supplier_article_col
                                            in self.current_df.columns
                                        ):
                                            supplier_matches = self.current_df[
                                                self.current_df[supplier_article_col]
                                                == article
                                            ]
                                            if len(supplier_matches) > 0:
                                                supplier_row = supplier_matches.iloc[0]
                                                match_with_details[
                                                    "supplier_article"
                                                ] = article
                                                match_with_details["supplier_color"] = (
                                                    self.safe_color_processing(
                                                        supplier_row.get("color", "")
                                                    )
                                                )
                                            else:
                                                match_with_details[
                                                    "supplier_article"
                                                ] = article
                                                match_with_details["supplier_color"] = (
                                                    ""
                                                )
                                        else:
                                            match_with_details["supplier_article"] = (
                                                article
                                            )
                                            match_with_details["supplier_color"] = ""
                                    except Exception as e:
                                        # Если произошла ошибка при поиске, используем fallback
                                        match_with_details["supplier_article"] = article
                                        match_with_details["supplier_color"] = ""
                            else:
                                match_with_details["supplier_article"] = article
                                match_with_details["supplier_color"] = ""

                            matches_with_details.append(match_with_details)

                        matches_df = pd.DataFrame(matches_with_details)
                        matches_df.to_excel(
                            writer, sheet_name="Совпадения", index=False
                        )

                        # Настраиваем ширину столбцов для Совпадений
                        worksheet = writer.sheets["Совпадения"]
                        # Ищем столбец с name и устанавливаем ширину 110
                        if "name" in matches_df.columns:
                            name_col_index = matches_df.columns.get_loc("name")
                            name_col_letter = chr(
                                65 + name_col_index
                            )  # A=65, B=66, C=67...
                            worksheet.column_dimensions[name_col_letter].width = 110

                        # Устанавливаем стандартную ширину для остальных столбцов
                        for i, col in enumerate(matches_df.columns):
                            col_letter = chr(65 + i)
                            if col != "name":  # name уже настроен выше
                                if "article" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif "price" in col.lower() or "diff" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif "color" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 20
                                else:
                                    worksheet.column_dimensions[col_letter].width = 18

                    # Лист с изменениями цен
                    if self.comparison_result["price_changes"]:
                        self.log_info(
                            f"📄 Создаем лист 'Изменения цен' ({len(self.comparison_result['price_changes'])} записей)..."
                        )

                        # Добавляем цвета в данные об изменениях цен
                        price_changes_with_colors = []
                        for change in self.comparison_result["price_changes"]:
                            change_with_colors = change.copy()
                            # Безопасно получаем article
                            article = change.get("article", "")
                            if not article:
                                article = change.get("code", "")
                            if not article:
                                article = "N/A"

                            # Добавляем цвета из базы данных
                            if (
                                article
                                and article != "N/A"
                                and self.base_df is not None
                            ):
                                try:
                                    # Получаем правильное название столбца артикула для базы
                                    base_article_col = self.get_base_article_column()
                                    # Проверяем, что столбец существует в базе
                                    if base_article_col in self.base_df.columns:
                                        base_matches = self.base_df[
                                            self.base_df[base_article_col] == article
                                        ]
                                        if len(base_matches) > 0:
                                            change_with_colors["base_color"] = (
                                                self.safe_color_processing(
                                                    base_matches.iloc[0].get(
                                                        "color", ""
                                                    )
                                                )
                                            )
                                        else:
                                            change_with_colors["base_color"] = ""
                                    else:
                                        change_with_colors["base_color"] = ""
                                except Exception as e:
                                    change_with_colors["base_color"] = ""
                            else:
                                change_with_colors["base_color"] = ""

                            # Добавляем цвета из данных поставщика
                            if (
                                article
                                and article != "N/A"
                                and self.current_df is not None
                            ):
                                try:
                                    # Получаем правильное название столбца артикула для поставщика
                                    supplier_article_col = (
                                        self.get_supplier_article_column()
                                    )
                                    # Проверяем, что столбец существует в поставщике
                                    if supplier_article_col in self.current_df.columns:
                                        supplier_matches = self.current_df[
                                            self.current_df[supplier_article_col]
                                            == article
                                        ]
                                        if len(supplier_matches) > 0:
                                            change_with_colors["supplier_color"] = (
                                                self.safe_color_processing(
                                                    supplier_matches.iloc[0].get(
                                                        "color", ""
                                                    )
                                                )
                                            )
                                        else:
                                            change_with_colors["supplier_color"] = ""
                                    else:
                                        change_with_colors["supplier_color"] = ""
                                except Exception as e:
                                    change_with_colors["supplier_color"] = ""
                            else:
                                change_with_colors["supplier_color"] = ""

                            price_changes_with_colors.append(change_with_colors)

                        price_changes_df = pd.DataFrame(price_changes_with_colors)
                        price_changes_df.to_excel(
                            writer, sheet_name="Изменения цен", index=False
                        )

                        # Настраиваем ширину столбцов для Изменений цен
                        worksheet = writer.sheets["Изменения цен"]
                        # Ищем столбец с name и устанавливаем ширину 110
                        if "name" in price_changes_df.columns:
                            name_col_index = price_changes_df.columns.get_loc("name")
                            name_col_letter = chr(65 + name_col_index)
                            worksheet.column_dimensions[name_col_letter].width = 110

                        # Устанавливаем стандартную ширину для остальных столбцов
                        for i, col in enumerate(price_changes_df.columns):
                            col_letter = chr(65 + i)
                            if col != "name":
                                if "article" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif (
                                    "price" in col.lower()
                                    or "diff" in col.lower()
                                    or "percent" in col.lower()
                                ):
                                    worksheet.column_dimensions[col_letter].width = 15
                                else:
                                    worksheet.column_dimensions[col_letter].width = 18

                    # Лист с новыми товарами
                    if self.comparison_result["new_items"]:
                        self.log_info(
                            f"📄 Создаем лист 'Новые товары' ({len(self.comparison_result['new_items'])} записей)..."
                        )

                        # Добавляем цвета и информацию о совпадениях в новые товары
                        new_items_with_colors = []
                        for item in self.comparison_result["new_items"]:
                            item_with_colors = item.copy()

                            # Если цвет уже есть в item, используем его
                            if "color" in item and item["color"]:
                                item_with_colors["color"] = item["color"]
                            else:
                                # Безопасно получаем article
                                article = item.get("article", "")
                                if not article:
                                    article = "N/A"

                                # Добавляем цвета из данных поставщика
                                if (
                                    article
                                    and article != "N/A"
                                    and self.current_df is not None
                                ):
                                    try:
                                        # Получаем правильное название столбца артикула для поставщика
                                        supplier_article_col = (
                                            self.get_supplier_article_column()
                                        )
                                        # Проверяем, что столбец существует в поставщике
                                        if (
                                            supplier_article_col
                                            in self.current_df.columns
                                        ):
                                            # Приводим артикул к строке для сравнения
                                            article_str = str(article).strip()
                                            supplier_matches = self.current_df[
                                                self.current_df[supplier_article_col]
                                                .astype(str)
                                                .str.strip()
                                                == article_str
                                            ]
                                            if len(supplier_matches) > 0:
                                                item_with_colors["color"] = (
                                                    self.safe_color_processing(
                                                        supplier_matches.iloc[0].get(
                                                            "color", ""
                                                        )
                                                    )
                                                )
                                            else:
                                                item_with_colors["color"] = ""
                                        else:
                                            item_with_colors["color"] = ""
                                    except Exception as e:
                                        item_with_colors["color"] = ""
                                else:
                                    item_with_colors["color"] = ""

                            # Добавляем информацию о возможном совпадении
                            if item.get("fuzzy_match_name"):
                                item_with_colors["Возможное_совпадение"] = item[
                                    "fuzzy_match_name"
                                ]
                                item_with_colors["Строка_в_базе"] = item[
                                    "fuzzy_match_row"
                                ]
                                item_with_colors["Цвет_совпадения"] = item[
                                    "fuzzy_match_color"
                                ]
                                item_with_colors["Цена_совпадения"] = item[
                                    "fuzzy_match_price"
                                ]
                                item_with_colors["Схожесть_%"] = (
                                    f"{item.get('fuzzy_match_similarity', 0) * 100:.1f}%"
                                )
                            else:
                                item_with_colors["Возможное_совпадение"] = "Не найдено"
                                item_with_colors["Строка_в_базе"] = ""
                                item_with_colors["Цвет_совпадения"] = ""
                                item_with_colors["Цена_совпадения"] = ""
                                item_with_colors["Схожесть_%"] = "0.0%"

                            new_items_with_colors.append(item_with_colors)

                        new_items_df = pd.DataFrame(new_items_with_colors)
                        new_items_df.to_excel(
                            writer, sheet_name="Новые товары", index=False
                        )

                        # Настраиваем ширину столбцов для Новых товаров
                        worksheet = writer.sheets["Новые товары"]
                        # Ищем столбец с name и устанавливаем ширину 110
                        if "name" in new_items_df.columns:
                            name_col_index = new_items_df.columns.get_loc("name")
                            name_col_letter = chr(65 + name_col_index)
                            worksheet.column_dimensions[name_col_letter].width = 110

                        # Устанавливаем стандартную ширину для остальных столбцов
                        for i, col in enumerate(new_items_df.columns):
                            col_letter = chr(65 + i)
                            if col != "name":
                                if "article" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif "price" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif "color" in col.lower() or "balance" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 20
                                elif "Возможное_совпадение" in col:
                                    worksheet.column_dimensions[col_letter].width = 60
                                elif "Схожесть_%" in col:
                                    worksheet.column_dimensions[col_letter].width = 12
                                elif (
                                    "Строка_в_базе" in col
                                    or "Цвет_совпадения" in col
                                    or "Цена_совпадения" in col
                                ):
                                    worksheet.column_dimensions[col_letter].width = 18
                                else:
                                    worksheet.column_dimensions[col_letter].width = 18

                    # Лист с совпадениями по кодам
                    if self.comparison_result.get("code_matches"):
                        self.log_info(
                            f"📄 Создаем лист 'Совпадения по кодам' ({len(self.comparison_result['code_matches'])} записей)..."
                        )

                        # Добавляем цвета и артикулы в совпадения по кодам
                        code_matches_with_details = []
                        for match in self.comparison_result["code_matches"]:
                            match_with_details = match.copy()
                            code = match.get("code", "")
                            base_index = match.get("base_index")

                            # Добавляем данные из базы
                            if (
                                base_index is not None
                                and self.base_df is not None
                                and base_index in self.base_df.index
                            ):
                                base_row = self.base_df.loc[base_index]
                                match_with_details["base_article"] = base_row.get(
                                    "article", ""
                                )
                                match_with_details["base_color"] = (
                                    self.safe_color_processing(
                                        base_row.get("color", "")
                                    )
                                )
                                match_with_details["base_price"] = (
                                    self.get_base_price_from_config(base_row)
                                )
                            else:
                                match_with_details["base_article"] = ""
                                match_with_details["base_color"] = ""
                                match_with_details["base_price"] = 0

                            # Добавляем данные из поставщика
                            supplier_article = match.get("supplier_article", "")
                            if supplier_article and self.current_df is not None:
                                try:
                                    # Получаем правильное название столбца артикула для поставщика
                                    supplier_article_col = (
                                        self.get_supplier_article_column()
                                    )
                                    # Проверяем, что столбец существует в поставщике
                                    if supplier_article_col in self.current_df.columns:
                                        supplier_matches = self.current_df[
                                            self.current_df[supplier_article_col]
                                            == supplier_article
                                        ]
                                    else:
                                        supplier_matches = pd.DataFrame()
                                except Exception as e:
                                    supplier_matches = pd.DataFrame()
                                if len(supplier_matches) > 0:
                                    supplier_row = supplier_matches.iloc[0]
                                    match_with_details["supplier_color"] = (
                                        self.safe_color_processing(
                                            supplier_row.get("color", "")
                                        )
                                    )
                                else:
                                    match_with_details["supplier_color"] = ""
                            else:
                                match_with_details["supplier_color"] = ""

                            code_matches_with_details.append(match_with_details)

                        code_matches_df = pd.DataFrame(code_matches_with_details)
                        code_matches_df.to_excel(
                            writer, sheet_name="Совпадения по кодам", index=False
                        )

                        # Настраиваем ширину столбцов для Совпадений по кодам
                        worksheet = writer.sheets["Совпадения по кодам"]
                        # Ищем столбцы с name и устанавливаем ширину 110
                        for col_name in ["name", "supplier_name", "base_name"]:
                            if col_name in code_matches_df.columns:
                                name_col_index = code_matches_df.columns.get_loc(
                                    col_name
                                )
                                name_col_letter = chr(65 + name_col_index)
                                worksheet.column_dimensions[name_col_letter].width = 110

                        # Устанавливаем стандартную ширину для остальных столбцов
                        for i, col in enumerate(code_matches_df.columns):
                            col_letter = chr(65 + i)
                            if col not in ["name", "supplier_name", "base_name"]:
                                if "article" in col.lower() or "code" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif "confidence" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                else:
                                    worksheet.column_dimensions[col_letter].width = 18

                    # Лист с совпадениями по кодам в скобках
                    if self.comparison_result.get("bracket_matches"):
                        self.log_info(
                            f"📄 Создаем лист 'Совпадения по кодам в скобках' ({len(self.comparison_result['bracket_matches'])} записей)..."
                        )

                        # Добавляем цвета и артикулы в совпадения по кодам в скобках
                        bracket_matches_with_details = []
                        for match in self.comparison_result["bracket_matches"]:
                            match_with_details = match.copy()
                            code = match.get("code", "")
                            base_index = match.get("base_index")

                            # Добавляем данные из базы
                            if (
                                base_index is not None
                                and self.base_df is not None
                                and base_index in self.base_df.index
                            ):
                                base_row = self.base_df.loc[base_index]
                                match_with_details["base_article"] = base_row.get(
                                    "article", ""
                                )
                                match_with_details["base_color"] = (
                                    self.safe_color_processing(
                                        base_row.get("color", "")
                                    )
                                )
                                match_with_details["base_price"] = (
                                    self.get_base_price_from_config(base_row)
                                )
                            else:
                                match_with_details["base_article"] = ""
                                match_with_details["base_color"] = ""
                                match_with_details["base_price"] = 0

                            # Добавляем данные из поставщика
                            supplier_article = match.get("supplier_article", "")
                            if supplier_article and self.current_df is not None:
                                try:
                                    # Получаем правильное название столбца артикула для поставщика
                                    supplier_article_col = (
                                        self.get_supplier_article_column()
                                    )
                                    # Проверяем, что столбец существует в поставщике
                                    if supplier_article_col in self.current_df.columns:
                                        supplier_matches = self.current_df[
                                            self.current_df[supplier_article_col]
                                            == supplier_article
                                        ]
                                        if len(supplier_matches) > 0:
                                            supplier_row = supplier_matches.iloc[0]
                                            match_with_details["supplier_color"] = (
                                                self.safe_color_processing(
                                                    supplier_row.get("color", "")
                                                )
                                            )
                                        else:
                                            match_with_details["supplier_color"] = ""
                                    else:
                                        match_with_details["supplier_color"] = ""
                                except Exception as e:
                                    match_with_details["supplier_color"] = ""
                            else:
                                match_with_details["supplier_color"] = ""

                            bracket_matches_with_details.append(match_with_details)

                        bracket_matches_df = pd.DataFrame(bracket_matches_with_details)
                        bracket_matches_df.to_excel(
                            writer,
                            sheet_name="Совпадения по кодам в скобках",
                            index=False,
                        )

                        # Настраиваем ширину столбцов для Совпадений по кодам в скобках
                        worksheet = writer.sheets["Совпадения по кодам в скобках"]
                        # Ищем столбцы с name и устанавливаем ширину 110
                        for col_name in ["name", "supplier_name", "base_name"]:
                            if col_name in bracket_matches_df.columns:
                                name_col_index = bracket_matches_df.columns.get_loc(
                                    col_name
                                )
                                name_col_letter = chr(65 + name_col_index)
                                worksheet.column_dimensions[name_col_letter].width = 110

                        # Устанавливаем стандартную ширину для остальных столбцов
                        for i, col in enumerate(bracket_matches_df.columns):
                            col_letter = chr(65 + i)
                            if col not in ["name", "supplier_name", "base_name"]:
                                if "article" in col.lower() or "code" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                elif "confidence" in col.lower():
                                    worksheet.column_dimensions[col_letter].width = 15
                                else:
                                    worksheet.column_dimensions[col_letter].width = 18

                    # Лист с предупреждениями (значительные изменения цен)
                    warnings_data = []

                    # Добавляем значительные изменения цен как предупреждения
                    for change in self.comparison_result.get("price_changes", []):
                        if (
                            abs(change.get("price_change_percent", 0))
                            > SIGNIFICANT_CHANGE_PERCENT
                        ):
                            warnings_data.append(
                                {
                                    "Тип предупреждения": "Значительное изменение цены",
                                    "Артикул": change.get("article", ""),
                                    "Наименование": change.get("name", ""),
                                    "Цена базы": change.get("base_price", 0),
                                    "Цена поставщика": change.get("supplier_price", 0),
                                    "Изменение %": f"{change.get('price_change_percent', 0):+.1f}%",
                                    "Разница": change.get("price_diff", 0),
                                    "Описание": f"Изменение цены превышает {SIGNIFICANT_CHANGE_PERCENT}%",
                                }
                            )

                    # Добавляем предупреждения о товарах без цены в базе
                    for match in self.comparison_result.get("matches", []):
                        if (
                            match.get("base_price", 0) <= 0
                            and match.get("supplier_price", 0) > 0
                        ):
                            warnings_data.append(
                                {
                                    "Тип предупреждения": "Отсутствует цена в базе",
                                    "Артикул": match.get("article", ""),
                                    "Наименование": match.get("name", ""),
                                    "Цена базы": match.get("base_price", 0),
                                    "Цена поставщика": match.get("supplier_price", 0),
                                    "Изменение %": "Новая цена",
                                    "Разница": match.get("supplier_price", 0),
                                    "Описание": "В базе нет цены, но есть у поставщика",
                                }
                            )

                    # Добавляем предупреждения о пропущенных кодах из changes_log
                    if hasattr(self, "changes_log") and self.changes_log:
                        for change in self.changes_log:
                            if change.get("type") == "article_skipped":
                                warnings_data.append(
                                    {
                                        "Тип предупреждения": "Пропущен код",
                                        "Артикул": change.get("code", ""),
                                        "Наименование": change.get("base_name", ""),
                                        "Цена базы": "N/A",
                                        "Цена поставщика": "N/A",
                                        "Изменение %": "N/A",
                                        "Разница": "N/A",
                                        "Описание": f"Код {change.get('code', '')}: {change.get('reason', '')} - строка {change.get('base_index', 0) + 2}, столбец {change.get('column', '')}",
                                    }
                                )

                    # Создаем лист Предупреждения если есть данные
                    if warnings_data:
                        self.log_info(
                            f"📄 Создаем лист 'Предупреждения' ({len(warnings_data)} записей)..."
                        )
                        warnings_df = pd.DataFrame(warnings_data)
                        warnings_df.to_excel(
                            writer, sheet_name="Предупреждения", index=False
                        )

                        # Настраиваем ширину столбцов для Предупреждений
                        worksheet = writer.sheets["Предупреждения"]
                        worksheet.column_dimensions["A"].width = (
                            25  # Тип предупреждения
                        )
                        worksheet.column_dimensions["B"].width = 15  # Артикул
                        worksheet.column_dimensions["C"].width = (
                            110  # Наименование (широкий)
                        )
                        worksheet.column_dimensions["D"].width = 15  # Цена базы
                        worksheet.column_dimensions["E"].width = 18  # Цена поставщика
                        worksheet.column_dimensions["F"].width = 15  # Изменение %
                        worksheet.column_dimensions["G"].width = 12  # Разница
                        worksheet.column_dimensions["H"].width = 40  # Описание
                    else:
                        self.log_info("ℹ️ Предупреждений для отчета не найдено")

                    # Лист с изменениями артикулов (если есть)
                    if hasattr(self, "changes_log") and self.changes_log:
                        self.log_info(
                            f"📄 Создаем лист 'Изменения артикулов' ({len(self.changes_log)} записей)..."
                        )

                        # Преобразуем данные об изменениях в удобный формат
                        changes_data = []
                        for change in self.changes_log:
                            if change["type"] == "article_added":
                                changes_data.append(
                                    {
                                        "Статус": "✅ ДОБАВЛЕН",
                                        "Тип совпадения": (
                                            "📦 Код в скобках"
                                            if change["match_type"] == "bracket"
                                            else "🔗 Общий код"
                                        ),
                                        "Код": change["code"],
                                        "Строка в базе": change["base_index"]
                                        + 2,  # +2 потому что база начинается с 1 + заголовок
                                        "Столбец": change["column"],
                                        "Товар в базе": (
                                            change["base_name"][:80] + "..."
                                            if len(change["base_name"]) > 80
                                            else change["base_name"]
                                        ),
                                        "Артикул в базе": change.get(
                                            "base_article", ""
                                        ),
                                        "Цвет в базе": self.safe_color_processing(
                                            change.get("base_color", "")
                                        ),
                                        "Товар поставщика": (
                                            change["supplier_name"][:80] + "..."
                                            if len(change["supplier_name"]) > 80
                                            else change["supplier_name"]
                                        ),
                                        "Артикул поставщика": change.get(
                                            "supplier_article", ""
                                        ),
                                        "Цвет поставщика": self.safe_color_processing(
                                            change.get("supplier_color", "")
                                        ),
                                        "Было значение": change["old_value"],
                                        "Стало значение": change["new_value"],
                                    }
                                )
                            elif change["type"] == "article_skipped":
                                changes_data.append(
                                    {
                                        "Статус": "⏭️ ПРОПУЩЕН",
                                        "Тип совпадения": (
                                            "📦 Код в скобках"
                                            if change["match_type"] == "bracket"
                                            else "🔗 Общий код"
                                        ),
                                        "Код": change["code"],
                                        "Строка в базе": change["base_index"] + 2,
                                        "Столбец": change["column"],
                                        "Товар в базе": (
                                            change["base_name"][:80] + "..."
                                            if len(change["base_name"]) > 80
                                            else change["base_name"]
                                        ),
                                        "Артикул в базе": change.get(
                                            "base_article", ""
                                        ),
                                        "Цвет в базе": self.safe_color_processing(
                                            change.get("base_color", "")
                                        ),
                                        "Товар поставщика": (
                                            change["supplier_name"][:80] + "..."
                                            if len(change["supplier_name"]) > 80
                                            else change["supplier_name"]
                                        ),
                                        "Артикул поставщика": change.get(
                                            "supplier_article", ""
                                        ),
                                        "Цвет поставщика": self.safe_color_processing(
                                            change.get("supplier_color", "")
                                        ),
                                        "Существующее значение": change[
                                            "existing_value"
                                        ],
                                        "Попытка записать": change["attempted_value"],
                                        "Причина пропуска": change["reason"],
                                    }
                                )

                        if changes_data:
                            changes_df = pd.DataFrame(changes_data)
                            changes_df.to_excel(
                                writer, sheet_name="Изменения артикулов", index=False
                            )

                            # Настраиваем ширину столбцов для Изменений артикулов
                            worksheet = writer.sheets["Изменения артикулов"]
                            worksheet.column_dimensions["A"].width = 15  # Статус
                            worksheet.column_dimensions["B"].width = (
                                18  # Тип совпадения
                            )
                            worksheet.column_dimensions["C"].width = 15  # Код
                            worksheet.column_dimensions["D"].width = 12  # Строка в базе
                            worksheet.column_dimensions["E"].width = 15  # Столбец
                            worksheet.column_dimensions["F"].width = 60  # Товар в базе
                            worksheet.column_dimensions["G"].width = (
                                15  # Артикул в базе
                            )
                            worksheet.column_dimensions["H"].width = 15  # Цвет в базе
                            worksheet.column_dimensions["I"].width = (
                                60  # Товар поставщика
                            )
                            worksheet.column_dimensions["J"].width = (
                                15  # Артикул поставщика
                            )
                            worksheet.column_dimensions["K"].width = (
                                15  # Цвет поставщика
                            )

                            # Для добавленных артикулов
                            if "Было значение" in changes_df.columns:
                                worksheet.column_dimensions["L"].width = (
                                    15  # Было значение
                                )
                                worksheet.column_dimensions["M"].width = (
                                    15  # Стало значение
                                )

                            # Для пропущенных артикулов
                            if "Существующее значение" in changes_df.columns:
                                worksheet.column_dimensions["L"].width = (
                                    20  # Существующее значение
                                )
                                worksheet.column_dimensions["M"].width = (
                                    20  # Попытка записать
                                )
                                worksheet.column_dimensions["N"].width = (
                                    50  # Причина пропуска
                                )

                            self.log_info(
                                f"✅ Создана вкладка 'Изменения артикулов' с {len(changes_data)} записями"
                            )
                    else:
                        self.log_info("ℹ️ Изменений артикулов для отчета не найдено")

                    # Лист с обновленными ценами (если есть)
                    if hasattr(self, "price_updates_log") and self.price_updates_log:
                        self.log_info(
                            f"📄 Создаем лист 'Обновленные цены' ({len(self.price_updates_log)} записей)..."
                        )

                        # Преобразуем данные об обновленных ценах в удобный формат
                        price_updates_data = []
                        for update in self.price_updates_log:
                            price_updates_data.append(
                                {
                                    "Артикул": update.get("article", ""),
                                    "Товар в базе": (
                                        str(update.get("base_name", ""))[:80] + "..."
                                        if len(str(update.get("base_name", ""))) > 80
                                        else str(update.get("base_name", ""))
                                    ),
                                    "Артикул в базе": str(
                                        update.get("base_article", "")
                                    ),
                                    "Цвет в базе": self.safe_color_processing(
                                        update.get("base_color", "")
                                    ),
                                    "Товар поставщика": (
                                        str(update.get("supplier_name", ""))[:80]
                                        + "..."
                                        if len(str(update.get("supplier_name", "")))
                                        > 80
                                        else str(update.get("supplier_name", ""))
                                    ),
                                    "Артикул поставщика": update.get(
                                        "supplier_article", ""
                                    ),
                                    "Цвет поставщика": self.safe_color_processing(
                                        update.get("supplier_color", "")
                                    ),
                                    "Старая цена": update.get("old_price", 0),
                                    "Новая цена": update.get("new_price", 0),
                                    "Изменение %": f"{update.get('price_change_percent', 0):+.1f}%",
                                    "Тип совпадения": update.get("match_type", ""),
                                }
                            )

                        if price_updates_data:
                            price_updates_df = pd.DataFrame(price_updates_data)
                            price_updates_df.to_excel(
                                writer, sheet_name="Обновленные цены", index=False
                            )

                            # Настраиваем ширину столбцов для Обновленных цен
                            worksheet = writer.sheets["Обновленные цены"]
                            worksheet.column_dimensions["A"].width = 15  # Артикул
                            worksheet.column_dimensions["B"].width = 60  # Товар в базе
                            worksheet.column_dimensions["C"].width = (
                                15  # Артикул в базе
                            )
                            worksheet.column_dimensions["D"].width = 15  # Цвет в базе
                            worksheet.column_dimensions["E"].width = (
                                60  # Товар поставщика
                            )
                            worksheet.column_dimensions["F"].width = (
                                15  # Артикул поставщика
                            )
                            worksheet.column_dimensions["G"].width = (
                                15  # Цвет поставщика
                            )
                            worksheet.column_dimensions["H"].width = 15  # Старая цена
                            worksheet.column_dimensions["I"].width = 15  # Новая цена
                            worksheet.column_dimensions["J"].width = 15  # Изменение %
                            worksheet.column_dimensions["K"].width = (
                                20  # Тип совпадения
                            )

                            self.log_info(
                                f"✅ Создана вкладка 'Обновленные цены' с {len(price_updates_data)} записями"
                            )
                    else:
                        self.log_info("ℹ️ Обновленных цен для отчета не найдено")

                self.log_info("✅ Excel файл создан успешно")

                self.log_info(f"📊 Отчет сохранен: {file_path}")
                self.log_info(f"   Листов создано: {len(summary_data)} + детализация")
                messagebox.showinfo("Успех", f"Отчет сохранен в {file_path}")
                self.set_status("Отчет сохранен", "success")

            except Exception as e:
                self.log_error(f"Ошибка сохранения отчета: {e}")
                messagebox.showerror("Ошибка", f"Не удалось сохранить отчет: {e}")
                self.set_status("Ошибка сохранения отчета", "error")
        else:
            self.log_info("ℹ️ Сохранение отчета отменено пользователем")

    def update_prices(self):
        """Обновление цен в базе данных с улучшенной индикацией"""
        try:
            self.start_progress("Обновляю цены", 5, "update")  # 5 шагов прогресса
            self.log_info("🔄 Начало обновления цен в базе данных...")

            # Очищаем область информации для вывода лога обновления
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(tk.END, "📊 ОБНОВЛЕНИЕ ЦЕН В БАЗЕ ДАННЫХ\n")
            self.info_text.insert(tk.END, "=" * 50 + "\n")
            self.root.update()

            # Шаг 1: Проверка данных
            self.update_progress(1, "Проверка данных")
            self.info_text.insert(tk.END, "🔍 Проверка входных данных...\n")
            self.root.update()

            if self.current_df is None:
                error_msg = "❌ Файл поставщика не загружен"
                self.info_text.insert(tk.END, error_msg + "\n")
                self.finish_progress(error_msg, auto_reset=False)
                messagebox.showwarning(
                    "Предупреждение", "Сначала загрузите файл поставщика"
                )
                return

            # Шаг 2: Загрузка базы данных (если нужно)
            self.update_progress(2, "Загрузка базы данных")
            self.info_text.insert(tk.END, "📂 Загрузка базы данных...\n")
            self.root.update()

            if self.base_df is None:
                self.info_text.insert(tk.END, "Автозагрузка базы данных...\n")
                data_dir = "data/input"
                result = load_largest_file(data_dir, "base")

                if result is None:
                    error_msg = "❌ Не удалось загрузить базу данных"
                    self.info_text.insert(tk.END, error_msg + "\n")
                    self.finish_progress(error_msg, auto_reset=False)
                    messagebox.showerror(
                        "Ошибка", "Не удалось загрузить базу данных из data/input"
                    )
                    return

                self.base_df, base_file_path = result
                self.base_file_name = os.path.basename(base_file_path)
                self.info_text.insert(
                    tk.END, f"✅ Загружена база: {self.base_file_name}\n"
                )

            # Шаг 3: Подготовка данных
            self.update_progress(3, "Подготовка данных")
            self.info_text.insert(tk.END, "\n🔧 Подготовка данных...\n")
            self.root.update()

            supplier_article_col = self.get_supplier_article_column()
            supplier_price_col = self.get_supplier_price_column()
            base_article_col = self.get_base_article_column()
            base_price_col = self.get_base_price_column()

            # Проверка столбцов
            missing_columns = []
            if supplier_article_col not in self.current_df.columns:
                missing_columns.append(f"'{supplier_article_col}' в файле поставщика")
            if supplier_price_col not in self.current_df.columns:
                missing_columns.append(f"'{supplier_price_col}' в файле поставщика")
            if base_article_col not in self.base_df.columns:
                missing_columns.append(f"'{base_article_col}' в базе данных")
            if base_price_col not in self.base_df.columns:
                missing_columns.append(f"'{base_price_col}' в базе данных")

            if missing_columns:
                error_msg = f"❌ Не найдены столбцы: {', '.join(missing_columns)}"
                self.info_text.insert(tk.END, error_msg + "\n")
                self.finish_progress(error_msg, auto_reset=False)
                messagebox.showerror(
                    "Ошибка", f"Не найдены столбцы:\n{', '.join(missing_columns)}"
                )
                return

            # Создаем словарь цен поставщика
            supplier_prices = {}
            for _, row in self.current_df.iterrows():
                article = (
                    str(row[supplier_article_col]).strip()
                    if pd.notna(row[supplier_article_col])
                    else None
                )
                price = row[supplier_price_col]
                if article and pd.notna(price) and price > 0:
                    supplier_prices[article] = float(price)

            self.info_text.insert(
                tk.END, f"📊 Загружено {len(supplier_prices)} цен поставщика\n"
            )
            self.info_text.insert(
                tk.END, f"📊 Всего строк в базе: {len(self.base_df)}\n\n"
            )
            self.info_text.insert(tk.END, "🔍 Поиск изменений цен...\n")
            self.root.update()

            # Шаг 4: Поиск изменений
            self.update_progress(4, "Поиск изменений цен")
            updated_count = 0
            skipped_count = 0
            price_updates = []

            # Добавляем заголовок таблицы изменений
            self.info_text.insert(tk.END, "\n📋 ИЗМЕНЕНИЯ ЦЕН:\n")
            self.info_text.insert(tk.END, "-" * 80 + "\n")
            self.info_text.insert(
                tk.END, "№   Артикул          Старая цена   Новая цена    Изменение\n"
            )
            self.info_text.insert(tk.END, "-" * 80 + "\n")
            self.root.update()

            for idx, base_row in self.base_df.iterrows():
                article = (
                    str(base_row[base_article_col]).strip()
                    if pd.notna(base_row[base_article_col])
                    else None
                )

                if not article or article not in supplier_prices:
                    skipped_count += 1
                    continue

                supplier_price = supplier_prices[article]
                base_price = base_row[base_price_col]

                # Пропускаем если цены практически одинаковы
                if abs(float(supplier_price) - float(base_price)) < 0.001:
                    skipped_count += 1
                    continue

                # Вычисляем изменение цены
                price_diff = float(supplier_price) - float(base_price)
                if float(base_price) != 0:
                    change_percent = (price_diff / float(base_price)) * 100
                else:
                    change_percent = 100.0

                # Форматируем вывод
                change_sign = "+" if price_diff >= 0 else ""
                change_color = "green" if price_diff < 0 else "red"

                # Добавляем запись в лог
                update_record = {
                    "article": article,
                    "old_price": base_price,
                    "new_price": supplier_price,
                    "change_percent": change_percent,
                    "base_index": idx,
                }
                price_updates.append(update_record)
                updated_count += 1

                # Выводим информацию в текстовое поле
                self.info_text.insert(tk.END, f"{updated_count:3d} {article:15} ")
                self.info_text.insert(
                    tk.END,
                    f"{float(base_price):10.2f} → {float(supplier_price):10.2f} ",
                )
                self.info_text.tag_config(change_color, foreground=change_color)
                self.info_text.insert(
                    tk.END,
                    f"{change_sign}{price_diff:+.2f} ({change_sign}{change_percent:+.1f}%)\n",
                    change_color,
                )

                # Обновляем прогресс каждые 10 записей
                if updated_count % 10 == 0:
                    self.root.update()

            self.info_text.insert(tk.END, "-" * 80 + "\n")
            self.info_text.insert(tk.END, f"✅ Найдено изменений: {updated_count}\n")
            self.info_text.insert(tk.END, f"⏩ Пропущено: {skipped_count}\n\n")
            self.root.update()

            if updated_count == 0:
                self.finish_progress(
                    "ℹ️ Нет изменений цен для обновления", auto_reset=False
                )
                messagebox.showinfo("Информация", "Нет изменений цен для обновления")
                return

            # Шаг 5: Обновление файла
            self.update_progress(5, "Сохранение изменений")
            self.info_text.insert(tk.END, "💾 Подготовка к сохранению изменений...\n")
            self.root.update()

            # Запрос подтверждения
            confirm = messagebox.askyesno(
                "Подтверждение",
                f"Будет обновлено {updated_count} цен. Продолжить?",
                icon="question",
            )

            if not confirm:
                self.info_text.insert(tk.END, "❌ Обновление отменено пользователем\n")
                self.finish_progress(
                    "⏹️ Обновление цен отменено пользователем", auto_reset=False
                )
                return

            try:
                base_file_path = os.path.join("data/input", self.base_file_name)
                self.info_text.insert(
                    tk.END, f"📝 Сохранение изменений в файл: {self.base_file_name}\n"
                )
                self.root.update()

                # Вызываем метод обновления Excel
                success = self.update_excel_prices_preserve_formatting(
                    base_file_path, None, price_updates, self.current_config
                )

                if success:
                    self.price_updated = True
                    self.price_updates_log = price_updates

                    # Итоговый отчет
                    self.info_text.insert(tk.END, "\n" + "=" * 50 + "\n")
                    self.info_text.insert(tk.END, "✅ ОБНОВЛЕНИЕ УСПЕШНО ЗАВЕРШЕНО\n")
                    self.info_text.insert(tk.END, f"Обновлено цен: {updated_count}\n")
                    self.info_text.insert(
                        tk.END, f"Пропущено записей: {skipped_count}\n"
                    )
                    self.info_text.insert(tk.END, f"Файл: {self.base_file_name}\n")

                    self.finish_progress(
                        f"✅ Цены успешно обновлены ({updated_count} шт.)",
                        auto_reset=False,
                    )
                    messagebox.showinfo(
                        "Успех",
                        f"Цены успешно обновлены!\n\nОбновлено: {updated_count}\nПропущено: {skipped_count}",
                    )
                    self.update_prices_button.config(state="disabled")
                else:
                    self.info_text.insert(
                        tk.END, "❌ Ошибка при сохранении изменений\n"
                    )
                    self.finish_progress("Ошибка обновления", auto_reset=False)
                    messagebox.showerror(
                        "Ошибка", "Не удалось обновить цены в Excel файле"
                    )

            except Exception as e:
                error_msg = f"❌ Ошибка сохранения: {str(e)}"
                self.info_text.insert(tk.END, error_msg + "\n")
                self.log_error(error_msg)
                self.finish_progress("Ошибка сохранения", auto_reset=False)
                messagebox.showerror("Ошибка", f"Не удалось сохранить изменения: {e}")

        except Exception as e:
            error_msg = f"❌ Неожиданная ошибка: {str(e)}"
            self.info_text.insert(tk.END, error_msg + "\n")
            self.log_error(error_msg)
            self.finish_progress("Ошибка обновления", auto_reset=False)
            messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")

    def create_simple_new_items_table(self, parent_frame, new_items, dialog):
        """
        Создать упрощенную таблицу для новых товаров с минимальными колонками
        """
        # Создаем основной фрейм
        self.set_status("🔧 Создание структуры таблицы новых товаров...", "loading")
        table_frame = ttk.Frame(parent_frame)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Создаем Treeview с колонками
        columns = (
            "supplier_article",
            "supplier_name",
            "supplier_color",
            "supplier_price",
            "found_base_name",
            "base_row_number",
            "base_color",
            "base_price",
        )
        tree = ttk.Treeview(
            table_frame, columns=columns, show="tree headings", height=10
        )

        # Настройка колонок
        self.set_status("📋 Настройка колонок таблицы...", "loading")
        tree.heading("#0", text="✓", anchor="w")
        tree.column("#0", width=40, minwidth=40, stretch=False)

        tree.heading("supplier_article", text="Артикул поставщика", anchor="w")
        tree.column("supplier_article", width=50, minwidth=50, stretch=True)

        tree.heading("supplier_name", text="Название", anchor="w")
        tree.column("supplier_name", width=300, minwidth=200, stretch=True)

        tree.heading("supplier_color", text="Цвет товара", anchor="w")
        tree.column("supplier_color", width=80, minwidth=60, stretch=True)

        tree.heading("supplier_price", text="Цена", anchor="w")
        tree.column("supplier_price", width=60, minwidth=60, stretch=True)

        tree.heading("found_base_name", text="Найденный товар в базе", anchor="w")
        tree.column("found_base_name", width=300, minwidth=200, stretch=True)

        tree.heading("base_row_number", text="Строка в базе", anchor="w")
        tree.column("base_row_number", width=50, minwidth=50, stretch=True)

        tree.heading("base_color", text="Цвет из базы", anchor="w")
        tree.column("base_color", width=80, minwidth=60, stretch=True)

        tree.heading("base_price", text="Цена из базы", anchor="w")
        tree.column("base_price", width=80, minwidth=60, stretch=True)

        # Скроллбары
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        h_scrollbar = ttk.Scrollbar(
            table_frame, orient="horizontal", command=tree.xview
        )
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Добавляем данные
        self.set_status(
            f"📊 Заполнение таблицы данными ({len(new_items)} элементов)...", "loading"
        )
        for i, item in enumerate(new_items):
            # Показываем полные названия товаров без обрезания
            supplier_name = item.get("name", "")
            supplier_article = str(item.get("article", "N/A"))

            # Улучшенная обработка цвета
            color_value = item.get("color", "")
            supplier_color = self.safe_color_processing(color_value)
            if not supplier_color:
                supplier_color = "N/A"

            # Безопасное форматирование цены
            price_value = item.get("price", 0)
            if price_value is None or pd.isna(price_value):
                supplier_price = "N/A"
            else:
                try:
                    supplier_price = f"${float(price_value):.2f}"
                except (ValueError, TypeError):
                    supplier_price = "N/A"

            # Ищем товар в базе по нечеткому сопоставлению
            found_base_name, base_row_number, base_color, base_price = (
                self.find_item_by_fuzzy_matching(supplier_name)
            )

            # Вставляем строку с чекбоксом
            item_id = tree.insert(
                "",
                "end",
                text="☐",  # По умолчанию сброшенный чекбокс
                values=(
                    supplier_article,
                    supplier_name,
                    supplier_color,
                    supplier_price,
                    found_base_name,
                    base_row_number,
                    base_color,
                    base_price,
                ),
            )

            # Создаем виртуальный чекбокс для совместимости
            checkbox = type("Checkbox", (), {})()
            checkbox.var = tk.BooleanVar(value=False)
            checkbox.match_data = item
            checkbox.match_type = "new_item"
            checkbox.item_id = item_id
            checkbox.tree = tree

            # Сохраняем информацию о найденном совпадении в базе
            if base_row_number != "N/A":
                try:
                    checkbox.base_row_number = int(base_row_number)
                except (ValueError, TypeError) as e:
                    self.log_error(
                        f"❌ Ошибка преобразования номера строки '{base_row_number}' для товара '{supplier_name}': {e}"
                    )
                    checkbox.base_row_number = None
            else:
                checkbox.base_row_number = None

            # Обновляем отображение при изменении состояния
            def update_display(checkbox=checkbox):
                if checkbox.var.get():
                    tree.item(checkbox.item_id, text="☑")
                else:
                    tree.item(checkbox.item_id, text="☐")

            checkbox.var.trace("w", lambda *args, cb=checkbox: update_display(cb))

            dialog.checkboxes.append(checkbox)
            dialog.new_item_checkboxes.append(checkbox)

        # Функция для переключения чекбокса по клику
        def on_item_click(event):
            x, y = event.x, event.y
            item = tree.identify_row(y)
            column = tree.identify_column(x)

            if item and column == "#0":  # Клик в колонке чекбокса (#0)
                for checkbox in dialog.checkboxes:
                    if hasattr(checkbox, "item_id") and checkbox.item_id == item:
                        old_value = checkbox.var.get()
                        checkbox.var.set(not old_value)
                        break

        # Функция для переключения чекбокса по двойному клику на название товара
        def on_item_double_click(event):
            x, y = event.x, event.y
            item = tree.identify_row(y)
            column = tree.identify_column(x)

            # Если клик по колонке с названием товара (supplier_name) или любой другой колонке кроме чекбокса
            if item and column != "#0":
                for checkbox in dialog.checkboxes:
                    if hasattr(checkbox, "item_id") and checkbox.item_id == item:
                        old_value = checkbox.var.get()
                        checkbox.var.set(not old_value)
                        break

        tree.bind("<Button-1>", on_item_click)
        tree.bind("<Double-Button-1>", on_item_double_click)

        # Упаковываем компоненты
        tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Создаем фрейм для кнопок этой вкладки
        tab_button_frame = ttk.Frame(parent_frame)
        tab_button_frame.pack(fill="x", pady=(10, 0))

        # Функции для этой конкретной вкладки
        # Кнопки для этой вкладки
        ttk.Button(
            tab_button_frame,
            text="✅ Выбрать все на вкладке",
            command=lambda: self.select_all_matches(dialog, dialog.new_item_checkboxes),
        ).pack(side="left", padx=5)

        ttk.Button(
            tab_button_frame,
            text="❌ Снять все на вкладке",
            command=lambda: self.deselect_all_matches(
                dialog, dialog.new_item_checkboxes
            ),
        ).pack(side="left", padx=5)

    def find_item_by_fuzzy_matching(self, supplier_name):
        """
        Поиск товара в базе по нечеткому сопоставлению названий

        Параметры:
        - supplier_name: название товара поставщика

        Возвращает:
        - (найденное название, номер строки, цвет, цена) или ("Не найдено", "N/A", "N/A", "N/A")
        """
        try:
            # Проверяем, что база загружена
            if self.base_df is None or self.base_df.empty:
                return "Не найдено", "N/A", "N/A", "N/A"

            if not supplier_name or supplier_name.strip() == "":
                return "Не найдено", "N/A", "N/A", "N/A"

            import difflib

            # Определяем название колонки для названий товаров в базе
            base_name_col = self._get_base_name_column(self.base_df)
            if not base_name_col:
                self.log_error(
                    "❌ Не удалось определить колонку с названиями товаров в базе"
                )
                return "Не найдено", "N/A", "N/A", "N/A"

            # Порог схожести (0.3 = 30%)
            similarity_threshold = TRSH
            best_match = None
            best_ratio = 0
            best_idx = None

            # Ищем товары в базе с наилучшим совпадением
            for idx, base_row in self.base_df.iterrows():
                base_name = str(base_row.get(base_name_col, "")).strip()
                if not base_name or base_name == "nan":
                    continue

                # Вычисляем схожесть названий
                ratio = difflib.SequenceMatcher(
                    None, supplier_name.lower(), base_name.lower()
                ).ratio()

                # Если схожесть выше порога и лучше предыдущего
                if ratio >= similarity_threshold and ratio > best_ratio:
                    best_ratio = ratio
                    best_match = base_name
                    best_idx = idx

            # Если нашли хорошее совпадение
            if best_match:
                base_color = self.safe_color_processing(
                    self.base_df.iloc[best_idx].get("color", "")
                )
                if not base_color:
                    base_color = "N/A"

                # Получаем цену из базы
                base_price_value = self.base_df.iloc[best_idx].get("price_usd", 0)
                if base_price_value is None or pd.isna(base_price_value):
                    base_price = "N/A"
                else:
                    try:
                        base_price = f"${float(base_price_value):.2f}"
                    except (ValueError, TypeError):
                        base_price = "N/A"

                # Получаем реальный номер строки в Excel файле
                # Используем iloc для получения реального индекса строки
                excel_row_number = (
                    self.base_df.index.get_loc(best_idx) + 2
                )  # +2 для Excel (заголовок + 1-индексация)

                return (
                    best_match,
                    str(excel_row_number),  # Реальный номер строки в Excel
                    base_color,
                    base_price,
                )

            return "Не найдено", "N/A", "N/A", "N/A"

        except Exception as e:
            self.log_error(f"❌ Ошибка нечеткого поиска: {e}")
            return "Не найдено", "N/A", "N/A", "N/A"

    def refresh_interface(self):
        """Обновление интерфейса"""
        self.log_info("🔄 Обновление интерфейса...")

        # Обновляем список доступных конфигураций
        self.load_available_configs()

        # Обновляем статус
        self.set_status("Интерфейс обновлён", "success")
        self.root.update()

        self.log_info("✅ Интерфейс обновлён")

    def show_help(self):
        """Показать справку по использованию"""
        help_text = """🚀 MiStockSync - Справка

📁 Загрузка файлов:
• Файлы Вити должны содержать 'JHT' в названии
• Файлы Димы должны содержать 'DiMi' в названии  
• База данных должна содержать 'BASE' в названии

🔍 Процесс работы:
1. Выберите файл поставщика (или поставьте галочку 'авто')
2. Нажмите 'Сравнить с базой' для анализа
3. Используйте 'Сохранить отчет' для Excel отчёта
4. 'Обновить цены' для применения изменений

⚙️ Фильтрация:
• Витя: только товары "Имеются в нал."
• Дима: исключает товары "Ожидается"
• Цены: исключает NaN, пустые и нулевые

📊 Папки:
• data/input - исходные файлы
• data/output - результаты работы
• logs/ - файлы логов"""

        messagebox.showinfo("Справка", help_text)
        self.log_info("❓ Показана справка пользователю")

    def show_log_window(self):
        """Показать окно с логами приложения"""
        try:
            # Создаем новое окно для логов
            log_window = tk.Toplevel(self.root)
            log_window.title("📋 Логи MiStockSync")
            log_window.geometry("900x600")
            log_window.resizable(True, True)

            # Заголовок
            header_frame = ttk.Frame(log_window)
            header_frame.pack(fill="x", padx=10, pady=5)

            ttk.Label(
                header_frame,
                text="📋 Логи работы приложения",
                font=("Arial", 14, "bold"),
            ).pack(anchor="w")

            # Панель поиска
            search_frame = ttk.Frame(log_window)
            search_frame.pack(fill="x", padx=10, pady=(0, 5))

            ttk.Label(search_frame, text="🔍 Поиск:").pack(side="left")

            search_var = tk.StringVar()
            search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
            search_entry.pack(side="left", padx=(5, 5))

            # Переменные для поиска
            search_results = []
            current_search_index = -1

            def search_text(direction="forward"):
                nonlocal search_results, current_search_index

                search_term = search_var.get().strip()
                if not search_term:
                    return

                # Если это новый поиск, ищем все вхождения
                if (
                    not search_results
                    or search_results
                    and search_results[0][0] != "1.0"
                ):
                    search_results = []
                    start_pos = "1.0"

                    while True:
                        pos = log_text.search(
                            search_term, start_pos, tk.END, nocase=True
                        )
                        if not pos:
                            break
                        end_pos = f"{pos}+{len(search_term)}c"
                        search_results.append((pos, end_pos))
                        start_pos = end_pos

                    if not search_results:
                        messagebox.showinfo("Поиск", f"Текст '{search_term}' не найден")
                        return

                    current_search_index = -1

                # Определяем следующий индекс
                if direction == "forward":
                    current_search_index = (current_search_index + 1) % len(
                        search_results
                    )
                else:  # backward
                    current_search_index = (current_search_index - 1) % len(
                        search_results
                    )

                # Выделяем найденный текст
                start_pos, end_pos = search_results[current_search_index]
                log_text.tag_remove("search_highlight", "1.0", tk.END)
                log_text.tag_add("search_highlight", start_pos, end_pos)
                log_text.tag_config(
                    "search_highlight", background="yellow", foreground="black"
                )

                # Прокручиваем к найденному тексту
                log_text.see(start_pos)

                # Обновляем статус в заголовке окна
                log_window.title(
                    f"📋 Логи MiStockSync - Найдено: {current_search_index + 1}/{len(search_results)}"
                )

            def search_forward():
                search_text("forward")

            def search_backward():
                search_text("backward")

            def clear_search():
                log_text.tag_remove("search_highlight", "1.0", tk.END)
                search_var.set("")
                search_results.clear()
                current_search_index = -1
                log_window.title("📋 Логи MiStockSync")

            def on_search_change(*args):
                # Очищаем поиск при изменении текста
                if not search_var.get().strip():
                    clear_search()

            # Кнопки поиска
            ttk.Button(search_frame, text="⬇️ Вперед", command=search_forward).pack(
                side="left", padx=(0, 5)
            )
            ttk.Button(search_frame, text="⬆️ Назад", command=search_backward).pack(
                side="left", padx=(0, 5)
            )
            ttk.Button(search_frame, text="❌ Очистить", command=clear_search).pack(
                side="left", padx=(0, 5)
            )

            # Привязываем Enter к поиску вперед
            def on_search_enter(event):
                search_forward()

            search_entry.bind("<Return>", on_search_enter)

            # Привязываем изменение текста к очистке поиска
            search_var.trace("w", on_search_change)

            # Горячие клавиши для поиска
            def on_key_press(event):
                if event.state & 4:  # Ctrl
                    if event.keysym == "f":
                        search_entry.focus()
                        return "break"
                    elif event.keysym == "F":  # Ctrl+Shift+F
                        search_backward()
                        return "break"

            log_window.bind("<Key>", on_key_press)

            # Основное текстовое поле с логами
            text_frame = ttk.Frame(log_window)
            text_frame.pack(fill="both", expand=True, padx=10, pady=5)

            # Текстовое поле с прокруткой
            log_text = tk.Text(
                text_frame,
                wrap=tk.WORD,
                state="normal",
                font=("Consolas", 10),
                bg="#f8f9fa",
                fg="#333333",
            )

            # Скроллбар
            scrollbar = ttk.Scrollbar(
                text_frame, orient="vertical", command=log_text.yview
            )
            log_text.configure(yscrollcommand=scrollbar.set)

            log_text.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # Функция обновления логов (определяем заранее)
            def refresh_logs():
                log_text.configure(state="normal")  # Разрешаем редактирование
                log_text.delete(1.0, tk.END)
                if hasattr(self, "log_file_path") and os.path.exists(
                    self.log_file_path
                ):
                    try:
                        with open(self.log_file_path, "r", encoding="utf-8") as f:
                            log_content = f.read()
                            log_text.insert(tk.END, log_content)
                            log_text.see(tk.END)
                    except Exception as e:
                        log_text.insert(tk.END, f"Ошибка чтения лог-файла: {e}\n")
                else:
                    # Если нет файла логов, показываем логи из памяти
                    log_text.insert(tk.END, "Логи из памяти приложения:\n")
                    log_text.insert(tk.END, "=" * 50 + "\n")
                    # Здесь можно добавить логи из self.logger если нужно
                log_text.configure(state="disabled")  # Блокируем редактирование

            # Контекстное меню для текстового поля
            context_menu = tk.Menu(log_text, tearoff=0)

            context_menu.add_command(
                label="📋 Копировать выделенное",
                command=lambda: self.copy_selected_text(log_text, log_window),
            )
            context_menu.add_command(
                label="📋 Копировать все",
                command=lambda: self.copy_all_text(log_text, log_window),
            )
            context_menu.add_separator()
            context_menu.add_command(label="🔄 Обновить", command=refresh_logs)

            def show_context_menu(event):
                context_menu.post(event.x_root, event.y_root)

            log_text.bind("<Button-3>", show_context_menu)  # Правый клик

            # Читаем логи из файла
            log_file = getattr(
                self,
                "log_file_path",
                os.path.join(
                    "logs", f"mistocksync_{datetime.now().strftime('%Y%m%d')}.log"
                ),
            )

            if os.path.exists(log_file):
                try:
                    with open(log_file, "r", encoding="utf-8") as f:
                        log_content = f.read()
                        log_text.insert(tk.END, log_content)
                        log_text.see(tk.END)  # Прокручиваем к концу
                except Exception as e:
                    log_text.insert(tk.END, f"Ошибка чтения лог-файла: {e}\n")
            else:
                log_text.insert(
                    tk.END,
                    f"Лог-файл не найден: {log_file}\nЛоги будут появляться здесь по мере работы приложения.\n",
                )

            # Кнопки управления
            button_frame = ttk.Frame(log_window)
            button_frame.pack(fill="x", padx=10, pady=5)

            # Кнопка обновления
            ttk.Button(button_frame, text="🔄 Обновить", command=refresh_logs).pack(
                side="left"
            )

            # Кнопка очистки
            def clear_logs():
                log_text.configure(state="normal")  # Разрешаем редактирование
                log_text.delete(1.0, tk.END)
                log_text.insert(tk.END, "Логи очищены.\n")
                log_text.configure(state="disabled")  # Блокируем редактирование

            ttk.Button(button_frame, text="🗑️ Очистить", command=clear_logs).pack(
                side="left", padx=(10, 0)
            )

            # Кнопка копирования всего текста
            ttk.Button(
                button_frame,
                text="📋 Копировать все",
                command=lambda: self.copy_all_text_with_notification(
                    log_text, log_window
                ),
            ).pack(side="left", padx=(10, 0))

            # Кнопка копирования выделенного текста
            ttk.Button(
                button_frame,
                text="📋 Копировать выделенное",
                command=lambda: self.copy_selected_text_with_notification(
                    log_text, log_window
                ),
            ).pack(side="left", padx=(10, 0))

            # Разделитель
            ttk.Separator(button_frame, orient="vertical").pack(
                side="left", fill="y", padx=10
            )

            # Кнопка закрытия
            ttk.Button(
                button_frame, text="❌ Закрыть", command=log_window.destroy
            ).pack(side="right")

            # Делаем поле только для чтения
            log_text.configure(state="disabled")

            # Обновляем статус
            self.set_status("📋 Окно логов открыто", "info")

        except Exception as e:
            self.log_error(f"Ошибка открытия окна логов: {e}")
            messagebox.showerror("Ошибка", f"Не удалось открыть окно логов: {e}")

    def create_advanced_status_bar(self, main_frame):
        """Создание продвинутого многосекционного статус-бара"""
        # Основной фрейм статус-бара
        self.status_frame = ttk.Frame(main_frame, relief=tk.SUNKEN, borderwidth=1)
        self.status_frame.grid(
            row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0)
        )
        self.status_frame.columnconfigure(0, weight=1)

        # Внутренний фрейм для компонентов
        inner_frame = ttk.Frame(self.status_frame)
        inner_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=5, pady=2)
        inner_frame.columnconfigure(0, weight=1)

        # Левая часть - основной статус с иконкой
        self.status_main = tk.StringVar(value="🚀 Готов к работе")
        self.status_label = ttk.Label(
            inner_frame, textvariable=self.status_main, anchor=tk.W
        )
        self.status_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))

        # Центр - прогресс-бар (скрыт по умолчанию)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            inner_frame, variable=self.progress_var, length=200, mode="determinate"
        )
        self.progress_bar.grid(row=0, column=1, padx=(10, 10))
        self.progress_bar.grid_remove()  # Изначально скрыт

        # Правая часть - дополнительная информация о файлах
        self.status_info = tk.StringVar(value="")
        self.info_label = ttk.Label(
            inner_frame, textvariable=self.status_info, anchor=tk.E
        )
        self.info_label.grid(row=0, column=2, padx=(10, 0))

        # Обновляем информацию о файлах при создании
        self.update_files_info()

        # Инициализация переменных для прогресса ⓘ
        self.is_progress_visible = False
        self.current_operation = None

        # Принудительное обновление GUI после создания статус-бара
        self.root.update()
        self.root.update_idletasks()

    def set_status(self, message, status_type="info", show_time=True):
        """Установка красивого статуса с иконками и цветами"""
        icons = {
            "loading": "⏳",
            "success": "✅",
            "error": "❌",
            "warning": "⚠️",
            "info": "🚀",
            "file": "📁",
            "save": "💾",
            "compare": "🔍",
            "update": "⏳",
            "report": "📊",
            "backup": "🛡️",
        }

        # Цвета для разных типов статусов
        colors = {
            "loading": "#9932CC",  # Orange
            "success": "#228B22",  # ForestGreen
            "error": "#DC143C",  # Crimson
            "warning": "#FFD700",  # Gold
            "info": "#4169E1",  # RoyalBlue
            "file": "#9932CC",  # DarkOrchid
            "save": "#9932CC",  # LimeGreen #32CD32
            "compare": "#9932CC",  # DodgerBlue #1E90FF
            "update": "#9932CC",  # BlueViolet
            "report": "#20B2AA",  # LightSeaGreen
            "backup": "#CD853F",  # Peru
        }

        icon = icons.get(status_type, "🚀")
        color = colors.get(status_type, "#000000")

        formatted_message = f"{icon} {message}"
        self.status_main.set(formatted_message)
        self.status_label.config(foreground=color)

        # Добавляем время если нужно
        if show_time:
            current_time = datetime.now().strftime("%H:%M:%S")
            self.status_info.set(f"🕐 {current_time}")

        # Принудительное обновление GUI
        self.root.update()
        self.root.update_idletasks()

        # Обновляем информацию о файлах
        self.update_files_info()

    def update_files_info(self):
        """Обновление информации о загруженных файлах в статус-баре"""
        info_parts = []

        # Информация о прайсе поставщика
        if self.current_df is not None:
            supplier_info = f"💼 {self.current_config or 'поставщик'}"
            if hasattr(self, "current_file_name") and self.current_file_name:
                supplier_info += f": {self.current_file_name}"
            info_parts.append(supplier_info)

        # Информация о базе данных
        if self.base_df is not None:
            base_info = "🏢 база"
            if hasattr(self, "base_file_name") and self.base_file_name:
                base_info += f": {self.base_file_name}"
            info_parts.append(base_info)

        # Формируем итоговую строку
        if info_parts:
            files_info = " | ".join(info_parts)
            self.status_info.set(f"📁 {files_info}")
        else:
            self.status_info.set("📁 Файлы не загружены")

    def start_progress(self, message, total_steps, operation_type="loading"):
        """Запуск прогресс-бара для длительной операции"""
        # Проверяем, что прогресс-бар существует
        if not hasattr(self, "progress_bar") or not self.progress_bar:
            self.log_error("❌ Прогресс-бар не инициализирован")
            return

        self.current_operation = {
            "message": message,
            "total": total_steps,
            "current": 0,
            "type": operation_type,
        }

        # Настраиваем прогресс-бар
        self.progress_var.set(0)
        self.progress_bar.config(maximum=total_steps)

        # Показываем прогресс-бар
        self.progress_bar.grid()
        self.is_progress_visible = True

        # Устанавливаем статус
        self.set_status(f"{message} (0/{total_steps})", operation_type, show_time=True)

        # Принудительное обновление GUI
        self.root.update()
        self.root.update_idletasks()

        # Дополнительное обновление для корректного отображения прогресс-бара
        self.progress_bar.update()
        self.status_frame.update()

    def update_progress(self, step, message=None):
        """Обновление прогресс-бара"""
        if not self.is_progress_visible or not self.current_operation:
            return

        # Проверяем, что прогресс-бар существует
        if not hasattr(self, "progress_bar") or not self.progress_bar:
            return

        self.current_operation["current"] = step
        self.progress_var.set(step)

        # Обновляем сообщение
        if message:
            display_message = message
        else:
            display_message = self.current_operation["message"]

        total = self.current_operation["total"]
        operation_type = self.current_operation["type"]

        # Вычисляем процент
        percent = int((step / total) * 100) if total > 0 else 0

        self.set_status(
            f"{display_message} ({step}/{total}) - {percent}%",
            operation_type,
            show_time=True,
        )

        # Принудительное обновление GUI
        self.root.update()
        self.root.update_idletasks()

    def finish_progress(self, success_message="Операция завершена", auto_reset=True):
        """Завершение прогресс-бара"""
        if not self.is_progress_visible:
            return

        # Проверяем, что прогресс-бар существует
        if not hasattr(self, "progress_bar") or not self.progress_bar:
            return

        # Скрываем прогресс-бар
        self.progress_bar.grid_remove()
        self.is_progress_visible = False

        # Показываем финальное сообщение
        self.set_status(success_message, "success", show_time=True)

        # Автосброс через 3 секунды
        if auto_reset:
            self.root.after(3000, lambda: self.set_status("Готов к работе", "info"))

        self.current_operation = None
        # Принудительное обновление GUI
        self.root.update()
        self.root.update_idletasks()

    def set_temp_status(self, message, status_type="info", duration=2000):
        """Временный статус с автосбросом"""
        old_status = self.status_main.get()
        old_color = self.status_label.cget("foreground")

        self.set_status(message, status_type)

        # Автоматический сброс
        def reset_status():
            self.status_main.set(old_status)
            self.status_label.config(foreground=old_color)

        self.root.after(duration, reset_status)

    def set_animated_status(
        self, base_message, status_type="loading", animation_chars="⠋⠙⠹⠸⠼⠴⠦⠧⠇⠏"
    ):
        """Установка анимированного статуса с вращающимися символами"""
        if not hasattr(self, "_animation_counter"):
            self._animation_counter = 0
        if not hasattr(self, "_animation_job"):
            self._animation_job = None

        # Останавливаем предыдущую анимацию
        if self._animation_job:
            self.root.after_cancel(self._animation_job)

        def animate():
            char = animation_chars[self._animation_counter % len(animation_chars)]
            animated_message = f"{char} {base_message}"
            self.set_status(animated_message, status_type, show_time=False)

            self._animation_counter += 1
            self._animation_job = self.root.after(
                100, animate
            )  # Обновляем каждые 100мс

        # Запускаем анимацию
        animate()

        return self._animation_job  # Возвращаем ID для возможности остановки

    def stop_animated_status(self):
        """Остановка анимированного статуса"""
        if hasattr(self, "_animation_job") and self._animation_job:
            self.root.after_cancel(self._animation_job)
            self._animation_job = None

    def add_to_base(self):
        """Добавление артикулов по кодам и новых товаров в базу данных"""
        self.log_info("🔄 Добавление данных в базу...")

        # Проверяем что есть результаты сравнения
        if self.comparison_result is None:
            self.log_error("❌ Сначала выполните сравнение с базой данных")
            messagebox.showwarning(
                "Предупреждение", "Сначала выполните сравнение с базой данных"
            )
            return

        # Получаем данные для добавления
        code_matches = self.comparison_result.get("code_matches", [])
        bracket_matches = self.comparison_result.get("bracket_matches", [])
        new_items = self.comparison_result.get("new_items", [])

        if not code_matches and not bracket_matches and not new_items:
            self.log_info("ℹ️ Нет данных для добавления в базу")
            messagebox.showinfo("Информация", "Нет новых данных для добавления в базу")
            return

        # Подсчитываем общее количество элементов для добавления
        total_items = len(code_matches) + len(bracket_matches) + len(new_items)

        # Запускаем прогресс-бар с анимированной индикацией
        self.start_progress("Добавление данных в базу", 5, "loading")
        self.set_animated_status("Добавление данных в базу", "loading")

        try:
            # Этап 1: Подготовка данных
            self.update_progress(1, f"Подготовка к добавлению {total_items} элементов")
            self.set_status("📋 Подготовка данных для добавления...", "loading")

            # Этап 2: Показ диалога выбора
            self.update_progress(2, "Открытие диалога выбора артикулов")
            self.set_status("🔍 Открытие диалога выбора артикулов...", "loading")

            # Показываем окно выбора артикулов для добавления
            self.show_add_articles_dialog(bracket_matches, code_matches, new_items)

            # Этап 3: Завершение
            self.update_progress(5, "Диалог выбора артикулов готов")
            self.set_status("✅ Диалог выбора артикулов готов", "success")

        except Exception as e:
            self.log_error(f"❌ Ошибка при добавлении в базу: {e}")
            self.set_status(f"❌ Ошибка: {e}", "error")
            messagebox.showerror("Ошибка", f"Не удалось добавить данные в базу: {e}")
        finally:
            # Завершаем прогресс
            self.finish_progress("Добавление данных завершено", auto_reset=True)
            self.stop_animated_status()

    def show_add_articles_dialog(self, bracket_matches, code_matches, new_items):
        """Показать диалог выбора артикулов для добавления с детальной проверкой"""

        # Обновляем прогресс - создание диалога
        self.update_progress(2, "Создание диалога выбора артикулов")
        self.set_status("🔧 Создание диалога выбора артикулов...", "loading")

        # Определяем поставщика для заголовка
        supplier_name = ""
        if self.current_config == "vitya":
            supplier_name = "Витя"
        elif self.current_config == "dimi":
            supplier_name = "Дима"
        else:
            supplier_name = self.current_config.upper()

        # Создаем главное окно диалога
        self.set_status("🏗️ Создание окна диалога...", "loading")
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Добавление артикулов поставщика {supplier_name}")
        dialog.resizable(True, True)

        # Загружаем сохраненные размеры или используем по умолчанию
        settings = self.load_settings()
        dialog_width = settings.get("add_articles_dialog_width", 800)
        dialog_height = settings.get("add_articles_dialog_height", 533)

        dialog.geometry(f"{dialog_width}x{dialog_height}")

        # Центрируем окно относительно основного окна
        self.set_status("📍 Позиционирование окна...", "loading")
        self.center_window(dialog, dialog_width, dialog_height, self.root)

        # Делаем окно модальным
        self.set_status("🔒 Настройка модальности окна...", "loading")
        dialog.transient(self.root)
        dialog.grab_set()

        # Обработчик изменения размера окна
        def on_dialog_resize(event):
            if event.widget == dialog:
                new_width = event.width
                new_height = event.height
                if new_width > 100 and new_height > 100:
                    settings["add_articles_dialog_width"] = new_width
                    settings["add_articles_dialog_height"] = new_height
                    self.save_settings(settings)

        # Обработчик закрытия окна
        def on_dialog_close():
            dialog.destroy()

        # Привязываем обработчики
        dialog.bind("<Configure>", on_dialog_resize)
        dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)

        # Заголовок
        self.set_status("📝 Создание заголовка диалога...", "loading")
        header_frame = ttk.Frame(dialog)
        header_frame.pack(fill="x", padx=10, pady=10)

        ttk.Label(
            header_frame,
            text=f"🔗 Добавление артикулов поставщика {supplier_name}",
            font=("Arial", 14, "bold"),
        ).pack(anchor="w")

        info_text = "Найдены товары по кодам в названии и артикулах других поставщиков.\nВыберите какие артикулы добавить в базу:"
        ttk.Label(header_frame, text=info_text, font=("Arial", 10)).pack(
            anchor="w", pady=(5, 0)
        )

        # Инициализируем список чекбоксов ПЕРЕД созданием вкладок
        dialog.checkboxes = []  # Список всех чекбоксов
        dialog.code_checkboxes = []  # Чекбоксы для кодов
        dialog.new_item_checkboxes = []  # Чекбоксы для новых товаров
        dialog.code_matches = bracket_matches + code_matches  # Объединяем все коды
        dialog.new_items = new_items

        # Создаем фреймы для разных типов совпадений
        self.set_status("📑 Создание вкладок диалога...", "loading")
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # Вкладка 1: Коды (объединенные)
        if dialog.code_matches:
            self.set_status(
                f"📊 Создание таблицы кодов ({len(dialog.code_matches)} элементов)...",
                "loading",
            )
            code_frame = ttk.Frame(notebook)
            notebook.add(code_frame, text=f"🔗 Коды ({len(dialog.code_matches)})")
            self.create_matches_table_with_buttons(
                code_frame, dialog.code_matches, "code", dialog
            )

        # Вкладка 2: Новые товары
        if new_items:
            self.set_status(
                f"📥 Создание таблицы новых товаров ({len(new_items)} элементов)...",
                "loading",
            )
            new_frame = ttk.Frame(notebook)
            notebook.add(new_frame, text=f"📥 Новые товары ({len(new_items)})")
            # Используем упрощенную таблицу с минимальными колонками
            self.create_simple_new_items_table(new_frame, new_items, dialog)

        # Нижняя панель с кнопками
        self.set_status("🔘 Создание панели управления...", "loading")
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Кнопки управления
        ttk.Button(
            button_frame,
            text="✅ Выбрать все",
            command=lambda: self.select_all_matches(dialog),
        ).pack(side="left", padx=5)

        ttk.Button(
            button_frame,
            text="❌ Снять все",
            command=lambda: self.deselect_all_matches(dialog),
        ).pack(side="left", padx=5)

        # Разделитель
        ttk.Separator(button_frame, orient="vertical").pack(
            side="left", fill="y", padx=10
        )

        # Кнопка добавления
        add_button = ttk.Button(
            button_frame,
            text="🔗 Добавить выбранные артикулы",
            command=lambda: self.process_selected_articles(
                dialog, dialog.code_matches, dialog.new_items
            ),
        )
        add_button.pack(side="right", padx=5)

        ttk.Button(button_frame, text="❌ Отмена", command=on_dialog_close).pack(
            side="right", padx=5
        )

        # Обновляем прогресс - диалог готов
        self.update_progress(3, "Диалог выбора артикулов готов")
        self.set_status("✅ Диалог выбора артикулов готов", "success")

    def create_matches_table_with_buttons(
        self, parent_frame, matches, match_type, dialog
    ):
        """Создать таблицу совпадений с кнопками управления для вкладки"""

        # Создаем основной фрейм для содержимого вкладки
        self.set_status(f"🔧 Создание структуры таблицы {match_type}...", "loading")
        main_frame = ttk.Frame(parent_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Создаем таблицу
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill="both", expand=True)

        # Создаем таблицу (используем существующую функцию)
        self.create_matches_table(table_frame, matches, match_type, dialog)

        # Создаем фрейм для кнопок этой вкладки
        tab_button_frame = ttk.Frame(main_frame)
        tab_button_frame.pack(fill="x", pady=(10, 0))

        # Функции для этой конкретной вкладки
        # Кнопки для этой вкладки
        ttk.Button(
            tab_button_frame,
            text="✅ Выбрать все на вкладке",
            command=lambda: self.select_all_matches(
                dialog,
                (
                    dialog.code_checkboxes
                    if match_type == "code"
                    else dialog.new_item_checkboxes
                ),
            ),
        ).pack(side="left", padx=5)

        ttk.Button(
            tab_button_frame,
            text="❌ Снять все на вкладке",
            command=lambda: self.deselect_all_matches(
                dialog,
                (
                    dialog.code_checkboxes
                    if match_type == "code"
                    else dialog.new_item_checkboxes
                ),
            ),
        ).pack(side="left", padx=5)

    def create_matches_table(self, parent_frame, matches, match_type, dialog):
        """Создать таблицу совпадений с чекбоксами используя Treeview для лучшего выравнивания"""

        # Создаем основной фрейм
        self.set_status(f"🔧 Создание таблицы совпадений {match_type}...", "loading")
        table_frame = ttk.Frame(parent_frame)
        table_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Создаем Treeview с колонками
        columns = (
            "code",
            "supplier_name",
            "supplier_article",
            "supplier_color",
            "supplier_price",
            "base_name",
            "base_article",
            "base_color",
            "base_price",
        )
        tree = ttk.Treeview(
            table_frame, columns=columns, show="tree headings", height=10
        )

        # Настройка колонок с возможностью изменения размера
        self.set_status("📋 Настройка колонок таблицы совпадений...", "loading")
        tree.heading("#0", text="✓", anchor="w")
        tree.column("#0", width=40, minwidth=40, stretch=False)

        tree.heading("code", text="Код", anchor="w")
        tree.column("code", width=120, minwidth=80, stretch=True)

        tree.heading("supplier_name", text="Товар поставщика", anchor="w")
        tree.column("supplier_name", width=250, minwidth=150, stretch=True)

        tree.heading("supplier_article", text="Артикул поставщика", anchor="w")
        tree.column("supplier_article", width=120, minwidth=80, stretch=True)

        tree.heading("supplier_color", text="Цвет", anchor="w")
        tree.column("supplier_color", width=80, minwidth=60, stretch=True)

        tree.heading("supplier_price", text="Цена", anchor="w")
        tree.column("supplier_price", width=80, minwidth=60, stretch=True)

        tree.heading("base_name", text="Найденный товар", anchor="w")
        tree.column("base_name", width=250, minwidth=150, stretch=True)

        tree.heading("base_article", text="Артикул", anchor="w")
        tree.column("base_article", width=120, minwidth=80, stretch=True)

        tree.heading("base_color", text="Цвет", anchor="w")
        tree.column("base_color", width=80, minwidth=60, stretch=True)

        tree.heading("base_price", text="Цена", anchor="w")
        tree.column("base_price", width=80, minwidth=60, stretch=True)

        # Скроллбары
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        h_scrollbar = ttk.Scrollbar(
            table_frame, orient="horizontal", command=tree.xview
        )
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Добавляем данные
        self.set_status(
            f"📊 Заполнение таблицы совпадений ({len(matches)} элементов)...", "loading"
        )
        for i, match in enumerate(matches):
            # Показываем полные названия товаров без обрезания
            supplier_name = match.get("supplier_name", "")
            base_name = match.get("base_name", "")

            code = match.get("code", "N/A")
            supplier_article = str(match.get("supplier_article", "N/A"))
            supplier_color = str(match.get("supplier_color", "N/A"))
            supplier_price = f"${match.get('supplier_price', 0):.2f}"

            base_article = str(match.get("base_article", "N/A"))
            base_color = str(match.get("base_color", "N/A"))

            # Получаем цену из базы (из соответствующих колонок)
            base_price = "N/A"
            if self.current_config == "vitya":
                base_price = f"${match.get('price_vitya_usd', 0):.2f}"
            elif self.current_config == "dimi":
                base_price = f"${match.get('price_dimi_usd', 0):.2f}"
            elif self.current_config == "mila":
                base_price = f"${match.get('price_mila_usd', 0):.2f}"
            else:
                base_price = f"${match.get('base_price', 0):.2f}"

            # Вставляем строку с чекбоксом в первой колонке
            item_id = tree.insert(
                "",
                "end",
                text="☐",  # По умолчанию сброшенный чекбокс
                values=(
                    code,
                    supplier_name,
                    supplier_article,
                    supplier_color,
                    supplier_price,
                    base_name,
                    base_article,
                    base_color,
                    base_price,
                ),
            )

            # Создаем виртуальный чекбокс для совместимости
            checkbox = type("Checkbox", (), {})()  # Создаем простой объект
            checkbox.var = tk.BooleanVar(value=False)  # По умолчанию сброшены
            checkbox.match_data = match
            checkbox.match_type = match_type
            checkbox.item_id = item_id
            checkbox.tree = tree

            # Обновляем отображение при изменении состояния
            def update_display(checkbox=checkbox):
                if checkbox.var.get():
                    tree.item(checkbox.item_id, text="☑")
                else:
                    tree.item(checkbox.item_id, text="☐")

            checkbox.var.trace("w", lambda *args, cb=checkbox: update_display(cb))

            dialog.checkboxes.append(checkbox)

            # Добавляем в список конкретной вкладки
            if match_type == "code":
                dialog.code_checkboxes.append(checkbox)
            elif match_type == "new_item":
                dialog.new_item_checkboxes.append(checkbox)

        # Функция для переключения чекбокса по клику
        def on_item_click(event):
            # Получаем координаты клика
            x, y = event.x, event.y
            item = tree.identify_row(y)
            column = tree.identify_column(x)

            if item and column == "#1":  # Клик в первой колонке (чекбокс)
                # Находим соответствующий чекбокс
                for checkbox in dialog.checkboxes:
                    if hasattr(checkbox, "item_id") and checkbox.item_id == item:
                        checkbox.var.set(not checkbox.var.get())
                        break

        tree.bind("<Button-1>", on_item_click)

        # Упаковываем компоненты
        tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

    def select_all_matches(self, dialog, checkbox_list=None):
        """Универсальная функция выбора всех чекбоксов"""
        if checkbox_list is None:
            checkbox_list = getattr(dialog, "checkboxes", [])

        for checkbox in checkbox_list:
            checkbox.var.set(True)

    def deselect_all_matches(self, dialog, checkbox_list=None):
        """Универсальная функция снятия выбора со всех чекбоксов"""
        if checkbox_list is None:
            checkbox_list = getattr(dialog, "checkboxes", [])

        for checkbox in checkbox_list:
            checkbox.var.set(False)

    def process_selected_articles(self, dialog, code_matches, new_items):
        """Обработать выбранные пользователем артикулы"""

        # Собираем выбранные совпадения
        selected_matches = []
        selected_new_items = []  # Новые товары для вставки пустых строк

        for checkbox in dialog.checkboxes:
            if checkbox.var.get():  # Если чекбокс выбран
                if checkbox.match_type == "new_item":
                    # Для новых товаров сохраняем информацию о номере строки в базе
                    selected_new_items.append(
                        {
                            "match_data": checkbox.match_data,
                            "match_type": checkbox.match_type,
                            "base_row_number": None,  # Будет заполнено ниже
                        }
                    )
                else:
                    # Для обычных совпадений
                    selected_matches.append(
                        {
                            "match_data": checkbox.match_data,
                            "match_type": checkbox.match_type,
                        }
                    )

        # Проверяем, есть ли выбранные элементы
        if not selected_matches and not selected_new_items:
            messagebox.showwarning(
                "Предупреждение", "Не выбрано ни одного совпадения для добавления"
            )
            return

        # Для новых товаров получаем номера строк в базе
        if selected_new_items:
            self.log_info(f"🔍 Обработка {len(selected_new_items)} новых товаров...")
            found_count = 0
            not_found_count = 0

            for new_item in selected_new_items:
                supplier_name = new_item["match_data"].get("name", "")
                if supplier_name:
                    # Ищем товар в базе по нечеткому сопоставлению
                    found_base_name, base_row_number, base_color, base_price = (
                        self.find_item_by_fuzzy_matching(supplier_name)
                    )
                    if base_row_number != "N/A":
                        new_item["base_row_number"] = int(base_row_number)
                        found_count += 1
                    else:
                        new_item["base_row_number"] = None
                        not_found_count += 1

            # Логируем общий результат вместо деталей по каждому товару
            if found_count > 0:
                self.log_info(f"✅ Найдено в базе: {found_count} товаров")
            if not_found_count > 0:
                self.log_info(f"⚠️ Не найдено в базе: {not_found_count} товаров")

        # Подтверждение
        total_items = len(selected_matches) + len(selected_new_items)
        confirmation_message = f"Обработать {total_items} выбранных элементов?\n\n"

        if selected_matches:
            confirmation_message += (
                f"🔗 Артикулов для добавления: {len(selected_matches)}\n"
            )
        if selected_new_items:
            confirmation_message += (
                f"📝 Новых товаров для вставки строк: {len(selected_new_items)}\n"
            )

        result = messagebox.askyesno("Подтверждение", confirmation_message)

        if not result:
            return

        # Сохраняем размеры окна добавления артикулов перед закрытием
        try:
            dialog_width = dialog.winfo_width()
            dialog_height = dialog.winfo_height()
            if dialog_width > 100 and dialog_height > 100:
                settings = self.load_settings()
                settings["add_articles_dialog_width"] = dialog_width
                settings["add_articles_dialog_height"] = dialog_height
                self.save_settings(settings)
        except Exception as e:
            self.log_error(f"❌ Ошибка сохранения размеров диалога: {e}")

        # Закрываем диалог и выполняем добавление
        dialog.destroy()

        try:
            self.set_status(
                "🔄 Подготовка к добавлению выбранных артикулов...", "loading"
            )
            self.update_progress(4, "Подготовка к добавлению артикулов")
            articles_added = 0

            # Определяем столбец артикула поставщика в базе
            self.set_status("🔍 Определение структуры базы данных...", "loading")
            if self.current_config == "vitya":
                supplier_article_col = "article_vitya"
            elif self.current_config == "dimi":
                supplier_article_col = "article_dimi"
            else:
                supplier_article_col = "article"

            # Получаем оригинальное имя столбца для Excel из конфигурации
            excel_column_name = self.get_excel_column_name_from_config(
                supplier_article_col
            )

            # Обрабатываем выбранные совпадения
            self.set_status(
                f"🔄 Начинаем обработку {len(selected_matches)} выбранных совпадений...",
                "loading",
            )
            self.log_info(
                f"🔄 Начинаем обработку {len(selected_matches)} выбранных совпадений..."
            )
            self.log_info(f"📊 Столбец для артикулов: {supplier_article_col}")

            self.update_progress(
                2, f"Обработка {len(selected_matches)} выбранных артикулов"
            )

            for i, selected in enumerate(selected_matches, 1):
                match = selected["match_data"]
                match_type = selected["match_type"]

                # Обновляем статус для каждого элемента
                self.set_status(
                    f"📝 Обработка {i}/{len(selected_matches)}: {match.get('code', 'N/A')}...",
                    "loading",
                )

                try:
                    base_idx = match.get("base_index")
                    # Получаем артикул из разных возможных полей
                    supplier_article = match.get("supplier_article") or match.get(
                        "article"
                    )
                    code = match.get("code")

                    # Проверяем что у нас есть все необходимые данные
                    if (
                        base_idx is not None
                        and supplier_article
                        and str(supplier_article).strip() not in ["", "nan", "None"]
                    ):
                        # Проверяем что столбец существует в базе
                        if supplier_article_col not in self.base_df.columns:
                            self.log_error(
                                f"❌ Столбец {supplier_article_col} не найден в базе данных!"
                            )
                            continue

                        # Проверяем что индекс существует
                        if base_idx not in self.base_df.index:
                            self.log_error(
                                f"❌ Индекс {base_idx} не найден в базе данных!"
                            )
                            continue

                        # Определяем тип данных для столбца
                        data_type = self.get_column_data_type(supplier_article_col)

                        try:
                            # Преобразуем значение к нужному типу
                            if data_type == "int":
                                value = int(supplier_article)
                            elif data_type == "float":
                                value = float(supplier_article)
                            else:
                                value = str(
                                    supplier_article
                                )  # Для строк и других типов

                            # Проверяем что в базе нет уже этого артикула
                            current_article = self.base_df.loc[
                                base_idx, supplier_article_col
                            ]

                            if pd.isna(current_article) or str(
                                current_article
                            ).strip() in ["", "nan"]:
                                # Добавляем артикул поставщика
                                old_value = self.base_df.loc[
                                    base_idx, supplier_article_col
                                ]
                                self.base_df.loc[base_idx, supplier_article_col] = value
                                articles_added += 1

                                # Логируем только общую информацию, детали будут в отчете
                                if match_type == "bracket":
                                    self.log_info(
                                        f"✅ 📦 Добавлен артикул для кода в скобках: {code}"
                                    )
                                else:
                                    self.log_info(
                                        f"✅ 🔗 Добавлен артикул для кода: {code}"
                                    )

                                # Сохраняем информацию об изменении для отчета
                                if not hasattr(self, "changes_log"):
                                    self.changes_log = []

                                change_info = {
                                    "type": "article_added",
                                    "base_index": base_idx,
                                    "code": code,
                                    "match_type": match_type,
                                    "column": excel_column_name,
                                    "old_value": (
                                        str(old_value)
                                        if not pd.isna(old_value)
                                        else "пусто"
                                    ),
                                    "new_value": str(value),
                                    "base_name": match.get("base_name", "N/A"),
                                    "supplier_name": match.get("supplier_name", "N/A"),
                                }
                                self.changes_log.append(change_info)

                            else:
                                # Подробная информация о причине отказа
                                existing_value = str(current_article).strip()
                                if existing_value == str(supplier_article).strip():
                                    reason = f"артикул уже содержит то же значение ({existing_value})"
                                else:
                                    reason = f"ячейка уже заполнена значением ({existing_value}), попытка перезаписать на ({supplier_article})"

                                self.log_info(f"⏭️ Код {code}: ПРОПУЩЕН - {reason}")

                                # Сохраняем информацию о пропуске для отчета
                                if not hasattr(self, "changes_log"):
                                    self.changes_log = []

                                skip_info = {
                                    "type": "article_skipped",
                                    "base_index": base_idx,
                                    "code": code,
                                    "match_type": match_type,
                                    "column": excel_column_name,
                                    "existing_value": existing_value,
                                    "attempted_value": str(supplier_article),
                                    "reason": reason,
                                    "base_name": match.get("base_name", "N/A"),
                                    "supplier_name": match.get("supplier_name", "N/A"),
                                }
                                self.changes_log.append(skip_info)

                        except ValueError as e:
                            self.log_error(
                                f"❌ Ошибка преобразования артикула {supplier_article} в тип {data_type}: {e}"
                            )
                            continue
                        except Exception as e:
                            self.log_error(
                                f"❌ Ошибка обработки артикула {supplier_article}: {e}"
                            )
                            continue
                    else:
                        self.log_error(
                            f"❌ Недостаточно данных: base_idx={base_idx}, supplier_article={supplier_article}"
                        )
                        self.log_error(f"   match keys: {list(match.keys())}")
                        self.log_error(
                            f"   supplier_article from match: {match.get('supplier_article')}"
                        )
                        self.log_error(f"   article from match: {match.get('article')}")

                except Exception as e:
                    self.log_error(
                        f"❌ Ошибка добавления артикула по коду {match.get('code', 'N/A')}: {e}"
                    )

                # Обновляем прогресс каждые 5 артикулов для более частого обновления
                if i % 5 == 0:
                    progress_percent = int((i / len(selected_matches)) * 100)
                    self.update_progress(
                        4,
                        f"Обработано {i}/{len(selected_matches)} артикулов ({progress_percent}%)",
                    )
                    self.root.update()

            # Показываем результаты
            self.log_info("✅ Добавление артикулов завершено")
            self.log_info(f"   🔗 Артикулов добавлено: {articles_added}")

            # Обрабатываем новые товары - вставляем пустые строки в Excel
            rows_inserted = 0
            if selected_new_items:
                self.log_info(
                    f"📝 Начинаем обработку {len(selected_new_items)} новых товаров..."
                )
                self.set_status(
                    f"📝 Подготовка к вставке {len(selected_new_items)} новых товаров...",
                    "loading",
                )
                self.update_progress(
                    5, f"Подготовка к вставке {len(selected_new_items)} новых товаров"
                )

                # Собираем номера строк для вставки
                row_numbers_to_insert = []
                processed_items = set()  # Для отслеживания уже обработанных товаров

                self.set_status(
                    f"🔍 Анализ {len(selected_new_items)} новых товаров...", "loading"
                )
                self.update_progress(
                    6, f"Анализ {len(selected_new_items)} новых товаров"
                )

                for i, new_item in enumerate(selected_new_items, 1):
                    base_row_number = new_item.get("base_row_number")
                    item_name = new_item["match_data"].get("name", "N/A")

                    # Обновляем статус для каждого товара
                    self.set_status(
                        f"🔍 Анализ {i}/{len(selected_new_items)}: {item_name[:40]}...",
                        "loading",
                    )

                    if base_row_number is not None:
                        # Преобразуем в число для проверки
                        try:
                            row_num = int(base_row_number)
                            if row_num > 0:
                                # Проверяем, не обрабатывали ли мы уже этот товар
                                item_key = f"{item_name}_{row_num}"
                                if item_key not in processed_items:
                                    row_numbers_to_insert.append(row_num)
                                    processed_items.add(item_key)
                                    self.log_info(
                                        f"📝 Новый товар '{item_name}' - вставка строки после {row_num}"
                                    )
                                else:
                                    self.log_info(
                                        f"⚠️ Товар '{item_name}' уже обработан для строки {row_num}, пропускаем"
                                    )
                            else:
                                self.log_error(
                                    f"❌ Некорректный номер строки {row_num} для товара '{item_name}'"
                                )
                        except (ValueError, TypeError) as e:
                            self.log_error(
                                f"❌ Ошибка преобразования номера строки '{base_row_number}' для товара '{item_name}': {e}"
                            )
                    else:
                        self.log_info(
                            f"⚠️ Новый товар '{item_name}' - не найден в базе, пропускаем"
                        )

                    # Обновляем прогресс каждые 3 товара
                    if i % 3 == 0 or i == len(selected_new_items):
                        progress_percent = int((i / len(selected_new_items)) * 100)
                        self.update_progress(
                            6,
                            f"Проанализировано {i}/{len(selected_new_items)} новых товаров ({progress_percent}%)",
                        )
                        self.root.update()

                self.log_info(
                    f"📊 Итого уникальных строк для вставки: {len(row_numbers_to_insert)}"
                )

                # Вставляем пустые строки в Excel
                if row_numbers_to_insert:
                    try:
                        # Определяем путь к файлу базы
                        base_file_path = "data/input"
                        original_path = None

                        self.log_info(
                            f"🔍 Ищем файл базы в директории: {base_file_path}"
                        )

                        if os.path.exists(base_file_path):
                            base_files = []
                            for file in os.listdir(base_file_path):
                                if file.endswith(
                                    (".xlsx", ".xls")
                                ) and not file.startswith("~"):
                                    full_path = os.path.join(base_file_path, file)
                                    file_size = os.path.getsize(full_path)
                                    base_files.append((full_path, file_size, file))
                                    self.log_info(
                                        f"📁 Найден файл: {file} ({file_size} байт)"
                                    )

                            if base_files:
                                base_files.sort(key=lambda x: x[1], reverse=True)
                                original_path = base_files[0][0]
                                self.log_info(
                                    f"🎯 Выбран файл базы: {os.path.basename(original_path)}"
                                )

                                self.set_status(
                                    f"💾 Сохранение в файл: {os.path.basename(original_path)}...",
                                    "loading",
                                )
                                self.update_progress(
                                    9,
                                    f"Сохранение в файл: {os.path.basename(original_path)}",
                                )
                            else:
                                self.log_error("❌ В директории нет Excel файлов")

                        if original_path:
                            # Проверяем доступность файла
                            self.set_status(
                                "🔒 Проверка прав доступа к файлу...", "loading"
                            )
                            self.update_progress(9, "Проверка прав доступа к файлу")

                            if os.access(original_path, os.R_OK | os.W_OK):
                                self.log_info(
                                    f"✅ Файл доступен для чтения и записи: {os.path.basename(original_path)}"
                                )

                                # Вставляем пустые строки
                                self.set_status(
                                    f"📝 Вставка {len(row_numbers_to_insert)} пустых строк в Excel...",
                                    "loading",
                                )
                                self.update_progress(
                                    7,
                                    f"Вставка {len(row_numbers_to_insert)} пустых строк в Excel",
                                )

                                self.insert_empty_rows_in_excel(
                                    original_path, row_numbers_to_insert
                                )
                                rows_inserted = len(row_numbers_to_insert)
                                self.log_info(
                                    f"✅ Вставлено {rows_inserted} пустых строк в Excel файл"
                                )
                            else:
                                self.log_error(
                                    f"❌ Нет прав доступа к файлу: {os.path.basename(original_path)}"
                                )
                        else:
                            self.log_error("❌ Не найден файл базы для вставки строк")

                    except Exception as e:
                        self.log_error(f"❌ Ошибка вставки пустых строк: {e}")
                        messagebox.showerror(
                            "Ошибка", f"Не удалось вставить пустые строки: {e}"
                        )

            # Обновляем список кандидатов для ИИ (убираем обработанные)
            processed_articles = []
            for selected in selected_matches:
                match = selected["match_data"]
                supplier_article = match.get("supplier_article") or match.get("article")
                if supplier_article and str(supplier_article).strip() not in [
                    "",
                    "nan",
                    "None",
                ]:
                    processed_articles.append(
                        str(supplier_article)
                    )  # Преобразуем в строку!

            # Обновляем comparison_result - убираем обработанные товары из new_items
            if self.comparison_result and "new_items" in self.comparison_result:
                original_count = len(self.comparison_result["new_items"])
                self.comparison_result["new_items"] = [
                    item
                    for item in self.comparison_result["new_items"]
                    if item.get("article") not in processed_articles
                ]
                new_count = len(self.comparison_result["new_items"])
                self.log_info(
                    f"📉 Кандидатов для ИИ: было {original_count}, стало {new_count}"
                )

            # Обновляем отображение в основном окне
            self.update_main_window_info(
                articles_added + rows_inserted,
                len(selected_matches) + len(selected_new_items),
                processed_articles,
            )

            # Завершаем прогресс
            self.set_status("✅ Обработка завершена!", "success")
            self.update_progress(10, "Обработка завершена")

            result_message = f"Обработка завершена!\n\n"
            result_message += f"🔗 Артикулов добавлено: {articles_added}\n"
            result_message += f"📝 Пустых строк вставлено: {rows_inserted}\n"
            result_message += f"📋 Выбрано для обработки: {len(selected_matches) + len(selected_new_items)}\n"
            if processed_articles:
                result_message += (
                    f"📝 Обработанные артикулы: {', '.join(processed_articles[:5])}"
                )
                if len(processed_articles) > 5:
                    result_message += f" и еще {len(processed_articles) - 5}...\n"
                else:
                    result_message += "\n"

            # Сохраняем изменения в Excel файл если что-то добавлено или вставлено
            if articles_added > 0 or rows_inserted > 0:
                self.set_status("💾 Сохранение изменений в Excel файл...", "loading")
                self.update_progress(8, "Сохранение изменений в Excel файл")
                self.log_info("💾 Сохранение изменений в Excel файл...")

                # Определяем путь к оригинальному файлу базы
                base_file_path = "data/input"
                original_path = None

                self.set_status("🔍 Поиск файла базы для сохранения...", "loading")
                self.update_progress(8, "Поиск файла базы для сохранения")

                if os.path.exists(base_file_path):
                    base_files = []
                    for file in os.listdir(base_file_path):
                        if file.endswith((".xlsx", ".xls")) and not file.startswith(
                            "~"
                        ):
                            full_path = os.path.join(base_file_path, file)
                            file_size = os.path.getsize(full_path)
                            base_files.append((full_path, file_size, file))

                    if base_files:
                        base_files.sort(key=lambda x: x[1], reverse=True)
                        original_path = base_files[0][0]

                if original_path:
                    try:
                        # Точечное обновление с сохранением форматирования
                        if hasattr(self, "changes_log") and self.changes_log:
                            self.update_excel_articles_preserve_formatting(
                                original_path, self.changes_log
                            )
                        else:
                            # Fallback - простое сохранение
                            self.base_df.to_excel(
                                original_path, index=False, engine="openpyxl"
                            )

                        self.log_info(
                            f"💾 База данных обновлена: {os.path.basename(original_path)}"
                        )
                        if articles_added > 0:
                            result_message += f"\n💾 Артикулы добавлены в базу"
                        if rows_inserted > 0:
                            result_message += f"\n💾 Пустые строки вставлены в Excel"
                    except Exception as save_error:
                        self.log_error(f"❌ Ошибка сохранения: {save_error}")
                        messagebox.showerror(
                            "Ошибка сохранения",
                            f"Не удалось сохранить изменения: {save_error}",
                        )
                        return
                else:
                    self.log_error("❌ Не найден файл базы для сохранения")

            messagebox.showinfo("Добавление завершено", result_message)

            # Устанавливаем флаг добавления товаров
            if articles_added > 0 or rows_inserted > 0:
                self.articles_added = True
                # Сбрасываем флаг обновления цен, так как теперь есть новые товары для обновления
                self.price_updated = False
                self.log_info(
                    "🔒 Кнопка 'Добавить в базу' деактивирована после добавления товаров"
                )
                self.log_info(
                    "🔄 Кнопка 'Обновить цены' активирована после добавления новых товаров"
                )

            # Обновляем состояние кнопок
            self.set_status("🔧 Обновление состояния интерфейса...", "loading")
            self.update_buttons_state(articles_added_this_run=(articles_added > 0))
            self.update_progress(4, "Завершение добавления артикулов")
            self.finish_progress(
                f"Артикулов добавлено: {articles_added}", auto_reset=True
            )
            self.set_status(
                f"✅ Успешно добавлено {articles_added} артикулов в базу", "success"
            )

        except Exception as e:
            self.log_error(f"❌ Ошибка добавления артикулов: {e}")
            self.finish_progress("Ошибка добавления артикулов", auto_reset=True)

    def update_main_window_info(
        self, articles_added, selected_count, processed_articles
    ):
        """Обновляет информацию в основном окне о добавленных артикулах"""

        # Добавляем информацию в основное текстовое поле
        from datetime import datetime

        timestamp = datetime.now().strftime("%H:%M:%S")

        info_text = f"\n{'='*60}\n"
        info_text += f"[{timestamp}] 🔗 ДОБАВЛЕНИЕ АРТИКУЛОВ ЗАВЕРШЕНО\n"
        info_text += f"{'='*60}\n"
        info_text += f"✅ Артикулов добавлено в базу: {articles_added}\n"
        info_text += f"📋 Выбрано пользователем: {selected_count}\n"

        if articles_added > 0:
            info_text += f"📝 Добавленные артикулы:\n"
            for i, article in enumerate(processed_articles[:10], 1):
                info_text += f"   {i}. {article}\n"
            if len(processed_articles) > 10:
                info_text += f"   ... и еще {len(processed_articles) - 10} артикулов\n"
        else:
            info_text += f"⚠️ Возможные причины:\n"
            info_text += f"   • Артикулы уже существуют в базе\n"
            info_text += f"   • Ошибка в данных или индексах\n"
            info_text += f"   • Проблема со столбцами базы данных\n"

        # Информация об обновленном списке кандидатов для ИИ
        if (
            hasattr(self, "comparison_result")
            and self.comparison_result
            and "new_items" in self.comparison_result
        ):
            remaining_count = len(self.comparison_result["new_items"])
            info_text += f"\n🤖 Кандидатов для ИИ обработки: {remaining_count}\n"

        info_text += f"{'='*60}\n"

        # Добавляем в текстовое поле основного окна, если оно существует
        if hasattr(self, "info_text"):
            self.info_text.insert(tk.END, info_text)
            self.info_text.see(tk.END)  # Прокручиваем к концу

        # Обновляем состояние кнопок (количество новых товаров могло измениться)
        # Передаем информацию о том, что артикулы были добавлены в этом сеансе
        self.update_buttons_state(articles_added_this_run=(articles_added > 0))

    def show_settings(self):
        """Показать окно настроек с автозагрузкой базы"""
        self.log_info("⚙️ Открытие окна настроек...")

        # Создаем окно настроек
        settings_window = tk.Toplevel(self.root)
        settings_window.title("Настройки MiStockSync")
        settings_window.resizable(False, False)

        # Устанавливаем иконку для окна
        self.set_window_icon(settings_window)

        # Центрируем окно относительно главного окна
        window_width = 450
        window_height = 480  # Увеличено из-за добавления настроек подтверждения выхода
        self.center_window(settings_window, window_width, window_height)

        # Делаем окно модальным
        settings_window.transient(self.root)
        settings_window.grab_set()

        # Заголовок
        ttk.Label(
            settings_window, text="⚙️ Настройки приложения", font=("Arial", 14, "bold")
        ).pack(pady=10)

        # Рамка с настройками
        settings_frame = ttk.LabelFrame(
            settings_window, text="Основные настройки", padding="10"
        )
        settings_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Настройка автозагрузки базы
        auto_load_frame = ttk.Frame(settings_frame)
        auto_load_frame.pack(fill="x", pady=10)

        ttk.Label(
            auto_load_frame,
            text="📊 Автозагрузка базы данных:",
            font=("Arial", 10, "bold"),
        ).pack(anchor="w")

        auto_load_var = tk.BooleanVar(value=self.auto_load_base_enabled)
        auto_load_check = ttk.Checkbutton(
            auto_load_frame,
            text="Автоматически загружать базу при сравнении",
            variable=auto_load_var,
        )
        auto_load_check.pack(anchor="w", padx=20, pady=5)

        ttk.Label(
            auto_load_frame,
            text="При включении база будет загружаться автоматически\nпри первом сравнении с поставщиком.",
            font=("Arial", 8),
            foreground="gray",
        ).pack(anchor="w", padx=20)

        # Разделитель
        ttk.Separator(settings_frame, orient="horizontal").pack(fill="x", pady=15)

        # Настройка размера шрифта
        font_frame = ttk.Frame(settings_frame)
        font_frame.pack(fill="x", pady=10)

        ttk.Label(
            font_frame,
            text="🔤 Размер шрифта по умолчанию:",
            font=("Arial", 10, "bold"),
        ).pack(anchor="w")

        font_size_var = tk.StringVar(value=self.current_font_size)

        font_options = [
            ("📝 Обычный", "normal"),
            ("📄 Средний", "medium"),
            ("📊 Крупный", "large"),
        ]

        for text, value in font_options:
            ttk.Radiobutton(
                font_frame, text=text, variable=font_size_var, value=value
            ).pack(anchor="w", padx=20, pady=2)

        ttk.Label(
            font_frame,
            text="Изменения применятся к главному окну информации.",
            font=("Arial", 8),
            foreground="gray",
        ).pack(anchor="w", padx=20, pady=(5, 0))

        # Разделитель
        ttk.Separator(settings_frame, orient="horizontal").pack(fill="x", pady=15)

        # Настройка подтверждения выхода
        exit_frame = ttk.Frame(settings_frame)
        exit_frame.pack(fill="x", pady=10)

        ttk.Label(
            exit_frame,
            text="🚪 Подтверждение выхода:",
            font=("Arial", 10, "bold"),
        ).pack(anchor="w")

        confirm_exit_var = tk.BooleanVar(value=self.settings.get("confirm_exit", True))
        confirm_exit_check = ttk.Checkbutton(
            exit_frame,
            text="Показывать окно подтверждения при закрытии приложения",
            variable=confirm_exit_var,
        )
        confirm_exit_check.pack(anchor="w", padx=20, pady=5)

        ttk.Label(
            exit_frame,
            text="При отключении приложение будет закрываться сразу,\nесли не было изменений в данных.",
            font=("Arial", 8),
            foreground="gray",
        ).pack(anchor="w", padx=20)

        # Разделитель
        ttk.Separator(settings_frame, orient="horizontal").pack(fill="x", pady=15)

        # Функции в разработке
        ttk.Label(
            settings_frame, text="🚧 В разработке:", font=("Arial", 10, "bold")
        ).pack(anchor="w")

        planned_features = [
            "• Настройка путей к файлам по умолчанию",
            "• Пороги для фильтрации цен",
            "• Параметры автосохранения",
            "• Настройки логирования",
            "• Цветовые схемы интерфейса",
        ]

        for feature in planned_features:
            ttk.Label(settings_frame, text=feature, font=("Arial", 9)).pack(
                anchor="w", padx=10
            )

        # Кнопки
        button_frame = ttk.Frame(settings_window)
        button_frame.pack(pady=10)

        def save_settings():
            """Сохранить настройки"""
            # Сохраняем автозагрузку базы
            self.auto_load_base_enabled = auto_load_var.get()
            self.settings["auto_load_base"] = auto_load_var.get()

            # Сохраняем размер шрифта
            new_font_size = font_size_var.get()
            if new_font_size != self.current_font_size:
                self.current_font_size = new_font_size
                self.settings["font_size"] = new_font_size
                # Применяем новый размер шрифта сразу
                self.apply_font_size(new_font_size)

            # Сохраняем настройку подтверждения выхода
            self.settings["confirm_exit"] = confirm_exit_var.get()

            # Сохраняем настройки в файл
            if self.save_settings(self.settings):
                self.log_info(
                    f"💾 Настройки сохранены: автозагрузка={auto_load_var.get()}, шрифт={new_font_size}, подтверждение выхода={confirm_exit_var.get()}"
                )
                messagebox.showinfo("Настройки", "Настройки успешно сохранены!")
            else:
                messagebox.showerror("Ошибка", "Не удалось сохранить настройки")

            settings_window.destroy()

        def cancel_settings():
            """Отменить изменения"""
            self.log_info("↩️ Изменения настроек отменены")
            settings_window.destroy()

        ttk.Button(button_frame, text="💾 Сохранить", command=save_settings).pack(
            side="left", padx=5
        )
        ttk.Button(button_frame, text="❌ Отмена", command=cancel_settings).pack(
            side="left", padx=5
        )

    def quit_application(self):
        """Выход из приложения с подтверждением (только при изменениях)

        Логика работы:
        1. Если были изменения в данных И включено подтверждение - показываем окно подтверждения
        2. Если были изменения, но подтверждение отключено - выходим с предупреждением в логах
        3. Если изменений не было - выходим без подтверждения
        """
        self.log_info("🚪 Запрос на выход из приложения...")

        # Проверяем настройку подтверждения выхода
        confirm_exit = self.settings.get("confirm_exit", True)

        # Проверяем, были ли изменения в данных
        has_changes = (hasattr(self, "price_updated") and self.price_updated) or (
            hasattr(self, "articles_added") and self.articles_added
        )

        if has_changes and confirm_exit:
            # Если были изменения и включено подтверждение - показываем окно
            result = messagebox.askyesno(
                "Подтверждение выхода",
                "Вы действительно хотите выйти из MiStockSync?\n\n"
                "⚠️ Обнаружены несохраненные изменения в данных!\n"
                "Все несохраненные данные будут потеряны.",
                icon="warning",
            )
        elif has_changes and not confirm_exit:
            # Если были изменения, но подтверждение отключено - выходим с предупреждением в логах
            result = True
            self.log_info(
                "⚠️ Обнаружены изменения, но подтверждение выхода отключено - выход без подтверждения"
            )
        else:
            # Если изменений не было - выходим без подтверждения
            result = True
            self.log_info("ℹ️ Изменений не обнаружено - выход без подтверждения")

        if result:
            self.log_info("👋 Завершение работы приложения...")

            # Сохраняем текущие размеры окна перед выходом
            try:
                current_width = self.root.winfo_width()
                current_height = self.root.winfo_height()

                # Обновляем настройки
                self.settings["main_window_width"] = current_width
                self.settings["main_window_height"] = current_height

                # Сохраняем настройки
                self.save_settings(self.settings)
                self.log_info(
                    f"💾 Размеры окна сохранены: {current_width}x{current_height}"
                )
            except Exception as e:
                self.log_error(f"❌ Ошибка сохранения размеров окна: {e}")

            self.logger.info("📋 Приложение закрыто пользователем")
            self.root.quit()
        else:
            self.log_info("↩️ Выход отменен пользователем")

    def show_about(self):
        """Показать информацию о программе"""
        self.log_info("ℹ️ Показ информации о программе...")

        # Создаем отдельное окно вместо простого messagebox
        about_window = tk.Toplevel(self.root)
        about_window.title("О программе")
        about_window.resizable(False, False)

        # Устанавливаем иконку для окна
        self.set_window_icon(about_window)

        # Центрируем окно относительно главного окна
        window_width = 320  # Увеличено с 300 из-за длинного текста
        window_height = 350  # Увеличено с 240 из-за дополнительного текста
        self.center_window(about_window, window_width, window_height)

        # Делаем окно модальным
        about_window.transient(self.root)
        about_window.grab_set()

        # Главный фрейм
        main_frame = ttk.Frame(about_window, padding="20")
        main_frame.pack(fill="both", expand=True)

        # Большая иконка приложения (эмодзи)
        ttk.Label(main_frame, text="🚀", font=("Arial", 48)).pack()

        # Название и версия
        ttk.Label(
            main_frame, text="MiStockSync v0.0.9", font=("Arial", 14, "bold")
        ).pack(pady=5)

        # Дата
        ttk.Label(
            main_frame,
            text=f"📅 {datetime.now().strftime('%Y-%m-%d')}",
            font=("Arial", 9),
        ).pack()

        # Краткое описание
        ttk.Label(
            main_frame,
            text="Синхронизация прайс-листов\nс базой данных товаров\n\n• Автозагрузка базы данных\n• Настройка размера шрифта\n• Сохранение пользовательских настроек",
            font=("Arial", 9),
            justify="center",
        ).pack(pady=10)

        # Кнопка закрытия
        ttk.Button(
            main_frame, text="✅ ОК", command=about_window.destroy, width=10
        ).pack(pady=10)

        self.log_info("ℹ️ Информация о программе показана")

    # === ФУНКЦИИ МЕНЮ "ПРАВКА" ===
    def cut_text(self):
        """Вырезать выделенный текст"""
        try:
            focused_widget = self.root.focus_get()
            if hasattr(focused_widget, "selection_get"):
                text = focused_widget.selection_get()
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                focused_widget.delete("sel.first", "sel.last")
                self.log_info("✂️ Текст вырезан в буфер обмена")
        except tk.TclError:
            self.log_info("⚠️ Нет выделенного текста для вырезания")

    def copy_text(self):
        """Копировать выделенный текст"""
        try:
            focused_widget = self.root.focus_get()
            if hasattr(focused_widget, "selection_get"):
                text = focused_widget.selection_get()
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
                self.log_info("📋 Текст скопирован в буфер обмена")
        except tk.TclError:
            self.log_info("⚠️ Нет выделенного текста для копирования")

    def copy_selected_text(self, text_widget, window):
        """Универсальная функция копирования выделенного текста"""
        try:
            selected_text = text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
            window.clipboard_clear()
            window.clipboard_append(selected_text)
        except tk.TclError:
            pass  # Нет выделения

    def copy_all_text(self, text_widget, window):
        """Универсальная функция копирования всего текста"""
        try:
            text_widget.configure(state="normal")
            all_text = text_widget.get(1.0, tk.END)
            text_widget.configure(state="disabled")
            window.clipboard_clear()
            window.clipboard_append(all_text)
        except Exception as e:
            pass

    def copy_selected_text_with_notification(self, text_widget, window):
        """Копирование выделенного текста с уведомлением"""
        try:
            selected_text = text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
            window.clipboard_clear()
            window.clipboard_append(selected_text)
            messagebox.showinfo(
                "Копирование", "Выделенный текст скопирован в буфер обмена!"
            )
        except tk.TclError:  # Нет выделения
            messagebox.showwarning(
                "Предупреждение", "Сначала выделите текст для копирования!"
            )
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка копирования: {e}")

    def copy_all_text_with_notification(self, text_widget, window):
        """Копирование всего текста с уведомлением"""
        try:
            text_widget.configure(state="normal")
            all_text = text_widget.get(1.0, tk.END)
            text_widget.configure(state="disabled")
            window.clipboard_clear()
            window.clipboard_append(all_text)
            messagebox.showinfo("Копирование", "Все логи скопированы в буфер обмена!")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка копирования: {e}")

    def select_all_text(self):
        """Выделить весь текст в активном поле"""
        try:
            focused_widget = self.root.focus_get()
            if focused_widget == self.info_text:
                # Для ScrolledText
                focused_widget.tag_add("sel", "1.0", "end")
                self.log_info("🔘 Весь текст выделен")
            elif hasattr(focused_widget, "select_range"):
                # Для Entry
                focused_widget.select_range(0, tk.END)
                self.log_info("🔘 Весь текст выделен")
        except:
            self.log_info("⚠️ Нет активного текстового поля")

    def invert_selection(self):
        """Инвертировать выделение в активном текстовом поле"""
        try:
            focused_widget = self.root.focus_get()
            if focused_widget == self.info_text:
                # Для ScrolledText - инвертируем выделение
                try:
                    current_selection = focused_widget.tag_ranges("sel")
                    if current_selection:
                        # Если есть выделение, снимаем его
                        focused_widget.tag_remove("sel", "1.0", "end")
                        self.log_info("🔘 Выделение снято")
                    else:
                        # Если нет выделения, выделяем весь текст
                        focused_widget.tag_add("sel", "1.0", "end")
                        self.log_info("🔘 Весь текст выделен")
                except:
                    # Если что-то пошло не так, выделяем весь текст
                    focused_widget.tag_add("sel", "1.0", "end")
                    self.log_info("🔘 Весь текст выделен")
            elif hasattr(focused_widget, "select_range"):
                # Для Entry - выделяем весь текст
                focused_widget.select_range(0, tk.END)
                self.log_info("🔘 Весь текст выделен")
            else:
                self.log_info("⚠️ Нет активного текстового поля")
        except Exception as e:
            self.log_info(f"⚠️ Ошибка инвертирования выделения: {e}")

    # === ФУНКЦИИ РАЗМЕРА ШРИФТА ===
    def change_font_size(self, size_type):
        """Изменить размер шрифта в интерфейсе"""
        if size_type in ["normal", "medium", "large"]:
            # Применяем новый размер шрифта
            self.apply_font_size(size_type)

            # Сохраняем в настройки
            self.current_font_size = size_type
            self.settings["font_size"] = size_type
            self.save_settings(self.settings)

            size_names = {"normal": "обычный", "medium": "средний", "large": "крупный"}

            self.log_info(f"🔤 Установлен {size_names[size_type]} размер шрифта")
        else:
            self.log_info("⚠️ Неизвестный размер шрифта")

    def apply_font_size(self, size_type):
        """Применение размера шрифта к текстовому полю"""
        sizes = {
            "normal": ("Arial", 9),
            "medium": ("Arial", 11),
            "large": ("Arial", 13),
        }

        if size_type in sizes and hasattr(self, "info_text"):
            font_family, font_size = sizes[size_type]
            self.info_text.configure(font=(font_family, font_size))

    def center_window(self, window, width, height, parent=None):
        """Центрирование окна относительно родительского окна или экрана"""
        if parent is None:
            parent = self.root

        # Получаем размеры и позицию родительского окна
        parent.update_idletasks()
        parent_x = parent.winfo_x()
        parent_y = parent.winfo_y()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()

        # Вычисляем позицию для центрирования относительно родительского окна
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2

        # Устанавливаем размер и позицию
        window.geometry(f"{width}x{height}+{x}+{y}")

    def set_window_icon(self, window):
        """Установка иконки для дочернего окна"""
        try:
            from PIL import Image, ImageTk

            icon = ImageTk.PhotoImage(Image.open("assets/icon.png"))
            window.iconphoto(False, icon)
        except Exception:
            # Если не удалось загрузить иконку, пропускаем
            pass

    def create_backup_base(self):
        """Создание резервной копии оригинального Excel файла базы с форматированием"""

        if self.base_df is None:
            self.log_error("❌ База данных не загружена для создания backup")
            return False

        try:
            # Создаем папку для backup если не существует
            backup_dir = "data/output"
            os.makedirs(backup_dir, exist_ok=True)

            # Создаем имя файла backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"BACKUP_base_{self.current_config}_{timestamp}.xlsx"
            backup_path = os.path.join(backup_dir, backup_filename)

            # Находим оригинальный файл базы в data/input
            data_dir = "data/input"
            base_files = []

            if os.path.exists(data_dir):
                for file in os.listdir(data_dir):
                    if file.endswith((".xlsx", ".xls")) and "base" in file.lower():
                        file_path = os.path.join(data_dir, file)
                        file_size = os.path.getsize(file_path)
                        base_files.append((file_path, file_size, file))

            if base_files:
                # Берем самый большой файл (это должна быть база)
                base_files.sort(key=lambda x: x[1], reverse=True)
                original_path = base_files[0][0]

                # Копируем оригинальный файл с форматированием
                import shutil

                shutil.copy(original_path, backup_path)

                self.log_info(f"💾 Резервная копия создана: {backup_filename}")
                self.log_info(f"📁 Путь: {backup_path}")
                self.log_info(f"📄 Оригинал: {os.path.basename(original_path)}")

                return True
            else:
                self.log_error("❌ Не найден оригинальный файл базы в data/input")
                return False

        except Exception as e:
            self.log_error(f"❌ Ошибка создания резервной копии: {e}")
            return False

    def update_excel_prices_preserve_formatting(
        self, original_path, backup_path, price_updates, supplier_config
    ):
        """
        Точечное обновление цен в Excel файле с сохранением всего форматирования
        Изменяются ТОЛЬКО значения ценовых ячеек, всё остальное остается как было
        """

        self.log_info("🔧 Точечное обновление цен с сохранением форматирования...")

        try:
            # Проверяем наличие openpyxl
            if not OPENPYXL_AVAILABLE:
                self.log_error(
                    "❌ Библиотека openpyxl не установлена. Используйте: pip install openpyxl"
                )
                return False

            # 1. Создаем backup только если указан путь
            if backup_path:
                os.makedirs("data/output", exist_ok=True)
                shutil.copy(original_path, backup_path)
                self.log_info(f"💾 Backup создан: {os.path.basename(backup_path)}")
            else:
                self.log_info("🔧 Обновление без создания backup")

            # 2. Открываем Excel файл через openpyxl (сохраняет форматирование)
            workbook = load_workbook(original_path)
            worksheet = workbook.active  # Берем первый лист

            # 3. Определяем столбец для обновления цен (реальные названия в базе)
            if supplier_config == "vitya":
                price_column_name = self.get_excel_column_name_from_config(
                    "price_vitya_usd"
                )
                article_column_name = self.get_excel_column_name_from_config(
                    "article_vitya"
                )
            elif supplier_config == "dimi":
                price_column_name = self.get_excel_column_name_from_config(
                    "price_dimi_usd"
                )
                article_column_name = self.get_excel_column_name_from_config(
                    "article_dimi"
                )
            else:
                self.log_error(f"❌ Неподдерживаемая конфигурация: {supplier_config}")
                return False

            # 4. Находим индексы столбцов в Excel файле (регистронезависимый поиск)
            header_row = 1  # Предполагаем что заголовки в первой строке
            price_col_idx = None
            article_col_idx = None

            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_value_str = str(cell_value).strip()
                    # Регистронезависимый поиск
                    if cell_value_str.lower() == price_column_name.lower():
                        price_col_idx = col_idx
                    elif cell_value_str.lower() == article_column_name.lower():
                        article_col_idx = col_idx

            if not price_col_idx or not article_col_idx:
                self.log_error(
                    f"❌ Не найдены столбцы в Excel: {price_column_name}, {article_column_name}"
                )
                # Показываем доступные столбцы для отладки
                available_columns = []
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=header_row, column=col_idx).value
                    if cell_value:
                        available_columns.append(str(cell_value).strip())
                self.log_error(f"📋 Доступные столбцы: {available_columns[:10]}...")
                return False

            self.log_info(
                f"📍 Найдены столбцы: {article_column_name} (col {article_col_idx}), {price_column_name} (col {price_col_idx})"
            )

            # Дополнительная проверка - показываем первые несколько значений в найденных столбцах
            self.log_info(f"🔍 Проверка столбца артикулов ({article_column_name}):")
            for row_idx in range(2, min(7, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_idx, column=article_col_idx).value
                self.log_info(
                    f"   Строка {row_idx}: {cell_value} (тип: {type(cell_value)})"
                )

            self.log_info(f"🔍 Проверка столбца цен ({price_column_name}):")
            for row_idx in range(2, min(7, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_idx, column=price_col_idx).value
                self.log_info(
                    f"   Строка {row_idx}: {cell_value} (тип: {type(cell_value)})"
                )

            # 5. Применяем только изменения цен
            updates_applied = 0

            self.log_info(
                f"🔍 Начинаем обработку {len(price_updates)} обновлений цен..."
            )

            # Показываем первые несколько обновлений для отладки
            for i, update in enumerate(price_updates[:5]):
                self.log_info(f"   Обновление {i+1}: {update}")

            for update in price_updates:
                article_to_find = str(update.get("article", "")).strip()
                new_price_raw = update.get("new_price", 0)

                # Приводим цену к правильному типу данных
                try:
                    new_price = (
                        float(new_price_raw) if new_price_raw is not None else 0.0
                    )
                except (ValueError, TypeError):
                    new_price = 0.0

                self.log_info(
                    f"🔍 Excel обновление: {article_to_find} → {new_price} (raw: {new_price_raw})"
                )

                if not article_to_find or new_price <= 0:
                    self.log_info(
                        f"   ⏭️ Пропускаем {article_to_find}: артикул пустой или цена <= 0"
                    )
                    continue

                # Ищем строку с нужным артикулом
                found_match = False
                for row_idx in range(2, worksheet.max_row + 1):  # Начинаем с 2-й строки
                    cell_value = worksheet.cell(
                        row=row_idx, column=article_col_idx
                    ).value

                    if cell_value is not None:
                        if supplier_config == "vitya":
                            # Для Вити сравниваем как int
                            try:
                                cell_value_int = (
                                    int(float(cell_value))
                                    if isinstance(cell_value, (int, float))
                                    else None
                                )
                                article_to_find_int = (
                                    int(float(article_to_find))
                                    if article_to_find
                                    else None
                                )

                                if (
                                    cell_value_int is not None
                                    and article_to_find_int is not None
                                    and cell_value_int == article_to_find_int
                                ):
                                    found_match = True
                                    self.log_info(
                                        f"   🔍 Найдено совпадение для Вити: {cell_value} == {article_to_find}"
                                    )
                                else:
                                    found_match = False
                            except (ValueError, TypeError) as e:
                                found_match = False
                                self.log_info(
                                    f"   ⚠️ Ошибка сравнения для Вити: {cell_value} vs {article_to_find} - {e}"
                                )
                        else:
                            # Для Димы сравниваем как строки
                            cell_value_str = str(cell_value).strip()
                            article_to_find_str = article_to_find.strip()
                            found_match = cell_value_str == article_to_find_str
                            if found_match:
                                self.log_info(
                                    f"   🔍 Найдено совпадение для Димы: '{cell_value_str}' == '{article_to_find_str}'"
                                )

                        if found_match:
                            # ОБНОВЛЯЕМ ТОЛЬКО ЗНАЧЕНИЕ ЯЧЕЙКИ (форматирование сохраняется!)
                            old_value = worksheet.cell(
                                row=row_idx, column=price_col_idx
                            ).value

                            # Проверяем, нужно ли обновлять цену
                            try:
                                old_value_float = (
                                    float(old_value) if old_value is not None else 0.0
                                )
                            except (ValueError, TypeError):
                                old_value_float = 0.0

                            price_diff = abs(new_price - old_value_float)
                            prices_equal = price_diff < 0.001

                            self.log_info(
                                f"🔍 Excel: {article_to_find}: old_value={old_value} ({type(old_value)}), new_price={new_price} ({type(new_price)}), diff={price_diff:.6f}, equal={prices_equal}"
                            )

                            if not prices_equal:
                                worksheet.cell(
                                    row=row_idx, column=price_col_idx, value=new_price
                                )
                                updates_applied += 1

                                self.log_info(
                                    f"   ✅ {article_to_find}: {old_value} → {new_price}"
                                )
                            else:
                                self.log_info(
                                    f"   ⏭️ {article_to_find}: цены одинаковые, пропускаем"
                                )
                            break

                if not found_match:
                    self.log_info(
                        f"   ❌ Артикул {article_to_find} не найден в Excel файле"
                    )

            # 6. Сохраняем файл (форматирование полностью сохраняется)
            self.log_info(f"💾 Сохраняем файл: {original_path}")
            try:
                workbook.save(original_path)
                workbook.close()
                self.log_info(f"✅ Файл успешно сохранен: {original_path}")
            except Exception as e:
                self.log_error(f"❌ Ошибка сохранения файла: {e}")
                return False

            self.log_info(f"✅ Применено {updates_applied} обновлений цен")
            self.log_info(
                f"🎨 Сохранено ВСЁ форматирование: размеры ячеек, цвета, картинки и т.д."
            )

            return True

        except Exception as e:
            self.log_error(f"❌ Ошибка обновления Excel файла: {e}")
            return False

    def update_excel_articles_preserve_formatting(self, file_path, changes_log):
        """
        Точечное обновление артикулов в Excel файле с сохранением форматирования

        Args:
            file_path: Путь к Excel файлу
            changes_log: Список изменений с информацией о том, что нужно обновить
        """
        try:
            from openpyxl import load_workbook

            # Загружаем рабочую книгу
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # Получаем заголовки для определения номеров столбцов
            headers = {}
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers[str(cell_value).lower().strip()] = col

            changes_made = 0

            for change in changes_log:
                if change["type"] == "article_added":
                    try:
                        # Определяем номер строки в Excel (base_index + 2, т.к. DataFrame index начинается с 0, а Excel с 1, плюс заголовок)
                        excel_row = change["base_index"] + 2

                        # Определяем столбец для записи - используем оригинальное имя из конфигурации
                        original_column_name = change[
                            "column"
                        ]  # Это уже оригинальное имя из конфигурации

                        # Ищем точное совпадение среди заголовков (регистронезависимый поиск)
                        excel_col = None
                        original_column_name_lower = (
                            original_column_name.lower().strip()
                        )
                        for header_name, col_num in headers.items():
                            if header_name == original_column_name_lower:
                                excel_col = col_num
                                break

                        if excel_col is not None:

                            # Получаем pandas название столбца из оригинального названия Excel
                            pandas_column_name = (
                                self.get_pandas_column_name_from_excel_name(
                                    change["column"]
                                )
                            )

                            # Преобразуем значение к нужному типу
                            data_type = self.get_column_data_type(pandas_column_name)

                            if data_type == "int":
                                value = int(change["new_value"])
                            elif data_type == "float":
                                value = float(change["new_value"])
                            else:
                                value = change["new_value"]

                            # Записываем новое значение в ячейку
                            cell = worksheet.cell(row=excel_row, column=excel_col)
                            old_value = cell.value
                            cell.value = value

                            changes_made += 1
                            self.log_info(
                                f"📝 Excel: строка {excel_row}, столбец '{original_column_name}' (pandas: '{pandas_column_name}'): '{old_value}' → '{value}' (тип: {data_type})"
                            )
                        else:
                            available_columns = list(headers.keys())
                            self.log_error(
                                f"❌ Столбец '{original_column_name}' не найден в Excel файле"
                            )
                            self.log_error(
                                f"📋 Доступные столбцы: {available_columns[:10]}..."
                            )  # Показываем первые 10

                    except Exception as e:
                        self.log_error(
                            f"❌ Ошибка обновления строки {change.get('base_index', 'N/A')}: {e}"
                        )

            # Сохраняем изменения
            if changes_made > 0:
                workbook.save(file_path)
                self.log_info(
                    f"💾 Внесено {changes_made} изменений в Excel файл с сохранением форматирования"
                )
            else:
                self.log_info("ℹ️ Нет изменений для записи в Excel")

        except Exception as e:
            self.log_error(f"❌ Ошибка точечного обновления Excel: {e}")
            # Fallback на обычное сохранение
            self.base_df.to_excel(file_path, index=False, engine="openpyxl")
            self.log_info("💾 Использовано резервное сохранение через pandas")

    def insert_empty_rows_in_excel(self, file_path, row_numbers):
        """
        Вставка пустых строк в Excel файл под указанными номерами строк

        Args:
            file_path: Путь к Excel файлу
            row_numbers: Список номеров строк (начиная с 1), после которых нужно вставить пустые строки
        """
        try:
            from openpyxl import load_workbook
            import os

            # Проверяем, что файл существует и доступен
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Файл не найден: {file_path}")

            # Проверяем права доступа
            if not os.access(file_path, os.R_OK | os.W_OK):
                raise PermissionError(f"Нет прав доступа к файлу: {file_path}")

            self.log_info(f"🔍 Открываем Excel файл: {os.path.basename(file_path)}")
            self.log_info(f"📊 Номера строк для вставки: {row_numbers}")

            # Загружаем рабочую книгу
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # Получаем максимальное количество строк
            max_row = worksheet.max_row
            self.log_info(f"📏 Максимальная строка в файле: {max_row}")

            # Проверяем корректность номеров строк
            valid_row_numbers = []
            for row_num in row_numbers:
                if isinstance(row_num, str):
                    try:
                        row_num = int(row_num)
                    except ValueError:
                        self.log_error(f"❌ Некорректный номер строки: {row_num}")
                        continue

                if 1 <= row_num <= max_row:
                    valid_row_numbers.append(row_num)
                else:
                    self.log_error(
                        f"❌ Номер строки {row_num} вне диапазона [1, {max_row}]"
                    )

            if not valid_row_numbers:
                self.log_error("❌ Нет корректных номеров строк для вставки")
                workbook.close()
                return 0

            # Убираем дубликаты и сортируем по убыванию
            unique_rows = list(set(valid_row_numbers))
            sorted_rows = sorted(unique_rows, reverse=True)

            self.log_info(f"📝 Уникальные строки для вставки: {sorted_rows}")

            rows_inserted = 0

            for i, row_num in enumerate(sorted_rows, 1):
                try:
                    # Обновляем статус для каждой строки
                    self.set_status(
                        f"📝 Вставка строки {i}/{len(sorted_rows)}: после строки {row_num}...",
                        "loading",
                    )

                    # Вставляем пустую строку после указанной строки
                    # openpyxl использует 1-индексацию
                    worksheet.insert_rows(row_num + 1)

                    self.log_info(
                        f"📝 Excel: вставлена пустая строка после строки {row_num}"
                    )
                    rows_inserted += 1

                    # Обновляем прогресс каждые 5 строк
                    if i % 5 == 0 or i == len(sorted_rows):
                        progress_percent = int((i / len(sorted_rows)) * 100)
                        self.update_progress(
                            7,
                            f"Вставлено {i}/{len(sorted_rows)} пустых строк ({progress_percent}%)",
                        )
                        self.root.update()

                except Exception as e:
                    self.log_error(f"❌ Ошибка вставки строки после {row_num}: {e}")
                    continue

            # Сохраняем изменения
            if rows_inserted > 0:
                try:
                    workbook.save(file_path)
                    self.log_info(
                        f"💾 Вставлено {rows_inserted} пустых строк в Excel файл"
                    )
                except Exception as save_error:
                    self.log_error(f"❌ Ошибка сохранения файла: {save_error}")
                    raise
            else:
                self.log_info("ℹ️ Нет строк для вставки")

            # Закрываем workbook
            workbook.close()

        except Exception as e:
            self.log_error(f"❌ Ошибка вставки пустых строк в Excel: {e}")
            raise


def main():
    """Главная функция приложения"""
    # Базовое логирование для main функции
    print("🚀 Запуск MiStockSync GUI...")
    print("📋 Инициализация интерфейса...")

    root = tk.Tk()

    # Настройка иконки приложения
    try:
        from PIL import Image, ImageTk

        icon = ImageTk.PhotoImage(Image.open("assets/icon.png"))
        root.iconphoto(False, icon)
        print("✅ Иконка приложения загружена")
    except Exception as e:
        print(f"⚠️ Не удалось загрузить иконку: {e}")

    # Устанавливаем заголовок
    root.title("🚀 MiStockSync - Управление прайсами")

    app = MiStockSyncApp(root)

    # Центрируем окно только если размеры не были загружены из конфигурации
    root.update_idletasks()
    current_width = root.winfo_width()
    current_height = root.winfo_height()

    # Проверяем, были ли размеры загружены из конфигурации
    if (
        current_width == 1000
        and current_height == 800
        and app.settings.get("main_window_width", 1000) == 1000
        and app.settings.get("main_window_height", 800) == 800
    ):
        # Размеры по умолчанию - центрируем окно
        x = (root.winfo_screenwidth() // 2) - (current_width // 2)
        y = (root.winfo_screenheight() // 2) - (current_height // 2)
        root.geometry(f"{current_width}x{current_height}+{x}+{y}")
        app.log_info("🎯 Окно отцентрировано (использованы размеры по умолчанию)")
    else:
        # Размеры загружены из конфигурации - центрируем по текущим размерам
        x = (root.winfo_screenwidth() // 2) - (current_width // 2)
        y = (root.winfo_screenheight() // 2) - (current_height // 2)
        root.geometry(f"+{x}+{y}")  # Только позиция, размер уже установлен
        app.log_info(
            f"📐 Размеры окна загружены из конфигурации: {current_width}x{current_height}"
        )

    # Добавляем обработчик закрытия окна
    def on_closing():
        app.quit_application()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    app.logger.info("🖥️ GUI интерфейс готов к работе")
    print("✅ Приложение готово к работе!")

    root.mainloop()


if __name__ == "__main__":
    main()
